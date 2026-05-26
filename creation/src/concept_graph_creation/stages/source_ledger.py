from __future__ import annotations

import hashlib
import json
import re
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


WORKBOOK_COLUMNS = {
    "week": "Week",
    "sort": "Sort",
    "row_type": "Type",
    "title": "Title",
    "date": "Date",
    "date_source": "Date source",
    "parent_class": "Parent class",
    "class_date": "Class date",
    "professor": "Professor",
    "axis": "Axis",
    "related_subjects": "Related subjects",
    "description": "Description",
    "url": "URL",
    "resource_code": "Resource code",
    "required": "Required",
    "grade_weight": "Grade weight",
}


def build_source_ledger(
    *,
    cg_pipeline_root: Path,
    workbook_path: Path,
    index_path: Path,
    subject_sheet: str,
    course_id: str,
    module_id: str,
    subject_id: str,
) -> dict[str, Any]:
    workbook_rows = _read_workbook_rows(workbook_path, subject_sheet)
    extraction_index = json.loads(index_path.read_text(encoding="utf-8"))
    records_by_id = {str(record["id"]): record for record in extraction_index["records"]}

    lessons = _build_lessons(workbook_rows)
    lessons_by_key = {(lesson["title"], lesson["date"]): lesson for lesson in lessons}
    lessons_by_title = {lesson["title"]: lesson for lesson in lessons}

    self_studies: list[dict[str, Any]] = []
    for row in workbook_rows:
        if row["row_type"] != "Self-study":
            continue

        self_study_id = str(row["workbook_row_number"])
        record = records_by_id.get(self_study_id)
        lesson = lessons_by_key.get((row["parent_class"], row["class_date"])) or lessons_by_title.get(row["parent_class"])
        status = _source_body_status(row, record)
        metadata_only_candidate = status == "unavailable_source_body"

        self_studies.append(
            {
                "self_study_id": self_study_id,
                "lesson_id": lesson["lesson_id"] if lesson else None,
                "source_body_status": status,
                "metadata_only_candidate": metadata_only_candidate,
                "exclusion_reason": "activity_only_without_teaching_signal"
                if status == "excluded_activity_only_self_study"
                else None,
                "workbook_metadata": _workbook_metadata(row),
                "source_body": _source_body(record),
                "ledger_warnings": _ledger_warnings(row, lesson),
            }
        )

    return {
        "artifact_type": "source_ledger",
        "schema_version": "source_ledger.v0",
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "course_id": course_id,
        "module_id": module_id,
        "subject_id": subject_id,
        "inputs": {
            "workbook_path": _relative_to(workbook_path, cg_pipeline_root),
            "index_path": _relative_to(index_path, cg_pipeline_root),
            "extraction_dir": extraction_index.get("extraction_dir", "extraction"),
            "workbook_sha256": _sha256_file(workbook_path),
            "index_sha256": _sha256_file(index_path),
        },
        "summary": {
            "lesson_count": len(lessons),
            "self_study_count": len(self_studies),
            "available_count": sum(1 for item in self_studies if item["source_body_status"] == "usable_source_body"),
            "unavailable_count": sum(
                1 for item in self_studies if item["source_body_status"] == "unavailable_source_body"
            ),
            "excluded_count": sum(
                1 for item in self_studies if item["source_body_status"] == "excluded_activity_only_self_study"
            ),
        },
        "lessons": lessons,
        "self_studies": self_studies,
    }


def validate_source_ledger(ledger: dict[str, Any], *, cg_pipeline_root: Path) -> list[str]:
    errors: list[str] = []
    if ledger.get("artifact_type") != "source_ledger":
        errors.append("source_ledger.artifact_type must be 'source_ledger'")
    if not ledger.get("lessons"):
        errors.append("source_ledger.lessons must not be empty")
    if not ledger.get("self_studies"):
        errors.append("source_ledger.self_studies must not be empty")

    lesson_ids = {lesson.get("lesson_id") for lesson in ledger.get("lessons", [])}
    seen_self_studies: set[str] = set()
    for item in ledger.get("self_studies", []):
        location = f"self_studies[{item.get('self_study_id', '?')}]"
        self_study_id = item.get("self_study_id")
        if not self_study_id:
            errors.append(f"{location}.self_study_id is required")
        elif self_study_id in seen_self_studies:
            errors.append(f"{location}.self_study_id is duplicated")
        seen_self_studies.add(str(self_study_id))

        if item.get("lesson_id") not in lesson_ids:
            errors.append(f"{location}.lesson_id must reference a lesson")

        metadata = item.get("workbook_metadata") or {}
        if not metadata.get("title"):
            errors.append(f"{location}.workbook_metadata.title is required")
        if "required" not in metadata:
            errors.append(f"{location}.workbook_metadata.required is required")

        status = item.get("source_body_status")
        source_body = item.get("source_body") or {}
        if status == "usable_source_body":
            path = source_body.get("path")
            if not path:
                errors.append(f"{location}.source_body.path is required for usable source bodies")
            else:
                full_path = cg_pipeline_root / path
                if not full_path.is_file():
                    errors.append(f"{location}.source_body.path does not exist: {path}")
                elif source_body.get("sha256") != _sha256_file(full_path):
                    errors.append(f"{location}.source_body.sha256 does not match file: {path}")
            if not source_body.get("word_count"):
                errors.append(f"{location}.source_body.word_count is required for usable source bodies")
        elif status == "unavailable_source_body":
            if item.get("metadata_only_candidate") is not True:
                errors.append(f"{location}.metadata_only_candidate must be true for unavailable source bodies")
            if not source_body.get("availability_failures"):
                errors.append(f"{location}.source_body.availability_failures is required for unavailable source bodies")
        elif status == "excluded_activity_only_self_study":
            if not item.get("exclusion_reason"):
                errors.append(f"{location}.exclusion_reason is required for excluded activity-only self-studies")
        else:
            errors.append(f"{location}.source_body_status is invalid: {status}")

    summary = ledger.get("summary") or {}
    if summary.get("self_study_count") != len(ledger.get("self_studies", [])):
        errors.append("source_ledger.summary.self_study_count does not match self_studies length")
    if summary.get("lesson_count") != len(ledger.get("lessons", [])):
        errors.append("source_ledger.summary.lesson_count does not match lessons length")
    return errors


def write_source_ledger(ledger: dict[str, Any], output_path: Path) -> Path:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(json.dumps(ledger, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    return output_path


def _read_workbook_rows(workbook_path: Path, subject_sheet: str) -> list[dict[str, Any]]:
    workbook = load_workbook(workbook_path, read_only=False, data_only=True)
    worksheet = workbook[subject_sheet]
    rows = list(worksheet.iter_rows(values_only=True))
    header = [str(value).strip() if value is not None else "" for value in rows[0]]
    header_index = {name: index for index, name in enumerate(header)}

    parsed: list[dict[str, Any]] = []
    for row_number, values in enumerate(rows[1:], start=2):
        if not any(value is not None and str(value).strip() for value in values):
            continue
        parsed_row: dict[str, Any] = {"workbook_row_number": row_number}
        for key, column_name in WORKBOOK_COLUMNS.items():
            parsed_row[key] = _clean_cell(values[header_index[column_name]])
        parsed.append(parsed_row)
    return parsed


def _build_lessons(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    lessons: list[dict[str, Any]] = []
    for row in rows:
        if row["row_type"] != "Class":
            continue
        title = row["title"]
        date = row["date"]
        lessons.append(
            {
                "lesson_id": f"lesson-{_date_slug(date)}-{_slugify(title)}",
                "display_code": f"L{len(lessons) + 1:02d}",
                "workbook_row_number": row["workbook_row_number"],
                "title": title,
                "date": date,
                "professor": row["professor"],
                "axis": row["axis"],
                "related_labels": _split_labels(row["related_subjects"]),
                "description": row["description"],
            }
        )
    return lessons


def _workbook_metadata(row: dict[str, Any]) -> dict[str, Any]:
    return {
        "workbook_row_number": row["workbook_row_number"],
        "week": row["week"],
        "sort": row["sort"],
        "title": row["title"],
        "description": row["description"],
        "url": row["url"],
        "resource_code": row["resource_code"],
        "required": str(row["required"]).strip().lower() == "yes",
        "grade_weight": float(row["grade_weight"] or 0),
        "related_labels": _split_labels(row["related_subjects"]),
        "parent_class": row["parent_class"],
        "class_date": row["class_date"],
    }


def _source_body(record: dict[str, Any] | None) -> dict[str, Any]:
    if not record:
        return {
            "index_record_present": False,
            "path": None,
            "sha256": None,
            "word_count": None,
            "availability_warnings": [],
            "availability_failures": ["missing_extraction_index_record"],
        }

    file_info = record.get("file") or {}
    return {
        "index_record_present": True,
        "type": record.get("type"),
        "status": record.get("status"),
        "path": file_info.get("path"),
        "sha256": file_info.get("sha256"),
        "word_count": file_info.get("word_count"),
        "bytes": file_info.get("bytes"),
        "source_markdown": record.get("source_markdown"),
        "artifact_dir": record.get("artifact_dir"),
        "availability_warnings": record.get("warnings") or [],
        "availability_failures": record.get("failures") or [],
    }


def _ledger_warnings(row: dict[str, Any], lesson: dict[str, Any] | None) -> list[str]:
    warnings: list[str] = []
    if lesson and row.get("class_date") and row.get("class_date") != lesson.get("date"):
        warnings.append("workbook_parent_class_date_mismatch")
    return warnings


def _source_body_status(row: dict[str, Any], record: dict[str, Any] | None) -> str:
    if record and record.get("available") is True and (record.get("file") or {}).get("path"):
        return "usable_source_body"
    if _is_activity_only_without_teaching_signal(row):
        return "excluded_activity_only_self_study"
    return "unavailable_source_body"


def _is_activity_only_without_teaching_signal(row: dict[str, Any]) -> bool:
    title = str(row.get("title") or "").strip().lower()
    description = str(row.get("description") or "").strip()
    return title.startswith(("atividade:", "questão")) and not description and not row.get("url") and not row.get("resource_code")


def _split_labels(value: Any) -> list[str]:
    if value is None:
        return []
    return [part.strip() for part in str(value).split(";") if part and part.strip()]


def _clean_cell(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return int(value)
    return value


def _date_slug(value: Any) -> str:
    text = str(value or "").strip()
    try:
        return datetime.strptime(text, "%d/%m/%Y").strftime("%Y-%m-%d")
    except ValueError:
        return _slugify(text)


def _slugify(value: str) -> str:
    text = str(value).lower()
    text = re.sub(r"[^a-z0-9]+", "-", text)
    return text.strip("-") or "untitled"


def _relative_to(path: Path, root: Path) -> str:
    resolved = path.resolve()
    try:
        return str(resolved.relative_to(root.resolve()))
    except ValueError:
        return str(resolved)


def _sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as file:
        for chunk in iter(lambda: file.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()
