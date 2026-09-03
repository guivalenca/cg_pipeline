"""Focused API/static tests for the Syllabus HTML surface."""

import asyncio
import hashlib
import json
import textwrap
import uuid
from io import BytesIO
from pathlib import Path
from zipfile import ZipFile

import psycopg
import pytest
from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook
from psycopg.types.json import Jsonb

from fake_companion import accepting_companion
from test_graph_revision import _seed_finished_build, _seed_subject
from universe import graph_revision
from universe.assets import LocalAssetStore
from universe.graph_identity import GRAPH_ID_CONFLICT_MESSAGE, subject_graph_id_for
from adalove_workbook import activity, write_adalove_workbook
from universe.web import app as web_app
from universe.web.acquisition_app import _markdown_renderer
from universe.web.app import (
    _diagnostic_message,
    _image_failure_message,
    _summarize_image_counts,
    create_app,
)


def test_markdown_renderer_typesets_inline_math_as_mathml():
    html = _markdown_renderer().render(r"Energy is $E = mc^2$.")

    assert "<math" in html
    assert 'xmlns="http://www.w3.org/1998/Math/MathML"' in html
    assert 'display="inline"' in html
    assert "<msup>" in html
    assert "$E = mc^2$" not in html


def test_markdown_renderer_typesets_block_math_as_mathml():
    html = _markdown_renderer().render("$$\n\\frac{a}{b} = c\n$$")

    assert "<math" in html
    assert 'xmlns="http://www.w3.org/1998/Math/MathML"' in html
    assert 'display="block"' in html
    assert "<mfrac>" in html
    assert "syl-math-block" in html


def test_markdown_renderer_keeps_malformed_math_readable():
    html = _markdown_renderer().render(r"Bad formula: $\frac{1}{2$ after text.")

    assert "syl-math-source" in html
    assert r"\frac{1}{2" in html
    assert "after text" in html
    assert "<math" not in html


def test_math_support_preserves_markdown_table_and_image_safety():
    html = _markdown_renderer().render(
        "| Formula | Value |\n"
        "| --- | --- |\n"
        "| Inline | $x^2$ |\n\n"
        "<script>alert('x')</script>\n\n"
        r"Unsafe $\href{javascript:alert(1)}{x}$" "\n\n"
        "![Tracker](https://tracker.example/pixel.png)\n\n"
        "![Local](/api/source-assets/asset-safe)"
    )

    assert "<table>" in html
    assert "<math" in html
    assert "<script>" not in html
    assert "&lt;script&gt;" in html
    assert 'href="javascript:' not in html
    assert "syl-math-source" in html
    assert '<img src="https://tracker.example' not in html
    assert "imagem remota não carregada" in html
    assert '<img src="/api/source-assets/asset-safe"' in html


def test_static_surface_has_only_the_syllabus_operator_route(test_database_url):
    with TestClient(_app(test_database_url)) as client:
        assert client.get("/", follow_redirects=False).headers["location"] == "/syllabi"
        surface = client.get("/syllabi")
        assert surface.status_code == 200
        assert "data-lesson-build-dialog" in surface.text

        script = client.get("/static/syllabi.js")
        assert script.status_code == 200
        assert "data-lesson-build-accept" in script.text
        assert "data-lesson-build-reject" in script.text
        assert "function graphRevisionMarkup" in script.text
        assert "data-companion-package" in script.text
        assert "Baixar pacote Companion" in script.text
        assert "await downloadCompanionPackage" in script.text
        lesson_build_handler = script.text.split(
            "lessonBuildDialog.addEventListener('click'", 1
        )[1].split("lessonBuildDialog.addEventListener('cancel'", 1)[0]
        assert "if (packageButton)" in lesson_build_handler
        assert "data-lesson-build-error" in lesson_build_handler
        assert "error.message" in lesson_build_handler

        assert client.get("/graph").status_code == 404


def _namespace(graph_ids=()):
    return {
        "schema_version": "companion_graph_namespace.v1",
        "institutions": [
            {"slug": "web-inteli", "name": "Inteli Web"},
            {"slug": "web-other", "name": "Other Web Institution"},
        ],
        "graph_ids": list(graph_ids),
    }


def _app(
    database_url: str,
    asset_store=None,
    *,
    graph_ids=(),
    companion_repo: Path | None = None,
):
    options = {}
    if asset_store is not None:
        options["asset_store_factory"] = lambda: asset_store
    options["companion_namespace_provider"] = lambda: _namespace(graph_ids)
    if companion_repo is not None:
        options["companion_repo"] = companion_repo
    return create_app(lambda: psycopg.connect(database_url), **options)


def _rejecting_companion(tmp_path: Path, code: str) -> Path:
    companion = tmp_path / "rejecting-companion"
    scripts = companion / "scripts"
    scripts.mkdir(parents=True)
    (scripts / "validate_graph_package.py").write_text(
        textwrap.dedent(
            f"""
            import json

            print(json.dumps({{
                "schema_version": "companion_graph_package_acceptance.v1",
                "accepted": False,
                "graph_id": None,
                "package_sha256": None,
                "issues": [{{"code": {code!r}}}],
            }}))
            raise SystemExit(2)
            """
        ),
        encoding="utf-8",
    )
    return companion


def _png_bytes(width=20, height=10):
    return (
        b"\x89PNG\r\n\x1a\n"
        + (13).to_bytes(4, "big")
        + b"IHDR"
        + width.to_bytes(4, "big")
        + height.to_bytes(4, "big")
        + b"\x08\x02\x00\x00\x00"
        + b"\x00\x00\x00\x00"
    )


def _workbook(
    path: Path,
    *,
    project: str,
    lesson: str,
    subject: str = "Negócios",
    source_title: str = "Fonte",
    source_url: str = "https://example.com/material",
    source_description: str = "Uma descrição útil.",
) -> Path:
    lesson_row = activity(
        title=lesson,
        kind="Class",
        week=1,
        order=1,
        subject=subject,
    )
    source_row = activity(
        title=source_title,
        kind="Self-study",
        week=1,
        order=2,
        parent_uuid=lesson_row["Activity UUID"],
        parent_title=lesson,
        subject=subject,
        description=source_description,
        url=source_url,
    )
    return write_adalove_workbook(path, [lesson_row, source_row], project=project)


def _five_subject_workbook(path: Path) -> Path:
    rows = []
    for index, (code, display_name) in enumerate(
        (
            ("COM", "Comunicação"),
            ("MTF", "Matemática"),
            ("NEG", "Negócios"),
            ("UEX", "User Experience"),
            ("LID", "Liderança"),
        ),
        1,
    ):
        lesson = activity(
            week=index,
            order=index * 2 - 1,
            kind="Class",
            title=f"Aula de {display_name}",
            subject=code,
        )
        rows.extend(
            [
                lesson,
                activity(
                    week=index,
                    order=index * 2,
                    kind="Self-study",
                    title=f"Fonte de {display_name}",
                    subject=None,
                    parent_uuid=lesson["Activity UUID"],
                    parent_title=lesson["Title"],
                    url=f"https://example.com/{code.lower()}",
                ),
            ]
        )
    rows.append(
        activity(
            week=6,
            order=11,
            kind="Orientation",
            title="Orientação não curricular",
            subject=None,
        )
    )
    return write_adalove_workbook(path, rows, project="GRAD PILOT")


def _publish_source(conn, source: dict, marker: str, *, content_hash: str = "a" * 64):
    snapshot_id = f"snapshot-{marker}"
    artifact_id = f"artifact-{marker}"
    conn.execute(
        "INSERT INTO source_snapshot"
        " (id, source_id, content_hash, status) VALUES (%s, %s, %s, 'ok')",
        (snapshot_id, source["source_id"], content_hash),
    )
    conn.execute(
        "INSERT INTO artifact (id, snapshot_id, kind, tool, body)"
        " VALUES (%s, %s, 'markdown', 'test', '# Fonte')",
        (artifact_id, snapshot_id),
    )
    return artifact_id


def _upload(
    client: TestClient,
    path: Path,
    name: str,
    syllabus_id: str | None = None,
    *,
    institution_id: str | None = None,
):
    data = {"name": name}
    if syllabus_id:
        data["syllabus_id"] = syllabus_id
    else:
        data.update(
            {
                "institution_id": institution_id or "web-inteli",
            }
        )
    with path.open("rb") as workbook:
        return client.post(
            "/api/syllabi/upload",
            data=data,
            files={
                "file": (
                    path.name,
                    workbook,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )


def test_malformed_workbook_has_a_clear_validation_error(test_database_url, tmp_path):
    path = tmp_path / "malformed.xlsx"
    workbook = Workbook()
    workbook.active.title = "Activities"
    workbook.save(path)
    workbook.close()

    with TestClient(_app(test_database_url)) as client:
        response = _upload(client, path, f"Malformed {uuid.uuid4().hex[:8]}")

    assert response.status_code == 422
    assert "aba" in response.json()["detail"].casefold()


def test_named_upload_creates_a_visible_syllabus_without_queueing(
    test_database_url, applied_migrations, tmp_path
):
    name = f"Syllabus Web {uuid.uuid4().hex[:8]}"
    path = _workbook(tmp_path / "first.xlsx", project="Ignored workbook title", lesson="Aula 1")

    with TestClient(_app(test_database_url)) as client:
        uploaded = _upload(
            client,
            path,
            name,
        )
        assert uploaded.status_code == 201, uploaded.text
        result = uploaded.json()
        assert result["unchanged"] is False

        index = client.get("/api/syllabi")
        entry = next(item for item in index.json()["syllabi"] if item["id"] == result["syllabus_id"])
        graph_id = subject_graph_id_for("web-inteli", name, "NEG")
        assert entry["title"] == name
        assert entry["display_name"] == name
        assert entry["institution"] == {"id": "web-inteli", "name": "Inteli Web"}
        assert entry["institution_slug"] == "web-inteli"
        assert entry["lesson_subjects"] == [
            {"code": "NEG", "display_name": "Negócios", "graph_id": graph_id}
        ]
        assert entry["metadata_complete"] is True
        assert entry["export_identities"] == [
            {
                "graph_id": graph_id,
                "display_name": f"{name} · Negócios",
                "institution_slug": "web-inteli",
                "lesson_subject_code": "NEG",
            }
        ]
        assert entry["latest"]["lesson_count"] == 1
        assert entry["latest"]["source_count"] == 1

        detail = client.get(f"/api/syllabi/{result['syllabus_id']}").json()
        assert detail["title"] == name
        assert detail["display_name"] == name
        assert detail["institution_slug"] == "web-inteli"
        assert detail["lessons"][0]["lesson_subject"] == {
            "code": "NEG",
            "display_name": "Negócios",
            "graph_id": graph_id,
        }
        assert detail["lessons"][0]["title"] == "Aula 1"
        assert detail["lessons"][0]["sources"][0]["has_markdown"] is False

        download = client.get(
            f"/api/syllabi/{result['syllabus_id']}/versions/{result['version_id']}/workbook"
        )
        assert download.status_code == 200
        assert download.content == path.read_bytes()
        assert "first.xlsx" in download.headers["content-disposition"]

    with psycopg.connect(test_database_url) as conn:
        assert conn.execute(
            "SELECT count(*) FROM acquisition_job WHERE source_id = %s",
            (detail["lessons"][0]["sources"][0]["source_id"],),
        ).fetchone()[0] == 0


def test_new_upload_reports_dropped_orientations(
    test_database_url, applied_migrations, tmp_path
):
    lesson_row = activity(title="Aula 1", kind="Class", week=1, order=1, subject="Negócios")
    orientation_row = activity(
        title="Sprint Planning", kind="Orientation", week=1, order=2, subject=None
    )
    path = write_adalove_workbook(
        tmp_path / "dropped.xlsx", [lesson_row, orientation_row], project="Dropped"
    )

    with TestClient(_app(test_database_url)) as client:
        uploaded = _upload(client, path, f"Syllabus Drop {uuid.uuid4().hex[:8]}")

    assert uploaded.status_code == 201, uploaded.text
    result = uploaded.json()
    assert result["dropped_summary"] == {
        "orientation_count": 1,
        "orientation_self_study_count": 0,
        "no_parent_count": 0,
        "total_count": 1,
    }
    assert [(item["title"], item["reason"]) for item in result["dropped"]] == [
        ("Sprint Planning", "orientation")
    ]


def test_new_syllabus_upload_requires_complete_valid_syllabus_metadata(
    test_database_url, applied_migrations, tmp_path
):
    path = _workbook(
        tmp_path / "metadata.xlsx",
        project="Project",
        lesson="Aula 1",
    )

    with TestClient(_app(test_database_url)) as client, path.open("rb") as workbook:
        missing = client.post(
            "/api/syllabi/upload",
            data={"name": "Sem metadados"},
            files={
                "file": (
                    path.name,
                    workbook,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
    assert missing.status_code == 422
    assert "instituição" in missing.json()["detail"]

    with TestClient(_app(test_database_url)) as client, path.open("rb") as workbook:
        invalid = client.post(
            "/api/syllabi/upload",
            data={
                "name": "Metadados inválidos",
                "institution_id": "unknown-institution",
            },
            files={
                "file": (
                    path.name,
                    workbook,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
    assert invalid.status_code == 422
    assert "instituição existente" in invalid.json()["detail"]

    with TestClient(_app(test_database_url)) as client, path.open("rb") as workbook:
        nonexistent_target = client.post(
            "/api/syllabi/upload",
            data={"name": "Alvo inexistente", "syllabus_id": "missing-target"},
            files={
                "file": (
                    path.name,
                    workbook,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
    assert nonexistent_target.status_code == 404
    assert "unknown syllabus" in nonexistent_target.json()["detail"]


def test_subject_graph_id_conflict_requires_a_new_name_and_ignores_form_override(
    test_database_url, applied_migrations, tmp_path
):
    name = f"Graph conflict {uuid.uuid4().hex[:8]}"
    expected = subject_graph_id_for("web-inteli", name, "NEG")
    path = _workbook(tmp_path / "graph-conflict.xlsx", project="Project", lesson="Aula")

    with TestClient(_app(test_database_url, graph_ids=(expected,))) as client:
        conflict = _upload(client, path, name)
        assert conflict.status_code == 409
        assert conflict.json()["detail"] == {
            "code": "graph_id_conflict",
            "message": GRAPH_ID_CONFLICT_MESSAGE,
            "graph_id": expected,
        }

        with path.open("rb") as workbook:
            attempted_override = client.post(
                "/api/syllabi/upload",
                data={
                    "name": name,
                    "institution_id": "web-inteli",
                    "graph_id": f"{expected}-2",
                },
                files={
                    "file": (
                        path.name,
                        workbook,
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )
                },
            )
        assert attempted_override.status_code == 409
        assert attempted_override.json()["detail"]["graph_id"] == expected
        assert client.get("/api/syllabi/graph-id-availability").status_code == 404


def test_new_version_rejects_an_occupied_subject_graph_id(
    test_database_url, applied_migrations, tmp_path
):
    syllabus_id = f"version-path-conflict-{uuid.uuid4().hex[:8]}"
    occupied_graph_id = subject_graph_id_for("web-inteli", syllabus_id, "NEG")
    version_id = f"{syllabus_id}:v0001"
    path = _workbook(
        tmp_path / "version-path-conflict.xlsx",
        project="Project",
        lesson="Aula atualizada",
    )
    with psycopg.connect(test_database_url) as conn:
        conn.execute(
            "INSERT INTO institution (id, name) VALUES ('web-inteli', 'Inteli Web')"
            " ON CONFLICT (id) DO NOTHING"
        )
        conn.execute(
            "INSERT INTO syllabus (id, title, institution_id) VALUES (%s, %s, 'web-inteli')",
            (syllabus_id, syllabus_id),
        )
        conn.execute(
            "INSERT INTO syllabus_version (id, syllabus_id, seq, origin)"
            " VALUES (%s, %s, 1, 'upload')",
            (version_id, syllabus_id),
        )
        conn.commit()

    with TestClient(_app(test_database_url, graph_ids=(occupied_graph_id,))) as client:
        conflict = _upload(client, path, syllabus_id, syllabus_id)

    assert conflict.status_code == 409
    assert conflict.json()["detail"] == {
        "code": "graph_id_conflict",
        "message": GRAPH_ID_CONFLICT_MESSAGE,
        "graph_id": occupied_graph_id,
    }


def test_subject_graph_id_conflict_with_an_existing_local_owner_is_readable(
    test_database_url, applied_migrations, tmp_path
):
    name = f"Local conflict {uuid.uuid4().hex[:8]}"
    expected = subject_graph_id_for("web-inteli", name, "NEG")
    owner_id = f"graph-owner-{uuid.uuid4().hex[:8]}"
    with psycopg.connect(test_database_url) as conn:
        conn.execute(
            "INSERT INTO syllabus (id, title) VALUES (%s, 'Graph owner')",
            (owner_id,),
        )
        conn.execute(
            "INSERT INTO syllabus_subject"
            " (syllabus_id, lesson_subject_code, graph_id) VALUES (%s, 'NEG', %s)",
            (owner_id, expected),
        )

    path = _workbook(tmp_path / "local-conflict.xlsx", project="Project", lesson="Aula")
    with TestClient(_app(test_database_url)) as client:
        conflict = _upload(client, path, name)

    assert conflict.status_code == 409
    assert conflict.json()["detail"] == {
        "code": "graph_id_conflict",
        "message": GRAPH_ID_CONFLICT_MESSAGE,
        "graph_id": expected,
    }


def test_graph_id_proposal_describes_a_template_and_existing_syllabus_owners(
    test_database_url, applied_migrations, tmp_path
):
    marker = uuid.uuid4().hex[:8]
    name = f"Proposal {marker}"
    path = _workbook(tmp_path / "proposal.xlsx", project="Project", lesson="Aula")
    template = subject_graph_id_for("web-inteli", name, "subject").removesuffix(
        "-subject"
    ) + "-<subject>"

    with TestClient(_app(test_database_url)) as client:
        before = client.get(
            "/api/syllabi/graph-id-proposal",
            params={"institution_id": "web-inteli", "name": name},
        )
        created = _upload(client, path, name)
        after = client.get(
            "/api/syllabi/graph-id-proposal",
            params={"institution_id": "web-inteli", "name": name},
        )
        duplicate = _upload(client, path, name)

    assert before.status_code == 200, before.text
    assert before.json() == {
        "display_name": name,
        "graph_id_template": template,
        "existing_syllabus": None,
        "syllabus_id_owner": None,
    }
    assert created.status_code == 201, created.text
    owner = {"id": created.json()["syllabus_id"], "title": name}
    assert after.json() == {
        "display_name": name,
        "graph_id_template": template,
        "existing_syllabus": owner,
        "syllabus_id_owner": owner,
    }
    assert duplicate.status_code == 409
    assert duplicate.json()["detail"] == {
        "code": "syllabus_already_exists",
        "message": "Este nome já existe. Você está adicionando uma versão a esse syllabus.",
        "syllabus_id": created.json()["syllabus_id"],
    }


def test_graph_id_proposal_finds_an_exact_historical_title_independently_of_its_id(
    test_database_url, applied_migrations, tmp_path
):
    marker = uuid.uuid4().hex[:8]
    name = f"Historical syllabus {marker}"
    historical_id = f"historical-{marker}"
    with psycopg.connect(test_database_url) as conn:
        conn.execute(
            "INSERT INTO syllabus (id, title) VALUES (%s, %s)",
            (historical_id, name),
        )

    with TestClient(_app(test_database_url)) as client:
        proposal = client.get(
            "/api/syllabi/graph-id-proposal",
            params={"institution_id": "web-inteli", "name": name},
        )
        path = _workbook(
            tmp_path / "historical-title.xlsx", project="Project", lesson="Aula"
        )
        submission = _upload(client, path, name)

    assert proposal.status_code == 200, proposal.text
    assert proposal.json()["existing_syllabus"] == {
        "id": historical_id,
        "title": name,
    }
    assert submission.status_code == 409
    assert submission.json()["detail"] == {
        "code": "syllabus_already_exists",
        "message": "Este nome já existe. Você está adicionando uma versão a esse syllabus.",
        "syllabus_id": historical_id,
    }


def test_graph_id_proposal_reports_a_different_owner_of_the_normalized_id(
    test_database_url, applied_migrations, tmp_path
):
    marker = uuid.uuid4().hex[:8]
    existing_name = f"Math 101 {marker}"
    colliding_name = f"Math-101-{marker}"
    path = _workbook(
        tmp_path / "normalized-collision.xlsx", project="Project", lesson="Aula"
    )

    with TestClient(_app(test_database_url)) as client:
        created = _upload(client, path, existing_name)
        assert created.status_code == 201, created.text
        proposal = client.get(
            "/api/syllabi/graph-id-proposal",
            params={"institution_id": "web-inteli", "name": colliding_name},
        )
        collision = _upload(client, path, colliding_name)

    owner = {"id": created.json()["syllabus_id"], "title": existing_name}
    assert proposal.status_code == 200, proposal.text
    assert proposal.json()["existing_syllabus"] is None
    assert proposal.json()["syllabus_id_owner"] == owner
    assert collision.status_code == 409
    assert collision.json()["detail"] == {
        "code": "graph_id_conflict",
        "message": GRAPH_ID_CONFLICT_MESSAGE,
        "graph_id": subject_graph_id_for("web-inteli", colliding_name, "NEG"),
    }


def test_graph_id_proposal_rejects_a_name_that_cannot_become_a_syllabus_id(
    test_database_url, applied_migrations
):
    with TestClient(_app(test_database_url)) as client:
        proposal = client.get(
            "/api/syllabi/graph-id-proposal",
            params={"institution_id": "web-inteli", "name": "1"},
        )

    assert proposal.status_code == 422
    assert "identificador do syllabus" in proposal.json()["detail"]


def test_metadata_incomplete_legacy_syllabus_remains_readable(
    test_database_url, applied_migrations
):
    syllabus_id = f"legacy-readable-{uuid.uuid4().hex[:8]}"
    version_id = f"{syllabus_id}:v0001"
    with psycopg.connect(test_database_url) as conn:
        conn.execute(
            "INSERT INTO syllabus (id, title) VALUES (%s, 'Legacy readable')",
            (syllabus_id,),
        )
        conn.execute(
            "INSERT INTO syllabus_version (id, syllabus_id, seq, origin)"
            " VALUES (%s, %s, 1, 'upload')",
            (version_id, syllabus_id),
        )
        conn.commit()

    with TestClient(_app(test_database_url)) as client:
        detail = client.get(f"/api/syllabi/{syllabus_id}")

    assert detail.status_code == 200, detail.text
    assert detail.json()["institution"] is None
    assert detail.json()["lesson_subjects"] == []
    assert detail.json()["institution_slug"] is None
    assert detail.json()["metadata_complete"] is False
    assert detail.json()["export_identities"] == []


def test_version_query_loads_that_versions_actual_lessons(
    test_database_url, applied_migrations, tmp_path
):
    name = f"Versioned Web {uuid.uuid4().hex[:8]}"
    first = _workbook(tmp_path / "v1.xlsx", project="Project", lesson="Aula original")
    second = _workbook(tmp_path / "v2.xlsx", project="Project", lesson="Aula atualizada")

    with TestClient(_app(test_database_url)) as client:
        v1 = _upload(client, first, name).json()
        v2_response = _upload(client, second, name, v1["syllabus_id"])
        assert v2_response.status_code == 201, v2_response.text
        v2 = v2_response.json()

        old = client.get(
            f"/api/syllabi/{v1['syllabus_id']}",
            params={"version_id": v1["version_id"]},
        ).json()
        current = client.get(
            f"/api/syllabi/{v1['syllabus_id']}",
            params={"version_id": v2["version_id"]},
        ).json()

    assert old["version"]["id"] == v1["version_id"]
    assert old["lessons"][0]["title"] == "Aula original"
    assert current["version"]["id"] == v2["version_id"]
    assert current["lessons"][0]["title"] == "Aula atualizada"
    assert [version["id"] for version in current["versions"]][:2] == [
        v2["version_id"],
        v1["version_id"],
    ]


def test_new_workbook_is_reconciled_before_it_becomes_the_current_version(
    test_database_url, applied_migrations, tmp_path
):
    name = f"Syllabus reconciliation web {uuid.uuid4().hex[:8]}"
    original = _workbook(
        tmp_path / "reconciliation-original.xlsx",
        project="Project",
        lesson="Aula",
        source_description="Descrição original",
    )
    incoming = _workbook(
        tmp_path / "reconciliation-incoming.xlsx",
        project="Project",
        lesson="Aula",
        source_description="Descrição nova",
    )
    with TestClient(_app(test_database_url)) as client:
        uploaded = _upload(client, original, name).json()
        with incoming.open("rb") as workbook:
            preview_response = client.post(
                f"/api/syllabi/{uploaded['syllabus_id']}/reconciliations",
                files={
                    "file": (
                        incoming.name,
                        workbook,
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )
                },
            )
        assert preview_response.status_code == 201, preview_response.text
        preview = preview_response.json()
        assert preview["base_version_id"] == uploaded["version_id"]
        assert preview["summary"]["action_count"] == 1
        changed_source = next(
            source
            for lesson in preview["lessons"]
            for source in lesson["sources"]
            if source["status"] == "changed"
        )

        still_current = client.get(
            f"/api/syllabi/{uploaded['syllabus_id']}"
        ).json()
        assert still_current["version"]["id"] == uploaded["version_id"]
        assert still_current["lessons"][0]["sources"][0]["description"] == "Descrição original"

        applied = client.post(
            f"/api/syllabi/{uploaded['syllabus_id']}/reconciliations/{preview['id']}/apply",
            json={
                "decisions": {changed_source["item_id"]: "transition"},
                "drafts": {},
            },
        )
        assert applied.status_code == 201, applied.text
        assert applied.json()["version_id"] != uploaded["version_id"]
        latest = client.get(f"/api/syllabi/{uploaded['syllabus_id']}").json()
        assert latest["lessons"][0]["sources"][0]["description"] == "Descrição nova"


def test_editor_api_saves_new_version_with_hidden_added_and_reordered_sources(
    test_database_url, applied_migrations, tmp_path
):
    name = f"Editable Web {uuid.uuid4().hex[:8]}"
    path = _workbook(tmp_path / "editable-web.xlsx", project="Project", lesson="Aula")

    with TestClient(_app(test_database_url)) as client:
        uploaded = _upload(client, path, name).json()
        detail = client.get(f"/api/syllabi/{uploaded['syllabus_id']}").json()
        lesson = detail["lessons"][0]
        original = lesson["sources"][0]
        with psycopg.connect(test_database_url) as conn:
            _publish_source(conn, original, f"editor-{uuid.uuid4().hex}")
        reviewed = client.patch(
            f"/api/syllabi/{uploaded['syllabus_id']}/sources/{original['reference_id']}/review",
            json={"validated": True, "complexity": "complex"},
        )
        assert reviewed.status_code == 200, reviewed.text
        payload = {
            "base_version_id": uploaded["version_id"],
            "note": "Reorganiza a aula e inclui uma nova fonte.",
            "lessons": [
                {
                    "id": lesson["id"],
                    "hidden": True,
                    "week": lesson["week"],
                    "kind": lesson["kind"],
                    "title": "Aula revisada",
                    "subject": lesson["subject"],
                    "date": lesson["date"],
                    "description": "Descrição da aula revisada",
                    "sources": [
                        {
                            "title": "Nova fonte",
                            "description": "Adicionada pela UI",
                            "url": "https://example.com/nova",
                            "media_type": "article",
                            "hidden": False,
                        },
                        {
                            "reference_id": original["reference_id"],
                            "title": original["title"],
                            "description": original["description"],
                            "url": original["url"],
                            "media_type": original["media_type"],
                            "hidden": True,
                        },
                    ],
                }
            ],
        }

        saved = client.post(
            f"/api/syllabi/{uploaded['syllabus_id']}/curate", json=payload
        )

        assert saved.status_code == 201, saved.text
        assert saved.json()["seq"] == 2
        refreshed = client.get(f"/api/syllabi/{uploaded['syllabus_id']}").json()
        assert refreshed["version"]["note"] == (
            "Reorganiza a aula e inclui uma nova fonte."
        )
        assert refreshed["lessons"][0]["title"] == "Aula revisada"
        assert refreshed["lessons"][0]["hidden"] is True
        assert refreshed["lessons"][0]["sources"][1]["review"] == {
            "validated": False,
            "complexity": "complex",
        }
        assert [source["title"] for source in refreshed["lessons"][0]["sources"]] == [
            "Nova fonte",
            "Fonte",
        ]
        assert refreshed["lessons"][0]["sources"][1]["hidden"] is True
        workbook = client.get(
            f"/api/syllabi/{uploaded['syllabus_id']}/versions/"
            f"{saved.json()['version_id']}/workbook"
        )
        assert workbook.status_code == 200
        assert workbook.content.startswith(b"PK")
        exported = load_workbook(BytesIO(workbook.content), read_only=True)
        sheet = exported["Activities"]
        rows = sheet.iter_rows()
        headers = [cell.value for cell in next(rows)]
        lesson_row = [cell.value for cell in next(rows)]
        assert lesson_row[headers.index("Hidden")] == "yes"
        exported.close()

        stale = client.post(
            f"/api/syllabi/{uploaded['syllabus_id']}/curate", json=payload
        )
        assert stale.status_code == 409
        assert "versão mais nova" in stale.json()["detail"]


def test_editor_api_requires_a_bounded_reason_for_a_new_version(
    test_database_url, applied_migrations, tmp_path
):
    path = _workbook(
        tmp_path / "version-reason.xlsx", project="Project", lesson="Aula"
    )
    with TestClient(_app(test_database_url)) as client:
        uploaded = _upload(client, path, f"Reason {uuid.uuid4().hex[:8]}").json()
        detail = client.get(f"/api/syllabi/{uploaded['syllabus_id']}").json()
        lessons = detail["lessons"]
        lessons[0]["title"] = "Aula revisada"

        missing = client.post(
            f"/api/syllabi/{uploaded['syllabus_id']}/curate",
            json={"base_version_id": uploaded["version_id"], "lessons": lessons},
        )
        too_long = client.post(
            f"/api/syllabi/{uploaded['syllabus_id']}/curate",
            json={
                "base_version_id": uploaded["version_id"],
                "note": "x" * 501,
                "lessons": lessons,
            },
        )

        assert missing.status_code == 422
        assert missing.json()["detail"] == "A razão da nova versão é obrigatória."
        assert too_long.status_code == 422
        assert too_long.json()["detail"] == (
            "A razão da nova versão deve ter no máximo 500 caracteres."
        )

        excessive_description = [dict(lesson) for lesson in lessons]
        excessive_description[0] = {
            **excessive_description[0],
            "description": "d" * 20_001,
        }
        rejected_description = client.post(
            f"/api/syllabi/{uploaded['syllabus_id']}/curate",
            json={
                "base_version_id": uploaded["version_id"],
                "note": "Atualiza a descrição.",
                "lessons": excessive_description,
            },
        )
        assert rejected_description.status_code == 422
        assert "description exceeds 20000 characters" in (
            rejected_description.json()["detail"]
        )


def test_source_review_can_be_marked_validated_and_simple(
    test_database_url, applied_migrations, tmp_path
):
    name = f"Source Review {uuid.uuid4().hex[:8]}"
    path = _workbook(
        tmp_path / "source-review.xlsx",
        project="Project",
        lesson="Aula",
        source_url=f"https://example.com/review-{uuid.uuid4().hex}",
    )

    with TestClient(_app(test_database_url)) as client:
        uploaded = _upload(client, path, name).json()
        detail = client.get(f"/api/syllabi/{uploaded['syllabus_id']}").json()
        source = detail["lessons"][0]["sources"][0]

        missing_publication = client.patch(
            f"/api/syllabi/{uploaded['syllabus_id']}/sources/{source['reference_id']}/review",
            json={"validated": True, "complexity": "simple"},
        )
        assert missing_publication.status_code == 422
        assert "publication" in missing_publication.json()["detail"].casefold()

        with psycopg.connect(test_database_url) as conn:
            _publish_source(conn, source, f"review-{uuid.uuid4().hex}")
        reviewed = client.patch(
            f"/api/syllabi/{uploaded['syllabus_id']}/sources/{source['reference_id']}/review",
            json={"validated": True, "complexity": "simple"},
        )

        assert reviewed.status_code == 200, reviewed.text
        assert reviewed.json()["review"] == {
            "validated": True,
            "complexity": "simple",
        }
        refreshed = client.get(f"/api/syllabi/{uploaded['syllabus_id']}").json()
        assert refreshed["lessons"][0]["sources"][0]["review"] == {
            "validated": True,
            "complexity": "simple",
        }

        with psycopg.connect(test_database_url) as conn:
            _publish_source(
                conn,
                source,
                f"zz-review-{uuid.uuid4().hex}",
                content_hash="b" * 64,
            )
        refreshed_after_publication = client.get(
            f"/api/syllabi/{uploaded['syllabus_id']}"
        ).json()

    assert refreshed_after_publication["lessons"][0]["sources"][0]["review"] == {
        "validated": False,
        "complexity": "simple",
    }


def test_syllabus_detail_summarizes_recorded_openrouter_cost_and_firecrawl_usage(
    test_database_url, applied_migrations, tmp_path
):
    name = f"Usage Summary {uuid.uuid4().hex[:8]}"
    path = _workbook(tmp_path / "usage-summary.xlsx", project="Project", lesson="Aula")

    with TestClient(_app(test_database_url)) as client:
        uploaded = _upload(client, path, name).json()
        detail = client.get(f"/api/syllabi/{uploaded['syllabus_id']}").json()
        source_id = detail["lessons"][0]["sources"][0]["source_id"]
        snapshot_id = f"snapshot-{uuid.uuid4().hex}"
        artifact_id = f"artifact-{uuid.uuid4().hex}"
        job_id = f"job-{uuid.uuid4().hex}"
        run_id = f"run-{uuid.uuid4().hex}"
        with psycopg.connect(test_database_url) as conn:
            conn.execute(
                "INSERT INTO source_snapshot"
                " (id, source_id, content_hash, status) VALUES (%s, %s, %s, 'ok')",
                (snapshot_id, source_id, "a" * 64),
            )
            conn.execute(
                "INSERT INTO artifact (id, snapshot_id, kind, tool, body)"
                " VALUES (%s, %s, 'markdown', 'firecrawl', '# Source')",
                (artifact_id, snapshot_id),
            )
            conn.execute(
                "INSERT INTO acquisition_job"
                " (id, source_id, status, provider, attempt_count, artifact_id, diagnostics,"
                "  finished_at) VALUES (%s, %s, 'succeeded', 'firecrawl/v2', 1, %s, %s, now())",
                (job_id, source_id, artifact_id, Jsonb({"provider_attempts": 2})),
            )
            conn.execute(
                "INSERT INTO run"
                " (id, stage, model, prompt_ref, prompt_sha, status, finished_at)"
                " VALUES (%s, 'passage-triage', 'openrouter/model', 'prompt/v1', %s,"
                "  'done', now())",
                (run_id, "b" * 64),
            )
            conn.execute(
                "INSERT INTO run_item"
                " (id, run_id, artifact_id, response, usage) VALUES (%s, %s, %s, %s, %s)",
                (
                    f"item-{uuid.uuid4().hex}",
                    run_id,
                    artifact_id,
                    '{}',
                    Jsonb({"cost": 0.12, "total_tokens": 100}),
                ),
            )
            conn.execute(
                "INSERT INTO source_image_analysis_call"
                " (id, markdown_artifact_id, prompt_ref, prompt_sha, requested_model,"
                "  status, usage, finished_at) VALUES (%s, %s, 'images/v1', %s,"
                "  'google/gemini', 'succeeded', %s, now())",
                (
                    f"image-call-{uuid.uuid4().hex}",
                    artifact_id,
                    "c" * 64,
                    Jsonb({"cost": 0.03, "total_tokens": 50}),
                ),
            )
            conn.commit()

        refreshed = client.get(f"/api/syllabi/{uploaded['syllabus_id']}").json()

        assert refreshed["usage"] == {
            "openrouter": {"cost_usd": 0.15, "calls": 2, "total_tokens": 150},
            "firecrawl": {"extractions": 1, "attempts": 2, "succeeded": 1, "failed": 0},
        }


def test_queue_endpoint_enqueues_only_the_clicked_source(
    test_database_url, applied_migrations, tmp_path
):
    name = f"Queue Web {uuid.uuid4().hex[:8]}"
    path = _workbook(
        tmp_path / "queue.xlsx",
        project="Project",
        lesson="Aula",
        source_url=f"https://example.com/queue-{uuid.uuid4().hex}",
    )

    with TestClient(_app(test_database_url)) as client:
        uploaded = _upload(client, path, name).json()
        detail = client.get(f"/api/syllabi/{uploaded['syllabus_id']}").json()
        source_id = detail["lessons"][0]["sources"][0]["source_id"]

        queued = client.post(f"/api/sources/{source_id}/queue")
        assert queued.status_code == 202, queued.text
        assert queued.json()["job"]["source_id"] == source_id
        assert queued.json()["job"]["status"] == "queued"

        refreshed = client.get(f"/api/syllabi/{uploaded['syllabus_id']}").json()
        visible = refreshed["lessons"][0]["sources"][0]
        assert visible["acquisition_capability"] == {
            "supported": True,
            "adapter": "firecrawl",
            "label": "Firecrawl",
        }
        assert visible["job"]["status"] == "queued"
        assert visible["has_markdown"] is False


@pytest.mark.parametrize(
    ("source_url", "source_description", "media_type", "adapter_name"),
    [
        (
            "https://www.youtube.com/watch?v=dQw4w9WgXcQ",
            "Assista ao vídeo.",
            "video",
            "vídeo",
        ),
    ],
)
def test_web_exposes_video_adapter_but_requires_preflight_before_queueing(
    test_database_url,
    applied_migrations,
    tmp_path,
    source_url,
    source_description,
    media_type,
    adapter_name,
):
    path = _workbook(
        tmp_path / f"unsupported-{media_type}.xlsx",
        project="Project",
        lesson="Aula",
        source_url=source_url,
        source_description=source_description,
    )

    with TestClient(_app(test_database_url)) as client:
        uploaded = _upload(client, path, f"Unsupported {uuid.uuid4().hex[:8]}").json()
        detail = client.get(f"/api/syllabi/{uploaded['syllabus_id']}").json()
        source = detail["lessons"][0]["sources"][0]
        source_id = source["source_id"]

        assert source["media_type"] == media_type
        assert source["acquisition_capability"] == {
            "supported": True,
            "adapter": "youtube",
            "label": "YouTube",
        }

        rejected = client.post(f"/api/sources/{source_id}/queue")
        assert rejected.status_code == 409
        assert "preflight" in rejected.json()["detail"].lower()

    with psycopg.connect(test_database_url) as conn:
        assert conn.execute(
            "SELECT count(*) FROM acquisition_job WHERE source_id = %s",
            (source_id,),
        ).fetchone()[0] == 0


def test_web_queues_a_concretely_scoped_book_through_browserbase(
    test_database_url, applied_migrations, tmp_path
):
    path = _workbook(
        tmp_path / "browserbase-book.xlsx",
        project="Project",
        lesson="Aula",
        source_url="https://integrada.minhabiblioteca.com.br/#/books/9788557170322",
        source_description="Leia as páginas 27-28.",
    )

    with TestClient(_app(test_database_url)) as client:
        uploaded = _upload(
            client, path, f"Browserbase book {uuid.uuid4().hex[:8]}"
        ).json()
        detail = client.get(f"/api/syllabi/{uploaded['syllabus_id']}").json()
        source = detail["lessons"][0]["sources"][0]

        assert source["media_type"] == "book"
        assert source["source_id"]
        assert source["scope"] == {"kind": "pages", "value": "27-28"}
        assert source["acquisition_capability"] == {
            "supported": True,
            "adapter": "browserbase-book",
            "label": "Browserbase + reconstrução ordenada",
        }

        queued = client.post(f"/api/sources/{source['source_id']}/queue")
        assert queued.status_code == 202, queued.text
        assert queued.json()["job"]["provider"] == "browserbase-book/v1"


def test_markdown_endpoint_renders_and_escapes_source_html(
    test_database_url, applied_migrations, tmp_path
):
    name = f"Markdown Web {uuid.uuid4().hex[:8]}"
    path = _workbook(tmp_path / "markdown.xlsx", project="Project", lesson="Aula")

    with TestClient(_app(test_database_url)) as client:
        uploaded = _upload(client, path, name).json()
        detail = client.get(f"/api/syllabi/{uploaded['syllabus_id']}").json()
        source_id = detail["lessons"][0]["sources"][0]["source_id"]

        markdown = (
            "# Título\n\n<script>alert('x')</script>\n\nTexto útil.\n\n"
            "![Tracker](https://tracker.example/pixel.png)"
        )
        digest = hashlib.sha256(markdown.encode()).hexdigest()
        snapshot_id = f"{source_id}:snap:web-{digest[:8]}"
        artifact_id = f"{snapshot_id}:markdown"
        with psycopg.connect(test_database_url) as conn:
            conn.execute(
                "INSERT INTO source_snapshot (id, source_id, content_hash, status)"
                " VALUES (%s, %s, %s, 'ok')",
                (snapshot_id, source_id, digest),
            )
            conn.execute(
                "INSERT INTO artifact (id, snapshot_id, kind, tool, body)"
                " VALUES (%s, %s, 'markdown', 'test-renderer', %s)",
                (artifact_id, snapshot_id, markdown),
            )
            conn.commit()

        rendered = client.get(f"/api/sources/{source_id}/markdown")
        assert rendered.status_code == 200
        body = rendered.json()
        assert body["artifact_id"] == artifact_id
        assert body["markdown"] == markdown
        assert "<h1>Título</h1>" in body["html"]
        assert "<script>" not in body["html"]
        assert "&lt;script&gt;" in body["html"]
        assert '<img src="https://tracker.example' not in body["html"]
        assert "imagem remota não carregada" in body["html"]
        assert 'href="https://tracker.example/pixel.png"' in body["html"]


def test_new_article_pipeline_withholds_intermediate_markdown_until_latest_clean_artifact(
    test_database_url, applied_migrations, tmp_path
):
    marker = uuid.uuid4().hex[:10]
    path = _workbook(
        tmp_path / "clean-publication.xlsx",
        project="Project",
        lesson="Aula",
        source_url=f"https://example.com/clean-publication-{marker}",
    )
    with TestClient(_app(test_database_url)) as client:
        uploaded = _upload(client, path, f"Clean publication {marker}").json()
        syllabus_url = f"/api/syllabi/{uploaded['syllabus_id']}"
        source_id = client.get(syllabus_url).json()["lessons"][0]["sources"][0]["source_id"]
        snapshot_id = f"snapshot-clean-{marker}"
        base_id = f"{snapshot_id}:markdown"
        enriched_id = f"{base_id}:images"
        job_id = f"acq-clean-{marker}"
        cleanup_job_id = f"cleanup-clean-{marker}"
        with psycopg.connect(test_database_url) as conn:
            conn.execute(
                "INSERT INTO source_snapshot (id, source_id, content_hash, status)"
                " VALUES (%s, %s, %s, 'ok')",
                (snapshot_id, source_id, f"hash-{marker}"),
            )
            conn.execute(
                "INSERT INTO artifact (id, snapshot_id, kind, tool, body)"
                " VALUES (%s, %s, 'markdown', 'firecrawl', '# Intermediate')",
                (base_id, snapshot_id),
            )
            conn.execute(
                "INSERT INTO artifact (id, snapshot_id, kind, tool, body, metadata)"
                " VALUES (%s, %s, 'markdown', 'article-image-association',"
                " '# Enriched but not clean', %s)",
                (enriched_id, snapshot_id, Jsonb({"source_markdown_artifact_id": base_id})),
            )
            conn.execute(
                "INSERT INTO acquisition_job"
                " (id, source_id, status, provider, artifact_id, diagnostics, finished_at)"
                " VALUES (%s, %s, 'succeeded', 'firecrawl/v2', %s, %s, now())",
                (job_id, source_id, base_id, Jsonb({"pipeline_requires_cleanup": True})),
            )
            conn.execute(
                "INSERT INTO source_cleanup_job"
                " (id, acquisition_job_id, source_id, source_artifact_id)"
                " VALUES (%s, %s, %s, %s)",
                (cleanup_job_id, job_id, source_id, enriched_id),
            )
            conn.commit()

        pending = client.get(syllabus_url).json()["lessons"][0]["sources"][0]
        assert pending["pipeline"]["status"] == "cleaning"
        assert pending["has_markdown"] is False
        blocked = client.get(f"/api/sources/{source_id}/markdown")
        assert blocked.status_code == 409
        assert "limpeza" in blocked.json()["detail"]

        run_id = f"web-cut-{marker}"
        cleanup_id = f"pc-clean-{marker}"
        clean_id = f"{enriched_id}:clean:{cleanup_id}"
        with psycopg.connect(test_database_url) as conn:
            conn.execute(
                "INSERT INTO run"
                " (id, stage, model, prompt_ref, prompt_sha, status, finished_at)"
                " VALUES (%s, 'passage-cuts', 'fake/model', 'passage-cuts/v001',"
                " 'abc', 'done', now())",
                (run_id,),
            )
            conn.execute(
                "INSERT INTO passage_cleanup"
                " (id, cuts_run_id, model, triage_prompt_ref, refine_prompt_ref,"
                " status, finished_at)"
                " VALUES (%s, %s, 'fake/model', 'passage-triage/v003',"
                " 'passage-refine/v002', 'done', now())",
                (cleanup_id, run_id),
            )
            conn.execute(
                "INSERT INTO artifact"
                " (id, snapshot_id, kind, tool, tool_version, body, metadata)"
                " VALUES (%s, %s, 'markdown', 'passage-cleanup', 'v1', %s, %s)",
                (
                    clean_id,
                    snapshot_id,
                    "# Canonical clean\n",
                    Jsonb({"source_markdown_artifact_id": enriched_id, "cleanup_id": cleanup_id}),
                ),
            )
            conn.execute(
                "UPDATE source_cleanup_job SET status = 'succeeded', cleanup_id = %s,"
                " canonical_artifact_id = %s, finished_at = now() WHERE id = %s",
                (cleanup_id, clean_id, cleanup_job_id),
            )
            conn.commit()

        ready = client.get(syllabus_url).json()["lessons"][0]["sources"][0]
        assert ready["pipeline"]["status"] == "ready"
        assert ready["has_markdown"] is True
        assert ready["markdown"]["artifact_id"] == clean_id
        published = client.get(f"/api/sources/{source_id}/markdown")
        assert published.status_code == 200
        assert published.json()["artifact_id"] == clean_id
        assert published.json()["markdown"] == "# Canonical clean\n"


def test_pdf_pipeline_withholds_enriched_markdown_until_clean_or_reports_attention(
    test_database_url, applied_migrations, tmp_path
):
    marker = uuid.uuid4().hex[:10]
    path = _workbook(
        tmp_path / "pdf-clean-publication.xlsx",
        project="Project",
        lesson="Aula",
        source_url=f"https://example.com/pdf-clean-publication-{marker}",
    )
    with TestClient(_app(test_database_url)) as client:
        uploaded = _upload(client, path, f"PDF publication {marker}").json()
        syllabus_url = f"/api/syllabi/{uploaded['syllabus_id']}"
        source_id = client.get(syllabus_url).json()["lessons"][0]["sources"][0]["source_id"]
        snapshot_id = f"snapshot-pdf-clean-{marker}"
        raw_id = f"{snapshot_id}:raw-markdown"
        enriched_id = f"{snapshot_id}:markdown"
        job_id = f"acq-pdf-clean-{marker}"
        with psycopg.connect(test_database_url) as conn:
            conn.execute(
                "INSERT INTO source_snapshot (id, source_id, content_hash, status)"
                " VALUES (%s, %s, %s, 'ok')",
                (snapshot_id, source_id, f"hash-{marker}"),
            )
            conn.execute(
                "INSERT INTO artifact (id, snapshot_id, kind, tool, body)"
                " VALUES (%s, %s, 'raw-markdown', 'pdftotext-page-layer', '# Raw PDF')",
                (raw_id, snapshot_id),
            )
            conn.execute(
                "INSERT INTO artifact (id, snapshot_id, kind, tool, body, metadata)"
                " VALUES (%s, %s, 'markdown', 'pdf-page-association',"
                " '# Enriched PDF but not clean', %s)",
                (enriched_id, snapshot_id, Jsonb({"raw_artifact_id": raw_id})),
            )
            conn.execute(
                "INSERT INTO acquisition_job"
                " (id, source_id, status, provider, artifact_id, diagnostics, finished_at)"
                " VALUES (%s, %s, 'succeeded', 'manual-upload/v1', %s, %s, now())",
                (
                    job_id,
                    source_id,
                    enriched_id,
                    Jsonb(
                        {
                            "input_mode": "pdf",
                            "pipeline_requires_cleanup": True,
                            "visual_incomplete": True,
                        }
                    ),
                ),
            )
            conn.commit()

        attention = client.get(syllabus_url).json()["lessons"][0]["sources"][0]
        assert attention["pipeline"]["status"] == "attention"
        assert attention["has_markdown"] is False
        blocked = client.get(f"/api/sources/{source_id}/markdown")
        assert blocked.status_code == 409
        assert "atenção" in blocked.json()["detail"].lower()

        cleanup_job_id = f"cleanup-pdf-clean-{marker}"
        with psycopg.connect(test_database_url) as conn:
            conn.execute(
                "UPDATE acquisition_job SET diagnostics = %s WHERE id = %s",
                (
                    Jsonb(
                        {
                            "input_mode": "pdf",
                            "pipeline_requires_cleanup": True,
                            "visual_incomplete": False,
                        }
                    ),
                    job_id,
                ),
            )
            conn.execute(
                "INSERT INTO source_cleanup_job"
                " (id, acquisition_job_id, source_id, source_artifact_id)"
                " VALUES (%s, %s, %s, %s)",
                (cleanup_job_id, job_id, source_id, enriched_id),
            )
            conn.commit()
        cleaning = client.get(syllabus_url).json()["lessons"][0]["sources"][0]
        assert cleaning["pipeline"]["status"] == "cleaning"
        assert cleaning["has_markdown"] is False

        run_id = f"web-pdf-cut-{marker}"
        cleanup_id = f"pc-pdf-clean-{marker}"
        clean_id = f"{enriched_id}:clean:{cleanup_id}"
        with psycopg.connect(test_database_url) as conn:
            conn.execute(
                "INSERT INTO run"
                " (id, stage, model, prompt_ref, prompt_sha, status, finished_at)"
                " VALUES (%s, 'passage-cuts', 'fake/model', 'passage-cuts/v001',"
                " 'abc', 'done', now())",
                (run_id,),
            )
            conn.execute(
                "INSERT INTO passage_cleanup"
                " (id, cuts_run_id, model, triage_prompt_ref, refine_prompt_ref,"
                " status, finished_at)"
                " VALUES (%s, %s, 'fake/model', 'passage-triage/v003',"
                " 'passage-refine/v002', 'done', now())",
                (cleanup_id, run_id),
            )
            conn.execute(
                "INSERT INTO artifact"
                " (id, snapshot_id, kind, tool, tool_version, body, metadata)"
                " VALUES (%s, %s, 'markdown', 'passage-cleanup', 'v1', %s, %s)",
                (
                    clean_id,
                    snapshot_id,
                    "# Clean PDF\n",
                    Jsonb(
                        {
                            "source_markdown_artifact_id": enriched_id,
                            "cleanup_id": cleanup_id,
                        }
                    ),
                ),
            )
            conn.execute(
                "UPDATE source_cleanup_job SET status = 'succeeded', cleanup_id = %s,"
                " canonical_artifact_id = %s, finished_at = now() WHERE id = %s",
                (cleanup_id, clean_id, cleanup_job_id),
            )
            conn.commit()

        ready = client.get(syllabus_url).json()["lessons"][0]["sources"][0]
        assert ready["pipeline"]["status"] == "ready"
        assert ready["markdown"]["artifact_id"] == clean_id
        published = client.get(f"/api/sources/{source_id}/markdown")
        assert published.status_code == 200
        assert published.json()["artifact_id"] == clean_id
        assert published.json()["markdown"] == "# Clean PDF\n"


def test_manual_upload_queues_ordered_external_assets_and_serves_them_safely(
    test_database_url, applied_migrations, tmp_path
):
    path = _workbook(
        tmp_path / "manual-web.xlsx",
        project="Project",
        lesson="Aula",
        source_url=f"https://example.com/manual-{uuid.uuid4().hex}",
    )
    store = LocalAssetStore(tmp_path / "source-assets")

    with TestClient(_app(test_database_url, store)) as client:
        uploaded = _upload(
            client, path, f"Manual web {uuid.uuid4().hex[:8]}"
        ).json()
        detail = client.get(f"/api/syllabi/{uploaded['syllabus_id']}").json()
        source_id = detail["lessons"][0]["sources"][0]["source_id"]
        first = _png_bytes(20, 10)
        second = _png_bytes(30, 15)

        queued = client.post(
            f"/api/sources/{source_id}/manual-upload",
            data={"kind": "images"},
            files=[
                ("files", ("page-2.png", first, "image/png")),
                ("files", ("page-1.png", second, "image/png")),
            ],
        )
        assert queued.status_code == 202, queued.text
        assert queued.json()["job"]["provider"] == "manual-upload/v1"
        assert queued.json()["job"]["status"] == "queued"

        duplicate = client.post(
            f"/api/sources/{source_id}/manual-upload",
            data={"kind": "pdf"},
            files={"files": ("source.pdf", b"%PDF-1.7\nfixture", "application/pdf")},
        )
        assert duplicate.status_code == 409

        with psycopg.connect(test_database_url) as conn:
            rows = conn.execute(
                "SELECT id, ordinal, filename, storage_key FROM source_asset"
                " WHERE acquisition_job_id = %s ORDER BY ordinal",
                (queued.json()["job"]["id"],),
            ).fetchall()
            columns = {
                row[0]
                for row in conn.execute(
                    "SELECT column_name FROM information_schema.columns"
                    " WHERE table_name = 'source_asset'"
                )
            }
        assert [(row[1], row[2]) for row in rows] == [
            (1, "page-2.png"),
            (2, "page-1.png"),
        ]
        assert "body" not in columns
        assert store.get(rows[0][3]) == first
        assert store.get(rows[1][3]) == second

        asset = client.get(f"/api/source-assets/{rows[0][0]}")
        assert asset.status_code == 200
        assert asset.content == first
        assert asset.headers["content-type"].startswith("image/png")
        assert asset.headers["x-content-type-options"] == "nosniff"
        assert asset.headers["cross-origin-resource-policy"] == "same-origin"
        assert "sandbox" in asset.headers["content-security-policy"]
        assert "immutable" in asset.headers["cache-control"]
        assert asset.headers["etag"]

        class BackendClientError(Exception):
            pass

        class UnavailableStore:
            def get(self, _key):
                raise BackendClientError("provider detail must not leak")

        with TestClient(_app(test_database_url, UnavailableStore())) as unavailable:
            failed = unavailable.get(f"/api/source-assets/{rows[0][0]}")
        assert failed.status_code == 503
        assert failed.json() == {
            "detail": "O arquivo está temporariamente indisponível."
        }
        assert "provider detail" not in failed.text


def test_manual_upload_rejects_mixed_media_before_creating_a_job(
    test_database_url, applied_migrations, tmp_path
):
    path = _workbook(
        tmp_path / "manual-invalid.xlsx",
        project="Project",
        lesson="Aula",
        source_url=f"https://example.com/manual-invalid-{uuid.uuid4().hex}",
    )
    store = LocalAssetStore(tmp_path / "invalid-assets")
    with TestClient(_app(test_database_url, store)) as client:
        uploaded = _upload(
            client, path, f"Manual invalid {uuid.uuid4().hex[:8]}"
        ).json()
        detail = client.get(f"/api/syllabi/{uploaded['syllabus_id']}").json()
        source_id = detail["lessons"][0]["sources"][0]["source_id"]
        rejected = client.post(
            f"/api/sources/{source_id}/manual-upload",
            data={"kind": "images"},
            files={"files": ("source.pdf", b"%PDF-1.7\nfixture", "application/pdf")},
        )
        assert rejected.status_code == 415
        assert "não misture" in rejected.json()["detail"]

    with psycopg.connect(test_database_url) as conn:
        assert conn.execute(
            "SELECT count(*) FROM acquisition_job WHERE source_id = %s",
            (source_id,),
        ).fetchone()[0] == 0


def test_downloaded_image_branch_remains_active_until_batch_analysis():
    assert _summarize_image_counts({"downloaded": 2}) == {
        "state": "processing",
        "total": 2,
        "queued": 0,
        "running": 0,
        "downloaded": 2,
        "useful": 0,
        "not_important": 0,
        "filtered": 0,
        "failed": 0,
        "active": True,
        "attention": 0,
    }


def test_image_branch_is_a_nonblocking_sidecar_of_the_latest_markdown(
    test_database_url, applied_migrations, tmp_path
):
    path = _workbook(
        tmp_path / "image-sidecar.xlsx",
        project="Project",
        lesson="Aula",
        source_url=f"https://example.com/image-sidecar-{uuid.uuid4().hex}",
    )
    store = LocalAssetStore(tmp_path / "article-assets")
    marker = uuid.uuid4().hex[:10]
    with TestClient(_app(test_database_url, store)) as client:
        uploaded = _upload(
            client, path, f"Image branch {uuid.uuid4().hex[:8]}"
        ).json()
        syllabus_url = f"/api/syllabi/{uploaded['syllabus_id']}"
        source_id = client.get(syllabus_url).json()["lessons"][0]["sources"][0]["source_id"]
        snapshot_id = f"snap-{marker}"
        artifact_id = f"art-{marker}"
        enriched_id = f"{artifact_id}:images"
        job_id = f"acq-{marker}"
        asset_id = f"asset-{marker}"
        analysis_id = f"analysis-{marker}"
        candidate_id = f"candidate-{marker}-1"
        image_body = _png_bytes(64, 32)
        stored = store.put(image_body)
        with psycopg.connect(test_database_url) as conn:
            conn.execute(
                "INSERT INTO source_snapshot (id, source_id, content_hash, status)"
                " VALUES (%s, %s, %s, 'ok')",
                (snapshot_id, source_id, f"hash-{marker}"),
            )
            conn.execute(
                "INSERT INTO artifact (id, snapshot_id, kind, tool, body)"
                " VALUES (%s, %s, 'markdown', 'firecrawl-v2', '# Article')",
                (artifact_id, snapshot_id),
            )
            conn.execute(
                "INSERT INTO acquisition_job"
                " (id, source_id, status, provider, artifact_id, finished_at)"
                " VALUES (%s, %s, 'succeeded', 'firecrawl', %s, now())",
                (job_id, source_id, artifact_id),
            )
            conn.execute(
                "INSERT INTO source_asset"
                " (id, acquisition_job_id, source_id, ordinal, kind, filename,"
                "  mime_type, sha256, byte_size, storage_key, metadata, original_url)"
                " VALUES (%s, %s, %s, 1, 'article_image', 'diagram.png',"
                "  'image/png', %s, %s, %s, '{}', 'https://cdn.example/diagram.png')",
                (asset_id, job_id, source_id, stored.sha256, len(image_body), stored.key),
            )
            conn.execute(
                "INSERT INTO source_asset_analysis"
                " (id, source_asset_id, purpose, status, prompt_version, result)"
                " VALUES (%s, %s, 'article_image_relevance', 'succeeded', 'test.v1', %s)",
                (
                    analysis_id,
                    asset_id,
                    Jsonb(
                        {
                            "image_id": candidate_id,
                            "retain": True,
                            "reason_code": "information",
                            "ocr": "Estratégia → Execução",
                            "description": "Diagrama que relaciona estratégia e execução.",
                            "limitations": None,
                        }
                    ),
                ),
            )
            conn.execute(
                "INSERT INTO source_image_candidate"
                " (id, acquisition_job_id, source_id, snapshot_id, markdown_artifact_id,"
                "  ordinal, original_url, alt_text, status, asset_id, analysis_id, finished_at)"
                " VALUES (%s, %s, %s, %s, %s, 1, 'https://cdn.example/diagram.png',"
                "  'Diagrama', 'useful', %s, %s, now())",
                (
                    candidate_id,
                    job_id,
                    source_id,
                    snapshot_id,
                    artifact_id,
                    asset_id,
                    analysis_id,
                ),
            )
            conn.execute(
                "INSERT INTO source_image_candidate"
                " (id, acquisition_job_id, source_id, snapshot_id, markdown_artifact_id,"
                "  ordinal, original_url, alt_text, status, failure_code, diagnostics, finished_at)"
                " VALUES (%s, %s, %s, %s, %s, 2, 'https://cdn.example/missing.png',"
                "  'Foto ausente', 'failed', 'image_not_found', %s, now())",
                (
                    f"candidate-{marker}-2",
                    job_id,
                    source_id,
                    snapshot_id,
                    artifact_id,
                    Jsonb({"category": "not_found", "http_status": 404}),
                ),
            )
            conn.execute(
                "INSERT INTO artifact"
                " (id, snapshot_id, kind, tool, tool_version, body, metadata, created_at)"
                " VALUES (%s, %s, 'markdown', 'article-image-association', 'test.v1',"
                "  %s, %s, now() + interval '1 second')",
                (
                    enriched_id,
                    snapshot_id,
                    f"# Article\n\n![Diagrama](/api/source-assets/{asset_id})\n",
                    Jsonb({"source_markdown_artifact_id": artifact_id}),
                ),
            )
            conn.commit()

        source = client.get(syllabus_url).json()["lessons"][0]["sources"][0]
        assert source["has_markdown"] is True
        assert source["markdown"]["artifact_id"] == enriched_id
        assert source["job"]["status"] == "succeeded"
        assert source["image_branch"] == {
            "state": "attention",
            "total": 2,
            "queued": 0,
            "running": 0,
            "downloaded": 0,
            "useful": 1,
            "not_important": 0,
            "filtered": 0,
            "failed": 1,
            "active": False,
            "attention": 1,
        }

        markdown = client.get(f"/api/sources/{source_id}/markdown")
        assert markdown.status_code == 200
        body = markdown.json()
        assert body["artifact_id"] == enriched_id
        assert f'<img src="/api/source-assets/{asset_id}"' in body["html"]
        assert body["image_branch"]["attention"] == 1
        assert body["images"][0]["asset_url"] == f"/api/source-assets/{asset_id}"
        assert body["images"][0]["analysis"]["description"].startswith("Diagrama")
        assert body["images"][0]["analysis"]["ocr"] == "Estratégia → Execução"
        assert body["images"][1]["error"] == "A imagem retornou 404 (não encontrada)."
        assert client.get(body["images"][0]["asset_url"]).content == image_body


def test_late_image_enrichment_of_an_old_snapshot_never_replaces_new_markdown(
    test_database_url, applied_migrations, tmp_path
):
    marker = uuid.uuid4().hex[:10]
    path = _workbook(
        tmp_path / "late-enrichment.xlsx",
        project="Project",
        lesson="Aula",
        source_url=f"https://example.com/late-enrichment-{marker}",
    )
    with TestClient(_app(test_database_url)) as client:
        uploaded = _upload(
            client, path, f"Late enrichment {uuid.uuid4().hex[:8]}"
        ).json()
        syllabus_url = f"/api/syllabi/{uploaded['syllabus_id']}"
        source_id = client.get(syllabus_url).json()["lessons"][0]["sources"][0]["source_id"]
        old_snapshot = f"old-snapshot-{marker}"
        old_artifact = f"old-artifact-{marker}"
        old_enriched = f"{old_artifact}:images"
        new_snapshot = f"new-snapshot-{marker}"
        new_artifact = f"new-artifact-{marker}"
        with psycopg.connect(test_database_url) as conn:
            conn.execute(
                "INSERT INTO source_snapshot"
                " (id, source_id, content_hash, status, created_at)"
                " VALUES (%s, %s, 'old-hash', 'ok', now()),"
                "        (%s, %s, 'new-hash', 'ok', now() + interval '1 second')",
                (old_snapshot, source_id, new_snapshot, source_id),
            )
            conn.execute(
                "INSERT INTO artifact"
                " (id, snapshot_id, kind, tool, body, metadata, created_at)"
                " VALUES (%s, %s, 'markdown', 'firecrawl', '# Old', '{}', now()),"
                "        (%s, %s, 'markdown', 'article-image-association',"
                "         '# Old enriched', %s, now() + interval '10 seconds'),"
                "        (%s, %s, 'markdown', 'firecrawl', '# New', '{}',"
                "         now() + interval '1 second')",
                (
                    old_artifact,
                    old_snapshot,
                    old_enriched,
                    old_snapshot,
                    Jsonb({"source_markdown_artifact_id": old_artifact}),
                    new_artifact,
                    new_snapshot,
                ),
            )
            conn.commit()

        source = client.get(syllabus_url).json()["lessons"][0]["sources"][0]
        assert source["markdown"]["artifact_id"] == new_artifact
        markdown = client.get(f"/api/sources/{source_id}/markdown").json()
        assert markdown["artifact_id"] == new_artifact
        assert markdown["markdown"] == "# New"



def test_provider_auth_failure_is_not_described_as_target_site_refusal():
    provider = _diagnostic_message(
        {"category": "access_denied", "http_status": 403},
        "http_status_4xx",
    )
    target = _diagnostic_message(
        {"category": "access_denied", "target_http_status": 403},
        "http_status_4xx",
    )
    explicit_provider = _diagnostic_message(
        {
            "category": "provider_authentication",
            "provider_http_status": 401,
            "provider_code": "AUTH_INVALID",
            "request_id": "fc-request-123",
            "provider_job_id": "scrape-456",
        },
        "http_status_4xx",
    )

    assert "Firecrawl" in provider
    assert "site alvo" not in provider.lower()
    assert "site alvo" in target.lower()
    assert "Firecrawl" in explicit_provider
    assert "AUTH_INVALID" in explicit_provider
    assert "fc-request-123" in explicit_provider
    assert "scrape-456" in explicit_provider


def test_image_provider_routing_failure_is_actionable_and_keeps_http_status():
    message = _image_failure_message(
        {"category": "model_routing_unavailable", "provider_http_status": 404},
        "image_analysis_failed",
    )

    assert "OpenRouter" in message
    assert "provider compatível" in message
    assert "404" in message


def test_background_worker_survives_an_arbitrary_backend_factory_error(
    monkeypatch
):
    class BackendClientError(Exception):
        pass

    calls = []
    stop = asyncio.Event()

    def iteration(_connect_factory, _asset_store_factory):
        calls.append(len(calls) + 1)
        if len(calls) == 1:
            raise BackendClientError("temporary object-store outage")
        stop.set()
        return False

    monkeypatch.setattr(web_app, "_work_one", iteration)
    monkeypatch.setattr(web_app, "acquisition_poll_seconds", lambda: 0)

    asyncio.run(web_app._worker_loop(lambda: None, lambda: None, stop))

    assert calls == [1, 2]


def test_six_sheet_workbook_exposes_five_curricular_subjects_and_sources(
    test_database_url, applied_migrations, tmp_path
):
    name = f"Pilot syllabus {uuid.uuid4().hex[:8]}"
    workbook_path = _five_subject_workbook(tmp_path / "observer-export.xlsx")
    workbook = load_workbook(workbook_path, read_only=True)
    assert len(workbook.sheetnames) == 6
    workbook.close()

    with TestClient(_app(test_database_url)) as client:
        uploaded = _upload(client, workbook_path, name)
        assert uploaded.status_code == 201, uploaded.text
        result = uploaded.json()
        detail = client.get(f"/api/syllabi/{result['syllabus_id']}").json()

    assert [subject["code"] for subject in detail["lesson_subjects"]] == [
        "COM",
        "LID",
        "MTF",
        "NEG",
        "UEX",
    ]
    assert {lesson["kind"] for lesson in detail["lessons"]} == {"Class"}
    assert len(detail["lessons"]) == 5
    assert all(len(lesson["sources"]) == 1 for lesson in detail["lessons"])
    assert all(
        not source["has_markdown"]
        for lesson in detail["lessons"]
        for source in lesson["sources"]
    )
    assert detail["metadata_complete"] is True
    assert {
        identity["lesson_subject_code"] for identity in detail["export_identities"]
    } == {"COM", "MTF", "NEG", "UEX", "LID"}

    with psycopg.connect(test_database_url) as conn:
        stored = conn.execute(
            "SELECT lesson_subject_code, graph_id FROM syllabus_subject"
            " WHERE syllabus_id = %s ORDER BY lesson_subject_code",
            (result["syllabus_id"],),
        ).fetchall()
    assert stored == [
        (code, subject_graph_id_for("web-inteli", name, code))
        for code in ("COM", "LID", "MTF", "NEG", "UEX")
    ]


def test_operator_can_start_and_read_a_selected_lesson_build(
    test_database_url, tmp_path
):
    path = _five_subject_workbook(tmp_path / "lesson-build.xlsx")
    with TestClient(
        _app(
            test_database_url,
            companion_repo=accepting_companion(
                tmp_path, require_replacement=False
            ),
        )
    ) as client:
        uploaded = _upload(client, path, f"Build {uuid.uuid4().hex[:8]}").json()
        detail = client.get(f"/api/syllabi/{uploaded['syllabus_id']}").json()
        lesson = detail["lessons"][0]
        source = lesson["sources"][0]
        with psycopg.connect(test_database_url) as conn:
            _publish_source(
                conn,
                source,
                f"build-{source['reference_id']}",
                content_hash="c" * 64,
            )
        reviewed = client.patch(
            f"/api/syllabi/{uploaded['syllabus_id']}/sources/{source['reference_id']}/review",
            json={"validated": True, "complexity": "simple"},
        )
        assert reviewed.status_code == 200
        offer = client.get(
            f"/api/syllabi/{uploaded['syllabus_id']}/versions/"
            f"{detail['version']['id']}/lessons/{lesson['id']}/lesson-build"
        )
        assert offer.status_code == 200
        assert offer.json()["references"] == [
            {
                "reference_id": source["reference_id"],
                "title": source["title"],
                "eligible": True,
                "selected": True,
            }
        ]
        start_url = (
            f"/api/syllabi/{uploaded['syllabus_id']}/versions/"
            f"{detail['version']['id']}/lessons/{lesson['id']}/lesson-builds"
        )
        omitted = client.post(
            start_url,
            json={"request_key": f"browser-omitted-{uuid.uuid4().hex}"},
        )
        assert omitted.status_code == 422
        assert omitted.json()["detail"]["code"] == "no_selected_references"
        started = client.post(
            start_url,
            json={
                "request_key": f"browser-{uuid.uuid4().hex}",
                "reference_ids": [source["reference_id"]],
            },
        )
        assert started.status_code == 201, started.text
        fetched = client.get(f"/api/lesson-builds/{started.json()['id']}")

        build_id = started.json()["id"]
        graph_id = lesson["lesson_subject"]["graph_id"]
        concept_id = f"concept-{lesson['id']}-accepted"
        fragment = {
            "artifact_type": "runtime_graph",
            "schema_version": "runtime_graph.v0",
            "generated_at": "2026-09-02T12:00:00+00:00",
            "subject": {
                "course_id": uploaded["syllabus_id"],
                "module_id": detail["version"]["id"],
                "pipeline_subject_id": lesson["subject"],
                "title": lesson["lesson_subject"]["display_name"],
                "language": "pt-BR",
                "professors": [],
            },
            "concepts": [
                {
                    "concept_id": concept_id,
                    "display_code": "COM-001",
                    "label": "Conceito aceito",
                    "knowledge_type": "conceptual",
                    "description": "Conteúdo revisado.",
                    "coverage_criteria": ["Explicar o conteúdo revisado."],
                    "common_misconceptions": [],
                    "dependencies": {"blocking": [], "hard": [], "soft": []},
                }
            ],
            "lessons": [
                {
                    "lesson_id": lesson["id"],
                    "display_code": lesson["id"],
                    "title": lesson["title"],
                    "description": "",
                    "segments": [
                        {
                            "segment_id": f"segment-{lesson['id']}-accepted",
                            "display_code": "L01-S01",
                            "label": "Conceito aceito",
                            "instructional_role": "teach",
                            "concept_ids": [concept_id],
                            "teaching_notes": "",
                            "self_study_resource_ids": [],
                            "self_study_resource_refs": [],
                        }
                    ],
                }
            ],
            "self_study_resources": [],
        }
        body = json.dumps(fragment, ensure_ascii=False, sort_keys=True, indent=2) + "\n"
        with psycopg.connect(test_database_url) as conn:
            conn.execute(
                "UPDATE lesson_build SET status = 'succeeded', is_active = false,"
                " finished_at = now() WHERE id = %s",
                (build_id,),
            )
            conn.execute(
                "UPDATE lesson_build_work SET status = 'succeeded', stage = NULL"
                " WHERE build_id = %s",
                (build_id,),
            )
            conn.execute(
                "INSERT INTO lesson_build_checkpoint"
                " (id, build_id, stage, family, path, body, content_sha256,"
                " stage_fingerprint, is_stage_result)"
                " VALUES (%s, %s, 'lesson-fragment', 'lesson_fragment',"
                " 'final_graph/runtime_graph.json', %s, %s, %s, true)",
                (
                    f"checkpoint-{build_id}",
                    build_id,
                    body,
                    hashlib.sha256(body.encode()).hexdigest(),
                    hashlib.sha256(f"fingerprint-{build_id}".encode()).hexdigest(),
                ),
            )
        accepted = client.post(
            f"/api/lesson-builds/{build_id}/accept",
            json={"actor": "founder"},
        )
        reviewed_build = client.get(f"/api/lesson-builds/{build_id}")
        graph_history = client.get(f"/api/graphs/{graph_id}")
        current_graph = client.get(f"/api/graphs/{graph_id}/graph.json")
        downloaded_graph = client.get(
            f"/api/graphs/{graph_id}/graph.json?download=true"
        )
        historical_graph = client.get(
            "/api/graph-revisions/"
            f"{accepted.json()['revision']['id']}/graph.json?download=true"
        )
        current_package = client.get(
            f"/api/graphs/{graph_id}/companion-package.zip"
        )
        selected_package = client.get(
            "/api/graph-revisions/"
            f"{accepted.json()['revision']['id']}/companion-package.zip"
        )

    assert fetched.status_code == 200
    assert fetched.json()["manifest"]["references"][0]["reference_id"] == source[
        "reference_id"
    ]
    assert accepted.status_code == 200, accepted.text
    assert accepted.json()["revision"]["number"] == 1
    assert reviewed_build.json()["review"]["decision"] == "accepted"
    assert reviewed_build.json()["graph_revision"]["id"] == accepted.json()[
        "revision"
    ]["id"]
    assert graph_history.json()["current_revision"]["number"] == 1
    assert current_graph.json()["graph_id"] == graph_id
    assert current_graph.json()["concepts"][0]["concept_id"] == concept_id
    assert "Content-Disposition" not in current_graph.headers
    assert "attachment;" in downloaded_graph.headers["Content-Disposition"]
    assert historical_graph.content == downloaded_graph.content
    assert current_package.status_code == 200, current_package.text
    assert current_package.headers["content-type"] == "application/zip"
    assert selected_package.status_code == 200, selected_package.text
    with ZipFile(BytesIO(selected_package.content)) as archive:
        assert set(archive.namelist()) == {
            f"{graph_id}/graph.json",
            f"{graph_id}/intro_notes.json",
        }


def test_companion_rejection_blocks_package_but_keeps_raw_graph_download(
    test_database_url, tmp_path
):
    with psycopg.connect(test_database_url) as conn:
        syllabus_id, version_id, graph_id = _seed_subject(conn, "web-package-blocked")
        conn.execute(
            "INSERT INTO institution (id, name) VALUES ('web-inteli', 'Inteli Web')"
            " ON CONFLICT (id) DO NOTHING"
        )
        conn.execute(
            "UPDATE syllabus SET institution_id = 'web-inteli' WHERE id = %s",
            (syllabus_id,),
        )
        conn.commit()
        build_id, _ = _seed_finished_build(
            conn,
            version_id=version_id,
            graph_id=graph_id,
            lesson_id="lesson-package-blocked",
            build_label="v1",
            lesson_seq=1,
        )
        graph_revision.accept(conn, build_id, actor="founder")

    with TestClient(
        _app(
            test_database_url,
            companion_repo=_rejecting_companion(
                tmp_path, "malformed_runtime_graph"
            ),
        )
    ) as client:
        raw_before = client.get(f"/api/graphs/{graph_id}/graph.json?download=true")
        package = client.get(f"/api/graphs/{graph_id}/companion-package.zip")
        raw_after = client.get(f"/api/graphs/{graph_id}/graph.json?download=true")

    assert raw_before.status_code == 200
    assert package.status_code == 422
    assert package.json()["detail"] == {
        "code": "companion_package_blocked",
        "message": (
            "Companion rejected the package (malformed_runtime_graph): graph.json "
            "does not satisfy Companion's runtime_graph.v0 contract. Inspect the raw "
            "Graph Revision and regenerate the affected Lesson Build."
        ),
    }
    assert raw_after.content == raw_before.content
