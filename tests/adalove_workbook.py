"""Small full-fidelity Adalove workbook factory for tests."""

import hashlib
from pathlib import Path

from openpyxl import Workbook

from universe.syllabus import (
    ADALOVE_ACTIVITY_COLUMNS,
    ADALOVE_MATERIAL_COLUMNS,
    ADALOVE_ORDER_AUDIT_COLUMNS,
    ADALOVE_SUBJECT_COLUMNS,
    HIDDEN_COLUMN,
)


def stable_uuid(kind: str, *parts: object) -> str:
    encoded = "\0".join(str(part or "") for part in parts).encode()
    return hashlib.sha256(f"{kind}\0".encode() + encoded).hexdigest()[:32]


def activity(
    *,
    title: str = "Lesson",
    kind: str = "Class",
    week: int = 1,
    order: int = 1,
    activity_uuid: str | None = None,
    folder_uuid: str | None = None,
    parent_uuid: str | None = None,
    parent_title: str | None = None,
    parent_inference: str | None = None,
    subject: str | None = "COM",
    subjects: list[str] | tuple[str, ...] = (),
    description: str | None = "Description",
    date: str | None = "06/08/2026",
    url: str | None = None,
    resource_code: str | None = None,
    hidden: bool = False,
    materials: list[dict] | None = None,
    **extra,
) -> dict:
    activity_uuid = activity_uuid or stable_uuid("activity", week, order)
    folder_uuid = folder_uuid or stable_uuid("folder", week)
    if kind == "Self-study" and parent_uuid and parent_inference is None:
        parent_inference = "inferred_from_activity_order"
    row = {
        "Activity order": order,
        "Week": week,
        "Type": kind,
        "Original label": kind,
        "Title": title,
        "Description": description,
        "Date": date,
        "Lesson Subject code": subject,
        "Related subjects": "; ".join(subjects),
        "Primary URL": url,
        "Resource code": resource_code,
        "Parent activity UUID": parent_uuid,
        "Parent title": parent_title,
        "Parent inference": parent_inference,
        "Activity UUID": activity_uuid,
        "Folder UUID": folder_uuid,
        "Section UUID": stable_uuid("section", "test"),
        "Active": "Sim",
        "Detail error": None,
        HIDDEN_COLUMN: "yes" if hidden else "no",
        "_subjects": list(subjects),
        "_materials": materials,
    }
    row.update(extra)
    return row


def material(
    *,
    url: str,
    label: str | None = None,
    source: str = "basic_activity_url",
    source_path: str = "basic_activity_url",
    resource_code: str | None = None,
    video: bool = False,
) -> dict:
    return {
        "Label": label,
        "URL": url,
        "Source": source,
        "Source path": source_path,
        "Resource code": resource_code,
        "Video": "Sim" if video else "Não",
    }


def write_adalove_workbook(
    path: Path,
    activities: list[dict],
    *,
    project: str = "TEST PROJECT",
) -> Path:
    workbook = Workbook()
    activity_sheet = workbook.active
    activity_sheet.title = "Activities"
    activity_sheet.append((*ADALOVE_ACTIVITY_COLUMNS, HIDDEN_COLUMN))
    subject_sheet = workbook.create_sheet("Subjects")
    subject_sheet.append(ADALOVE_SUBJECT_COLUMNS)
    material_sheet = workbook.create_sheet("Materials")
    material_sheet.append(ADALOVE_MATERIAL_COLUMNS)
    audit_sheet = workbook.create_sheet("Order audit")
    audit_sheet.append(ADALOVE_ORDER_AUDIT_COLUMNS)

    for raw in activities:
        row = dict(raw)
        related_subjects = list(row.pop("_subjects", []))
        related_materials = row.pop("_materials", None)
        activity_sheet.append(
            [row.get(column) for column in (*ADALOVE_ACTIVITY_COLUMNS, HIDDEN_COLUMN)]
        )
        for index, subject in enumerate(related_subjects, 1):
            subject_row = {
                "Activity order": row["Activity order"],
                "Week": row["Week"],
                "Activity UUID": row["Activity UUID"],
                "Activity title": row["Title"],
                "Lesson Subject code": row.get("Lesson Subject code"),
                "Subject UUID": stable_uuid("subject", row["Activity UUID"], index),
                "Related subject": subject,
            }
            subject_sheet.append(
                [subject_row.get(column) for column in ADALOVE_SUBJECT_COLUMNS]
            )
        if related_materials is None and row["Type"] == "Self-study" and row.get("Primary URL"):
            related_materials = [
                material(
                    url=row["Primary URL"],
                    label=row["Title"],
                    resource_code=row.get("Resource code"),
                    video="youtube" in row["Primary URL"] or "youtu.be" in row["Primary URL"],
                )
            ]
        for raw_material in related_materials or []:
            material_row = {
                "Activity order": row["Activity order"],
                "Week": row["Week"],
                "Activity UUID": row["Activity UUID"],
                "Activity title": row["Title"],
                **raw_material,
            }
            material_sheet.append(
                [material_row.get(column) for column in ADALOVE_MATERIAL_COLUMNS]
            )
        audit_row = {
            "Activity order": row["Activity order"],
            "Week": row["Week"],
            "Order key": f"{row['Week']}:{row['Activity order']}",
            "Duplicate order key": "no",
            "Missing orders in week": None,
            "Activity UUID": row["Activity UUID"],
            "Folder UUID": row["Folder UUID"],
            "Type": row["Type"],
            "Title": row["Title"],
            "Parent inference": row.get("Parent inference"),
            "Detail error": row.get("Detail error"),
        }
        audit_sheet.append(
            [audit_row.get(column) for column in ADALOVE_ORDER_AUDIT_COLUMNS]
        )

    read_me = workbook.create_sheet("Read me")
    read_me.append(("Field", "Value / note"))
    read_me.append(("Project", project))
    errors = workbook.create_sheet("Errors")
    errors.append(("Activity UUID", "Error"))
    workbook.save(path)
    workbook.close()
    return path
