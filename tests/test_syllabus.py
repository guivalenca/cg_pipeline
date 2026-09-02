"""Syllabus facts: Adalove intake, immutable versions and source identity."""

import hashlib
import re
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from universe.graph_identity import GraphIdConflict, subject_graph_id_for
from universe.syllabus import (
    LESSON_SUBJECTS,
    _lesson_subjects_by_syllabus,
    SyllabusVersionConflict,
    XLSX_MIME,
    build_parser,
    canonical_url,
    curate_syllabus,
    get_syllabus_history,
    get_syllabus_version,
    get_syllabus_workbook,
    import_workbook,
    list_syllabi,
    parse_workbook,
    source_identity,
)
from adalove_workbook import activity, material, stable_uuid, write_adalove_workbook

SUBJECT_CODES = {
    "Computação": "COM",
    "User Experience": "UEX",
    "Liderança": "LID",
    "Negócios": "NEG",
    "Matemática": "MTF",
}
KINDS = {
    "Encontro de instrução": "Class",
    "Encontro": "Class",
    "Autoestudo": "Self-study",
    "Encontro de orientação": "Orientation",
    "Desenvolvimento projeto": "Deliverable",
    "Avaliação / pesquisa": "Evaluation",
}


def syllabus_row(
    *,
    project="INTERNAL WORKBOOK TITLE",
    week=1,
    seq=1,
    title="Lesson",
    kind="Class",
    description="Description",
    url=None,
    parent=None,
    parent_order=1,
    subject="COM",
    subjects=None,
    code=None,
    date="06/08/2026",
    hidden=False,
    materials=None,
):
    week = int(str(week).replace("Semana", "").strip())
    seq = int(seq)
    canonical_kind = KINDS.get(kind, kind)
    canonical_subject = SUBJECT_CODES.get(subject, subject)
    topic_values = []
    if subjects:
        topic_values = [
            line.removeprefix(",").strip()
            for line in str(subjects).splitlines()
            if line.removeprefix(",").strip()
        ]
    parent_uuid = (
        stable_uuid("activity", week, parent_order)
        if canonical_kind == "Self-study" and parent
        else None
    )
    row = activity(
        title=title,
        kind=canonical_kind,
        week=week,
        order=seq,
        parent_uuid=parent_uuid,
        parent_title=parent,
        subject=canonical_subject,
        subjects=topic_values,
        description=description,
        date=date,
        url=url,
        resource_code=code,
        hidden=hidden,
        materials=materials,
    )
    row["_project"] = project
    return row


def write_syllabus(path, rows):
    project = next((row.get("_project") for row in rows if row.get("_project")), "TEST")
    clean_rows = [{key: value for key, value in row.items() if key != "_project"} for row in rows]
    return write_adalove_workbook(path, clean_rows, project=project)


class TestWorkbookAdapters:
    def test_lesson_subject_catalog_has_the_shared_five_subjects(self):
        assert [
            (subject.code, subject.accepted_spellings, subject.display_name)
            for subject in LESSON_SUBJECTS
        ] == [
            ("COM", ("COM", "Computação"), "Computação"),
            ("LID", ("LID", "Liderança"), "Liderança"),
            ("NEG", ("NEG", "Negócios"), "Negócios"),
            ("UEX", ("UEX", "User Experience"), "User Experience"),
            ("MTF", ("MTF", "Matemática", "Matemática e Física"), "Matemática"),
        ]

    def test_every_accepted_eixo_spelling_maps_to_its_subject_code(self, tmp_path):
        path = tmp_path / "spellings.xlsx"
        spellings = [
            ("Computação", "COM"),
            ("User Experience", "UEX"),
            ("Liderança", "LID"),
            ("Negócios", "NEG"),
            ("Matemática", "MTF"),
            ("Matemática e Física", "MTF"),
            ("uex", "UEX"),
        ]
        write_syllabus(
            path,
            [
                syllabus_row(title=f"Aula {seq}", seq=seq, subject=spelling)
                for seq, (spelling, _) in enumerate(spellings, 1)
            ],
        )

        parsed = parse_workbook(path)

        assert [lesson["subject"] for lesson in parsed["lessons"]] == [
            code for _, code in spellings
        ]

    def test_full_fidelity_sheets_become_lessons_and_material_sources(self, tmp_path):
        path = tmp_path / "adalove.xlsx"
        write_syllabus(
            path,
            [
                syllabus_row(
                    title="Aula de arquitetura",
                    subjects="Arquitetura de nuvem\n,Modelos de serviço em nuvem",
                ),
                syllabus_row(
                    seq=2,
                    title="Leitura de arquitetura",
                    kind="Self-study",
                    parent="Aula de arquitetura",
                    url="https://example.com/architecture",
                    materials=[
                        material(url="https://example.com/architecture", label="Artigo"),
                        material(url="https://youtu.be/h4gw6gCP5ls", label="Vídeo", video=True),
                    ],
                ),
            ],
        )

        parsed = parse_workbook(path)

        assert parsed["lesson_count"] == 1
        lesson = parsed["lessons"][0]
        assert lesson["subjects"] == [
            "Arquitetura de nuvem",
            "Modelos de serviço em nuvem",
        ]
        assert [source["url"] for source in lesson["source_references"]] == [
            "https://example.com/architecture",
            "https://youtu.be/h4gw6gCP5ls",
        ]
        assert lesson["activity_uuid"] == stable_uuid("activity", 1, 1)
        assert lesson["folder_uuid"] == stable_uuid("folder", 1)
        assert (lesson["week_order"], lesson["activity_order"]) == (1, 1)
        assert lesson["source_references"][0]["parent_inference"] == "inferred_from_activity_order"
        assert "adalove_order_audit" in lesson["fields"]

    def test_orientation_and_its_self_study_are_reported_and_dropped(self, tmp_path):
        path = tmp_path / "orientation.xlsx"
        write_syllabus(
            path,
            [
                syllabus_row(title="Aula", seq=1),
                syllabus_row(
                    title="Orientação", seq=2, kind="Orientation", subject=None
                ),
                syllabus_row(
                    title="Material da orientação", seq=3, kind="Self-study",
                    parent="Orientação", parent_order=2,
                    url="https://example.com/orientation", subject=None,
                ),
            ],
        )

        parsed = parse_workbook(path)

        assert [lesson["title"] for lesson in parsed["lessons"]] == ["Aula"]
        assert parsed["dropped_summary"] == {
            "orientation_count": 1,
            "orientation_self_study_count": 1,
            "no_parent_count": 0,
            "total_count": 2,
        }
        assert [item["reason"] for item in parsed["dropped"]] == [
            "orientation", "parent_orientation",
        ]

    def test_self_study_without_preceding_anchor_is_dropped_not_rejected(self, tmp_path):
        path = tmp_path / "orphan-self-study.xlsx"
        write_syllabus(
            path,
            [
                activity(
                    title="Leitura antes da aula",
                    kind="Self-study",
                    week=1,
                    order=1,
                    parent_inference="no_preceding_anchor_in_week",
                    url="https://example.com/orphan",
                ),
                syllabus_row(title="Aula", seq=2),
            ],
        )

        parsed = parse_workbook(path)

        assert [lesson["title"] for lesson in parsed["lessons"]] == ["Aula"]
        assert parsed["source_count"] == 0
        assert parsed["dropped_summary"] == {
            "orientation_count": 0,
            "orientation_self_study_count": 0,
            "no_parent_count": 1,
            "total_count": 1,
        }
        assert parsed["dropped"] == [
            {
                "activity_uuid": stable_uuid("activity", 1, 1),
                "type": "Self-study",
                "title": "Leitura antes da aula",
                "parent_activity_uuid": None,
                "parent_inference": "no_preceding_anchor_in_week",
                "reason": "no_parent",
            }
        ]

    def test_self_study_without_parent_or_inference_is_still_rejected(self, tmp_path):
        path = tmp_path / "unlabeled-self-study.xlsx"
        write_syllabus(
            path,
            [
                activity(
                    title="Leitura solta",
                    kind="Self-study",
                    week=1,
                    order=1,
                    url="https://example.com/loose",
                ),
                syllabus_row(title="Aula", seq=2),
            ],
        )

        with pytest.raises(ValueError, match="sem pai inferido e identificado"):
            parse_workbook(path)

    @pytest.mark.parametrize(
        ("subject", "message"),
        [
            ("Marketing", "COM Computação"),
            ("SI", "MTF Matemática"),
        ],
    )
    def test_unknown_eixo_error_explains_the_column_and_every_accepted_value(
        self, tmp_path, subject, message
    ):
        path = tmp_path / "unmapped-subject.xlsx"
        write_syllabus(path, [syllabus_row(subject=subject)])

        with pytest.raises(ValueError, match=message):
            parse_workbook(path)

    def test_self_study_with_unknown_lesson_subject_is_rejected(self, tmp_path):
        path = tmp_path / "self-study-with-unmapped-subject.xlsx"
        write_syllabus(
            path,
            [
                syllabus_row(title="Aula de SQL"),
                syllabus_row(
                    seq=2,
                    title="Leitura de SQL",
                    kind="Self-study",
                    parent="Aula de SQL",
                    subject="Marketing",
                    url="https://example.com/sql",
                ),
            ],
        )

        with pytest.raises(ValueError) as error:
            parse_workbook(path)

        assert str(error.value) == (
            "A linha 3 da aba Activities tem o Eixo 'Marketing' na coluna "
            "'Lesson Subject code'. O Eixo identifica a área curricular da aula. "
            "Use um destes valores: COM Computação, LID Liderança, NEG Negócios, "
            "UEX User Experience, MTF Matemática."
        )

    def test_workbook_without_full_fidelity_sheets_is_rejected(self, tmp_path):
        path = tmp_path / "incomplete.xlsx"
        workbook = Workbook()
        workbook.active.title = "Summary"
        workbook.active.append(("Week", "Activity"))
        workbook.save(path)
        workbook.close()

        with pytest.raises(ValueError, match="Faltam estas abas: Activities"):
            parse_workbook(path)

    def test_order_audit_duplicate_is_actionable(self, tmp_path):
        path = tmp_path / "duplicate-order.xlsx"
        write_syllabus(path, [syllabus_row()])
        workbook = load_workbook(path)
        sheet = workbook["Order audit"]
        headers = [cell.value for cell in sheet[1]]
        sheet.cell(2, headers.index("Duplicate order key") + 1, "yes")
        workbook.save(path)
        workbook.close()

        with pytest.raises(ValueError, match="chave '1:1' como duplicada"):
            parse_workbook(path)


class TestSourceIdentity:
    def test_articles_drop_fragment_and_tracking(self):
        assert canonical_url(" https://EXAMPLE.com/read?a=1&utm_source=x#part ") == (
            "https://example.com/read?a=1"
        )
        assert source_identity("https://example.com/read#part") == {
            "kind": "article",
            "canonical_url": "https://example.com/read",
        }

    @pytest.mark.parametrize(
        "invalid_url",
        (
            "medium.com/read",
            "/relative/path",
            "mailto:reader@example.com",
            "javascript:alert(1)",
        ),
    )
    def test_articles_require_an_absolute_http_url(self, invalid_url):
        assert canonical_url(invalid_url) == ""
        assert source_identity(invalid_url) is None

    def test_youtube_identity_ignores_timestamp_and_url_form(self):
        watch = source_identity("https://www.youtube.com/watch?v=h4gw6gCP5ls&t=273s")
        short = source_identity("https://youtu.be/h4gw6gCP5ls?t=4")

        assert watch == short == {
            "kind": "video",
            "provider": "youtube",
            "video_id": "h4gw6gCP5ls",
        }

    def test_youtube_identity_repairs_textual_query_separators_from_xlsx(self):
        identity = source_identity(
            "https://www.youtube.com/watch?v=h4gw6gCP5ls%20e%20"
            "list=PL9iw99lS3Prg0hPSCiOz9AXeEmj8W8fL8%20e%20index=14%20e%20t=49s"
        )

        assert identity == {
            "kind": "video",
            "provider": "youtube",
            "video_id": "h4gw6gCP5ls",
        }

    @pytest.mark.parametrize(
        "url",
        (
            "https://www.youtube.com/watch?v=not-a-real-id",
            "https://youtu.be/too-short",
            "https://www.youtube.com/watch?v=h4gw6gCP5ls%20unexpected",
        ),
    )
    def test_youtube_identity_rejects_invalid_provider_ids(self, url):
        assert source_identity(url) is None

    def test_books_are_resource_code_plus_scope_not_gateway_url(self):
        gateway = "https://philos.sophia.com.br/terminal/9418"
        first = source_identity(
            gateway,
            media_kind="book",
            resource_code="978-85-224-8504-8",
            scope_kind="pages",
            scope_value="11 à 18",
        )
        other_scope = source_identity(
            gateway,
            media_kind="book",
            resource_code="9788522485048",
            scope_kind="pages",
            scope_value="19-20",
        )
        other_book = source_identity(
            gateway,
            media_kind="book",
            resource_code="9788569726760",
            scope_kind="pages",
            scope_value="11-18",
        )

        assert first == {
            "kind": "book",
            "resource_code": "9788522485048",
            "scope": {"kind": "pages", "value": "11-18"},
        }
        assert first != other_scope
        assert first != other_book

    def test_incomplete_book_reference_does_not_mint_a_false_source(self):
        assert source_identity(
            "https://philos.sophia.com.br/terminal/9418",
            media_kind="book",
            resource_code="9788522485048",
        ) is None


class TestVersionedImport:
    def test_import_persists_only_deliverable_materials_as_sources(self, db, tmp_path):
        path = tmp_path / "activity-materials.xlsx"
        write_syllabus(
            path,
            [
                syllabus_row(
                    title="Aula de SQL",
                    kind="Class",
                    materials=[
                        material(
                            url="https://example.com/class-slides",
                            label="Slides da aula",
                        )
                    ],
                ),
                syllabus_row(
                    seq=2,
                    title="Entrega de SQL",
                    kind="Deliverable",
                    materials=[
                        material(
                            url="https://example.com/deliverable-brief",
                            label="Briefing da entrega",
                        )
                    ],
                ),
                syllabus_row(
                    seq=3,
                    title="Avaliação de SQL",
                    kind="Evaluation",
                    materials=[
                        material(
                            url="https://example.com/evaluation-rubric",
                            label="Rubrica da avaliação",
                        )
                    ],
                ),
            ],
        )

        result = import_workbook(
            db, path, "Activity materials", require_syllabus_metadata=False
        )
        lessons = {
            lesson["kind"]: lesson
            for lesson in get_syllabus_version(db, result["syllabus_id"])["lessons"]
        }

        assert result["reference_count"] == 1
        assert lessons["Class"]["sources"] == []
        assert lessons["Evaluation"]["sources"] == []
        assert [source["url"] for source in lessons["Deliverable"]["sources"]] == [
            "https://example.com/deliverable-brief"
        ]
        assert lessons["Deliverable"]["sources"][0]["activity_uuid"] == stable_uuid(
            "activity", 1, 2
        )
        assert lessons["Deliverable"]["sources"][0]["parent_activity_uuid"] is None

    def test_import_cli_has_no_manual_graph_id_override(self):
        with pytest.raises(SystemExit):
            build_parser().parse_args(
                [
                    "import",
                    "syllabus.xlsx",
                    "--name",
                    "Derived identity",
                    "--graph-id",
                    "graph-manual-override",
                ]
            )

    def test_new_named_import_requires_durable_metadata_by_default(self, db, tmp_path):
        path = tmp_path / "requires-metadata.xlsx"
        write_syllabus(path, [syllabus_row(project="Metadata required")])

        with pytest.raises(ValueError, match="instituição"):
            import_workbook(db, path, "Metadata required")

        assert db.execute(
            "SELECT 1 FROM syllabus WHERE id = 'metadata-required'"
        ).fetchone() is None

    def test_lesson_subjects_carry_the_display_name_the_companion_themes_by(
        self, db, tmp_path
    ):
        db.execute(
            "INSERT INTO institution (id, name) VALUES ('inteli', 'Inteli')"
            " ON CONFLICT (id) DO NOTHING"
        )
        path = tmp_path / "subject-display-names.xlsx"
        write_syllabus(
            path,
            [
                syllabus_row(title="UX", seq=1, subject="User Experience"),
                syllabus_row(title="Cálculo", seq=2, subject="MTF"),
                syllabus_row(title="Física", seq=3, subject="Matemática e Física"),
            ],
        )
        result = import_workbook(
            db, path, "Subject display names", institution_id="inteli"
        )

        subjects = _lesson_subjects_by_syllabus(db, [result["syllabus_id"]])

        assert subjects[result["syllabus_id"]] == [
            {
                "code": "MTF",
                "display_name": "Matemática",
                "graph_id": "graph-inteli-subject-display-names-mtf",
            },
            {
                "code": "UEX",
                "display_name": "User Experience",
                "graph_id": "graph-inteli-subject-display-names-uex",
            },
        ]

    def test_new_version_keeps_stored_subject_graph_id_even_when_companion_lists_it(
        self, db, tmp_path
    ):
        db.execute(
            "INSERT INTO institution (id, name) VALUES ('inteli', 'Inteli')"
            " ON CONFLICT (id) DO NOTHING"
        )
        path = tmp_path / "own-graph-id.xlsx"
        write_syllabus(path, [syllabus_row(project="Own graph id")])
        first = import_workbook(db, path, "Own graph id", institution_id="inteli")
        minted = db.execute(
            "SELECT graph_id FROM syllabus_subject"
            " WHERE syllabus_id = %s AND lesson_subject_code = 'COM'",
            (first["syllabus_id"],),
        ).fetchone()[0]
        assert minted == subject_graph_id_for("inteli", "Own graph id", "COM")

        write_syllabus(
            path,
            [syllabus_row(project="Own graph id", title="Lesson renamed")],
        )
        second = import_workbook(
            db,
            path,
            "Own graph id",
            syllabus_id=first["syllabus_id"],
            occupied_graph_ids={minted},
        )

        assert second["seq"] == 2
        assert second["version_id"] != first["version_id"]
        assert db.execute(
            "SELECT graph_id FROM syllabus_subject"
            " WHERE syllabus_id = %s AND lesson_subject_code = 'COM'",
            (first["syllabus_id"],),
        ).fetchone()[0] == minted

    def test_new_subject_rejects_an_occupied_graph_id(
        self, db, tmp_path
    ):
        db.execute(
            "INSERT INTO institution (id, name) VALUES ('inteli', 'Inteli')"
            " ON CONFLICT (id) DO NOTHING"
        )
        db.execute(
            "INSERT INTO syllabus (id, title, institution_id)"
            " VALUES ('no-graph-id', 'No graph id', 'inteli')"
        )
        db.execute(
            "INSERT INTO syllabus_version (id, syllabus_id, seq, origin)"
            " VALUES ('no-graph-id:v0001', 'no-graph-id', 1, 'upload')"
        )
        db.commit()
        derived = subject_graph_id_for("inteli", "No graph id", "COM")
        path = tmp_path / "no-graph-id.xlsx"
        write_syllabus(path, [syllabus_row(project="No graph id")])

        with pytest.raises(GraphIdConflict) as raised:
            import_workbook(
                db,
                path,
                "No graph id",
                syllabus_id="no-graph-id",
                occupied_graph_ids={derived},
            )
        db.rollback()

        assert raised.value.graph_id == derived
        assert db.execute(
            "SELECT count(*) FROM syllabus_subject WHERE syllabus_id = 'no-graph-id'"
        ).fetchone()[0] == 0
        assert db.execute(
            "SELECT count(*) FROM syllabus_version WHERE syllabus_id = 'no-graph-id'"
        ).fetchone()[0] == 1

    def test_name_is_manual_and_uploaded_xlsx_is_retained_exactly(self, db, tmp_path):
        path = tmp_path / "input.xlsx"
        write_syllabus(
            path,
            [
                syllabus_row(
                    project="DO NOT USE AS NAME",
                    subjects="Arquitetura de nuvem\n,Modelos de serviço em nuvem",
                )
            ],
        )
        original = path.read_bytes()
        run_count = db.execute("SELECT count(*) FROM run").fetchone()[0]

        result = import_workbook(
            db, path, "SI módulo 7 2026", require_syllabus_metadata=False
        )

        assert result["syllabus_id"] == "si-modulo-7-2026"
        assert result["seq"] == 1
        assert result["lesson_count"] == 1
        assert result["reference_count"] == 0
        assert db.execute("SELECT count(*) FROM run").fetchone()[0] == run_count
        stored = get_syllabus_workbook(db, result["version_id"])
        assert stored == {
            "file_name": "input.xlsx",
            "mime_type": XLSX_MIME,
            "sha256": hashlib.sha256(original).hexdigest(),
            "body": original,
        }
        assert list_syllabi(db)[0]["title"] == "SI módulo 7 2026"
        assert get_syllabus_version(db, result["syllabus_id"])["lessons"][0][
            "subjects"
        ] == ["Arquitetura de nuvem", "Modelos de serviço em nuvem"]

    def test_import_reports_orientations_that_were_not_stored(self, db, tmp_path):
        path = tmp_path / "dropped-orientations.xlsx"
        write_syllabus(
            path,
            [
                syllabus_row(title="Aula", seq=1),
                syllabus_row(
                    title="Orientação", seq=2, kind="Orientation", subject=None
                ),
                syllabus_row(
                    title="Material da orientação",
                    seq=3,
                    kind="Self-study",
                    parent="Orientação",
                    parent_order=2,
                    url="https://example.com/orientation",
                    subject=None,
                ),
            ],
        )

        result = import_workbook(
            db, path, "Orientation drop report", require_syllabus_metadata=False
        )

        assert result["dropped_summary"] == {
            "orientation_count": 1,
            "orientation_self_study_count": 1,
            "no_parent_count": 0,
            "total_count": 2,
        }
        assert result["lesson_count"] == 1
        assert result["reference_count"] == 0
        assert [
            lesson["title"]
            for lesson in get_syllabus_version(db, result["syllabus_id"])["lessons"]
        ] == ["Aula"]

    def test_import_reports_self_study_without_preceding_anchor(self, db, tmp_path):
        path = tmp_path / "orphan-self-study.xlsx"
        write_syllabus(
            path,
            [
                activity(
                    title="Leitura antes da aula",
                    kind="Self-study",
                    week=1,
                    order=1,
                    parent_inference="no_preceding_anchor_in_week",
                    url="https://example.com/orphan",
                ),
                syllabus_row(title="Aula", seq=2),
            ],
        )

        result = import_workbook(
            db, path, "Orphan self-study report", require_syllabus_metadata=False
        )

        assert result["dropped_summary"] == {
            "orientation_count": 0,
            "orientation_self_study_count": 0,
            "no_parent_count": 1,
            "total_count": 1,
        }
        assert [item["reason"] for item in result["dropped"]] == ["no_parent"]
        assert result["lesson_count"] == 1
        assert result["reference_count"] == 0
        stored = get_syllabus_version(db, result["syllabus_id"])["lessons"]
        assert [(lesson["title"], lesson["sources"]) for lesson in stored] == [("Aula", [])]

    def test_same_file_is_idempotent_within_one_syllabus(self, db, tmp_path):
        path = tmp_path / "same.xlsx"
        write_syllabus(path, [syllabus_row(project="UNRELATED")])

        first = import_workbook(
            db, path, "Idempotent syllabus", require_syllabus_metadata=False
        )
        second = import_workbook(
            db, path, "Idempotent syllabus", require_syllabus_metadata=False
        )

        assert second["unchanged"] is True
        assert second["version_id"] == first["version_id"]
        assert second["lesson_count"] == 1
        assert second["reference_count"] == 0
        assert second["source_count"] == 0
        assert second["new_source_count"] == 0
        assert get_syllabus_history(db, first["syllabus_id"])["versions"][0]["seq"] == 1

    def test_import_counts_reused_references_as_sources_without_calling_them_new(
        self, db, tmp_path
    ):
        path = tmp_path / "reused-source.xlsx"
        lesson = syllabus_row(project="IGNORED")
        source = syllabus_row(
            project="IGNORED",
            seq="2",
            title="Assigned article",
            kind="Autoestudo",
            parent="Lesson",
            url="https://example.com/reused",
        )
        write_syllabus(path, [lesson, source])
        first = import_workbook(
            db, path, "Reused source syllabus", require_syllabus_metadata=False
        )

        source["Description"] = "Description changed without changing source identity"
        write_syllabus(path, [lesson, source])
        second = import_workbook(
            db, path, "Reused source syllabus", require_syllabus_metadata=False
        )
        unchanged = import_workbook(
            db, path, "Reused source syllabus", require_syllabus_metadata=False
        )

        assert first["source_count"] == first["reference_count"] == 1
        assert first["new_source_count"] == 1
        assert second["source_count"] == second["reference_count"] == 1
        assert second["new_source_count"] == 0
        assert unchanged["source_count"] == unchanged["reference_count"] == 1
        assert unchanged["new_source_count"] == 0

    def test_upload_event_allocator_ignores_uuid_style_event_suffixes(
        self, db, tmp_path
    ):
        db.execute(
            "INSERT INTO curation_event (id, actor, action, subject)"
            " VALUES (%s, 'worker', 'source_acquisition_queued', '{}'::jsonb)",
            ("ce-acq-deadbeef999999999999999999999999",),
        )
        db.commit()
        path = tmp_path / "curation-id.xlsx"
        write_syllabus(path, [syllabus_row(project="IGNORED")])

        result = import_workbook(
            db, path, "Strict curation id syllabus", require_syllabus_metadata=False
        )

        event_id = db.execute(
            "SELECT id FROM curation_event"
            " WHERE action = 'syllabus_upload' AND subject->>'version_id' = %s",
            (result["version_id"],),
        ).fetchone()[0]
        assert re.fullmatch(r"ce\d+", event_id)

    def test_latest_and_history_are_full_immutable_versions(self, db, tmp_path):
        path = tmp_path / "versions.xlsx"
        lesson = syllabus_row(project="IGNORED")
        source = syllabus_row(
            project="IGNORED",
            seq="2",
            title="Assigned article",
            kind="Autoestudo",
            parent="Lesson",
            url="https://example.com/v1",
        )
        write_syllabus(path, [lesson, source])
        first = import_workbook(
            db, path, "Versioned syllabus", require_syllabus_metadata=False
        )
        first_view = get_syllabus_version(db, first["syllabus_id"])

        source["Primary URL"] = "https://example.com/v2"
        source["_materials"] = [
            material(url="https://example.com/v2", label="Assigned article")
        ]
        source["Description"] = "Changed description"
        write_syllabus(path, [lesson, source])
        second = import_workbook(
            db, path, "Versioned syllabus", require_syllabus_metadata=False
        )

        latest = get_syllabus_version(db, first["syllabus_id"])
        historical = get_syllabus_version(db, first["syllabus_id"], first["version_id"])
        history = get_syllabus_history(db, first["syllabus_id"])
        assert [version["seq"] for version in history["versions"]] == [2, 1]
        assert latest["version"]["id"] == second["version_id"]
        assert latest["lessons"][0]["sources"][0]["url"] == "https://example.com/v2"
        assert historical["lessons"][0]["sources"][0]["url"] == "https://example.com/v1"
        assert first_view["lessons"] == historical["lessons"]

    def test_removed_reference_is_hidden_latest_but_source_and_artifact_survive(self, db, tmp_path):
        path = tmp_path / "removal.xlsx"
        lesson = syllabus_row(project="IGNORED")
        source = syllabus_row(
            project="IGNORED",
            seq="2",
            title="Assigned article",
            kind="Autoestudo",
            parent="Lesson",
            url="https://example.com/lasting",
        )
        write_syllabus(path, [lesson, source])
        first = import_workbook(
            db, path, "Removal syllabus", require_syllabus_metadata=False
        )
        first_view = get_syllabus_version(db, first["syllabus_id"])
        source_id = first_view["lessons"][0]["sources"][0]["source_id"]
        db.execute(
            "INSERT INTO source_snapshot"
            " (id, source_id, content_hash, status) VALUES (%s, %s, %s, 'ok')",
            ("snap-lasting", source_id, "sha-lasting"),
        )
        db.execute(
            "INSERT INTO artifact (id, snapshot_id, kind, tool, body)"
            " VALUES ('art-lasting', 'snap-lasting', 'markdown', 'test', '# Preserved')"
        )
        db.commit()

        write_syllabus(path, [lesson])
        import_workbook(
            db, path, "Removal syllabus", require_syllabus_metadata=False
        )

        latest = get_syllabus_version(db, first["syllabus_id"])
        historical = get_syllabus_version(db, first["syllabus_id"], first["version_id"])
        assert latest["lessons"][0]["sources"] == []
        assert historical["lessons"][0]["sources"][0]["source_id"] == source_id
        assert db.execute("SELECT body FROM artifact WHERE id = 'art-lasting'").fetchone()[0] == "# Preserved"

    def test_sophia_gateway_mints_distinct_sources_by_code_and_scope(self, db, tmp_path):
        path = tmp_path / "books.xlsx"
        gateway = "https://philos.sophia.com.br/terminal/9418"
        write_syllabus(
            path,
            [
                syllabus_row(title="Books lesson"),
                syllabus_row(
                    seq=2,
                    kind="Self-study",
                    title="Book A",
                    parent="Books lesson",
                    description="Páginas 10-20",
                    url=gateway,
                    code="9780000000001",
                ),
                syllabus_row(
                    seq=3,
                    kind="Self-study",
                    title="Book B",
                    parent="Books lesson",
                    description="Páginas 10-20",
                    url=gateway,
                    code="9780000000002",
                ),
                syllabus_row(
                    seq=4,
                    kind="Self-study",
                    title="Book A later",
                    parent="Books lesson",
                    description="Páginas 21-30",
                    url=gateway,
                    code="9780000000001",
                ),
            ],
        )

        result = import_workbook(
            db, path, "Books syllabus", require_syllabus_metadata=False
        )
        view = get_syllabus_version(db, result["syllabus_id"])
        source_ids = {source["source_id"] for source in view["lessons"][0]["sources"]}

        assert len(source_ids) == 3
        assert None not in source_ids

    def test_book_without_scope_remains_visible_without_source(self, db, tmp_path):
        path = tmp_path / "incomplete-book.xlsx"
        gateway = "https://philos.sophia.com.br/terminal/9418"
        write_syllabus(
            path,
            [
                syllabus_row(title="Books lesson"),
                syllabus_row(
                    seq=2,
                    kind="Self-study",
                    title="Book without scope",
                    parent="Books lesson",
                    description="Read this book",
                    url=gateway,
                    code="9780000000001",
                ),
            ],
        )

        result = import_workbook(
            db, path, "Incomplete books", require_syllabus_metadata=False
        )
        source = get_syllabus_version(db, result["syllabus_id"])["lessons"][0]["sources"][0]

        assert source["media_type"] == "book"
        assert source["resource_code"] == "9780000000001"
        assert source["scope"] is None
        assert source["source_id"] is None


class TestSyllabusCuration:
    def _import_editable(self, db, tmp_path):
        path = tmp_path / "editable.xlsx"
        write_syllabus(
            path,
            [
                syllabus_row(title="Aula editável", description="Descrição original"),
                syllabus_row(
                    seq=2,
                    kind="Self-study",
                    title="Fonte A",
                    parent="Aula editável",
                    description="Descrição A",
                    url="https://example.com/a",
                ),
                syllabus_row(
                    seq=3,
                    kind="Self-study",
                    title="Fonte B",
                    parent="Aula editável",
                    description="Descrição B",
                    url="https://example.com/b",
                ),
            ],
        )
        return import_workbook(
            db, path, "Syllabus editável", require_syllabus_metadata=False
        )

    def test_editor_authors_new_complete_version_and_preserves_old_facts(self, db, tmp_path):
        imported = self._import_editable(db, tmp_path)
        before = get_syllabus_version(db, imported["syllabus_id"])
        source_a = before["lessons"][0]["sources"][0]
        source_b = before["lessons"][0]["sources"][1]
        payload = [
            {
                "id": before["lessons"][0]["id"],
                "week": 2,
                "kind": "Class",
                "title": "Aula revisada",
                "subject": "NEG",
                "subjects": ["Custos fixos", "Custos variáveis"],
                "date": "2026-08-13",
                "description": "Descrição revisada",
                "sources": [
                    {
                        "reference_id": source_b["reference_id"],
                        "title": "Fonte B revisada",
                        "description": "Descrição B revisada",
                        "url": source_b["url"],
                        "media_type": "article",
                        "hidden": True,
                    },
                    {
                        "title": "Fonte nova",
                        "description": "Adicionada manualmente",
                        "url": "https://example.com/new",
                        "media_type": "article",
                        "hidden": False,
                    },
                ],
            }
        ]

        result = curate_syllabus(
            db, imported["syllabus_id"], imported["version_id"], payload,
            note="Reorganiza fontes e revisa a aula.",
        )

        assert result["seq"] == 2
        assert result["unchanged"] is False
        assert result["reference_count"] == 2
        latest = get_syllabus_version(db, imported["syllabus_id"])
        historical = get_syllabus_version(
            db, imported["syllabus_id"], imported["version_id"]
        )
        assert latest["lessons"][0]["title"] == "Aula revisada"
        assert latest["lessons"][0]["subjects"] == [
            "Custos fixos",
            "Custos variáveis",
        ]
        assert [source["title"] for source in latest["lessons"][0]["sources"]] == [
            "Fonte B revisada",
            "Fonte nova",
        ]
        assert latest["lessons"][0]["sources"][0]["hidden"] is True
        assert historical["lessons"][0]["title"] == "Aula editável"
        assert [source["title"] for source in historical["lessons"][0]["sources"]] == [
            "Fonte A",
            "Fonte B",
        ]
        assert db.execute("SELECT count(*) FROM source WHERE id = %s", (source_a["source_id"],)).fetchone()[0] == 1

    def test_curated_xlsx_round_trips_order_edits_and_visibility(self, db, tmp_path):
        imported = self._import_editable(db, tmp_path)
        before = get_syllabus_version(db, imported["syllabus_id"])
        sources = before["lessons"][0]["sources"]
        payload = [
            {
                **{key: before["lessons"][0].get(key) for key in ("id", "week", "kind", "title", "subject", "description")},
                "subjects": ["Custos fixos", "Rentabilidade, ROI, EBITDA"],
                "date": "2026-08-06",
                "sources": [
                    {
                        **{key: sources[1].get(key) for key in (
                            "reference_id", "title", "description", "url", "media_type",
                            "resource_code", "scope_kind", "scope_value",
                        )},
                        "hidden": True,
                    },
                    {
                        **{key: sources[0].get(key) for key in (
                            "reference_id", "title", "description", "url", "media_type",
                            "resource_code", "scope_kind", "scope_value",
                        )},
                        "hidden": False,
                    },
                ],
            }
        ]
        curated = curate_syllabus(
            db, imported["syllabus_id"], imported["version_id"], payload,
            note="Reordena fontes e atualiza a visibilidade.",
        )
        workbook = get_syllabus_workbook(db, curated["version_id"])
        exported = tmp_path / workbook["file_name"]
        exported.write_bytes(workbook["body"])

        parsed = parse_workbook(exported)

        assert parsed["format"] == "adalove-observer"
        assert parsed["lessons"][0]["subjects"] == [
            "Custos fixos",
            "Rentabilidade, ROI, EBITDA",
        ]
        assert [source["title"] for source in parsed["lessons"][0]["source_references"]] == [
            "Fonte B",
            "Fonte A",
        ]
        assert parsed["lessons"][0]["source_references"][0]["is_hidden"] is True
        assert parsed["lessons"][0]["source_references"][1]["is_hidden"] is False

    def test_curated_deliverable_material_round_trips_on_its_activity(self, db, tmp_path):
        original = tmp_path / "deliverable.xlsx"
        write_syllabus(
            original,
            [
                syllabus_row(
                    title="Entrega de SQL",
                    kind="Deliverable",
                    materials=[
                        material(
                            url="https://example.com/deliverable-brief",
                            label="Briefing da entrega",
                        )
                    ],
                )
            ],
        )
        imported = import_workbook(
            db, original, "Deliverable round trip", require_syllabus_metadata=False
        )
        before = get_syllabus_version(db, imported["syllabus_id"])
        deliverable = before["lessons"][0]
        curated = curate_syllabus(
            db,
            imported["syllabus_id"],
            imported["version_id"],
            [
                {
                    "id": deliverable["id"],
                    "week": deliverable["week"],
                    "kind": deliverable["kind"],
                    "title": "Entrega de SQL revisada",
                    "subject": deliverable["subject"],
                    "subjects": deliverable["subjects"],
                    "date": str(deliverable["date"]),
                    "description": deliverable["description"],
                    "hidden": deliverable["hidden"],
                    "sources": deliverable["sources"],
                }
            ],
            note="Revisa o título da entrega.",
        )
        workbook = get_syllabus_workbook(db, curated["version_id"])
        compiled = tmp_path / workbook["file_name"]
        compiled.write_bytes(workbook["body"])

        parsed = parse_workbook(compiled)

        assert [
            {
                "kind": lesson["kind"],
                "title": lesson["title"],
                "activity_uuid": lesson["activity_uuid"],
                "sources": [source["url"] for source in lesson["source_references"]],
            }
            for lesson in parsed["lessons"]
        ] == [
            {
                "kind": "Deliverable",
                "title": "Entrega de SQL revisada",
                "activity_uuid": deliverable["activity_uuid"],
                "sources": ["https://example.com/deliverable-brief"],
            }
        ]

    def test_stale_editor_cannot_overwrite_a_newer_version(self, db, tmp_path):
        imported = self._import_editable(db, tmp_path)
        before = get_syllabus_version(db, imported["syllabus_id"])
        lesson = before["lessons"][0]
        payload = [{
            "id": lesson["id"], "week": lesson["week"], "kind": lesson["kind"],
            "title": "Primeira mudança", "subject": lesson["subject"],
            "date": str(lesson["date"] or ""), "description": lesson["description"],
            "sources": lesson["sources"],
        }]
        curate_syllabus(
            db, imported["syllabus_id"], imported["version_id"], payload,
            note="Primeira mudança concorrente.",
        )

        with pytest.raises(SyllabusVersionConflict, match="versão mais nova"):
            curate_syllabus(
                db, imported["syllabus_id"], imported["version_id"], payload,
                note="Tentativa sobre versão antiga.",
            )

    def test_saving_an_unchanged_projection_does_not_mint_noise_version(self, db, tmp_path):
        imported = self._import_editable(db, tmp_path)
        before = get_syllabus_version(db, imported["syllabus_id"])
        history_before = get_syllabus_history(db, imported["syllabus_id"])["versions"]
        payload = []
        for lesson in before["lessons"]:
            payload.append(
                {
                    "id": lesson["id"],
                    "week": lesson["week"],
                    "kind": lesson["kind"],
                    "title": lesson["title"],
                    "subject": lesson["subject"],
                    "date": str(lesson["date"] or ""),
                    "description": lesson["description"],
                    "sources": lesson["sources"],
                }
            )

        result = curate_syllabus(
            db, imported["syllabus_id"], imported["version_id"], payload
        )

        assert result["unchanged"] is True
        assert result["version_id"] == imported["version_id"]
        assert get_syllabus_history(db, imported["syllabus_id"])["versions"] == history_before

    def _import_with_dropped_orientation(self, db, tmp_path):
        path = tmp_path / "gapped-orders.xlsx"
        write_syllabus(
            path,
            [
                syllabus_row(
                    title="Orientação", seq=1, kind="Orientation", subject=None
                ),
                syllabus_row(title="Aula", seq=2),
                syllabus_row(
                    seq=3,
                    kind="Self-study",
                    title="Leitura",
                    parent="Aula",
                    parent_order=2,
                    url="https://example.com/reading",
                ),
                syllabus_row(
                    title="Entrega",
                    seq=4,
                    kind="Deliverable",
                    subject=None,
                    materials=[
                        material(
                            url="https://example.com/brief", label="Briefing"
                        )
                    ],
                ),
            ],
        )
        return import_workbook(
            db, path, "Syllabus com ordem do Adalove", require_syllabus_metadata=False
        )

    @staticmethod
    def _editor_payload(lessons):
        return [
            {
                "id": lesson["id"],
                "week": lesson["week"],
                "kind": lesson["kind"],
                "title": lesson["title"],
                "subject": lesson["subject"],
                "subjects": lesson["subjects"],
                "date": str(lesson["date"] or ""),
                "description": lesson["description"],
                "hidden": lesson["hidden"],
                "sources": lesson["sources"],
            }
            for lesson in lessons
        ]

    @staticmethod
    def _order_keys(version):
        return [
            (
                lesson["week_order"],
                lesson["activity_order"],
                [
                    (
                        source["week_order"],
                        source["activity_order"],
                        source["parent_activity_uuid"],
                        source["parent_inference"],
                    )
                    for source in lesson["sources"]
                ],
            )
            for lesson in version["lessons"]
        ]

    def test_noop_editor_save_keeps_adalove_order_keys_after_dropped_orientation(
        self, db, tmp_path
    ):
        imported = self._import_with_dropped_orientation(db, tmp_path)
        before = get_syllabus_version(db, imported["syllabus_id"])
        class_uuid = before["lessons"][0]["activity_uuid"]
        expected_keys = [
            (1, 2, [(1, 3, class_uuid, "inferred_from_activity_order")]),
            (1, 4, [(1, 4, None, None)]),
        ]
        assert self._order_keys(before) == expected_keys

        result = curate_syllabus(
            db,
            imported["syllabus_id"],
            imported["version_id"],
            self._editor_payload(before["lessons"]),
        )

        assert result["unchanged"] is True
        assert result["version_id"] == imported["version_id"]
        assert self._order_keys(get_syllabus_version(db, imported["syllabus_id"])) == expected_keys

    def test_editor_edit_keeps_adalove_order_keys_and_deliverable_material_parent(
        self, db, tmp_path
    ):
        imported = self._import_with_dropped_orientation(db, tmp_path)
        before = get_syllabus_version(db, imported["syllabus_id"])
        class_uuid = before["lessons"][0]["activity_uuid"]
        payload = self._editor_payload(before["lessons"])
        payload[0]["title"] = "Aula revisada"
        payload[1]["sources"][0]["title"] = "Briefing revisado"

        curated = curate_syllabus(
            db, imported["syllabus_id"], imported["version_id"], payload,
            note="Revisa títulos sem mexer na ordem.",
        )

        assert curated["unchanged"] is False
        latest = get_syllabus_version(db, imported["syllabus_id"])
        assert [lesson["title"] for lesson in latest["lessons"]] == ["Aula revisada", "Entrega"]
        assert latest["lessons"][1]["sources"][0]["title"] == "Briefing revisado"
        assert self._order_keys(latest) == [
            (1, 2, [(1, 3, class_uuid, "inferred_from_activity_order")]),
            (1, 4, [(1, 4, None, None)]),
        ]

    def test_editor_mints_order_keys_only_for_authored_rows(self, db, tmp_path):
        imported = self._import_with_dropped_orientation(db, tmp_path)
        before = get_syllabus_version(db, imported["syllabus_id"])
        class_uuid = before["lessons"][0]["activity_uuid"]
        payload = self._editor_payload(before["lessons"])
        payload[0]["sources"].append(
            {
                "title": "Fonte nova",
                "url": "https://example.com/new",
                "media_type": "article",
            }
        )

        curate_syllabus(
            db, imported["syllabus_id"], imported["version_id"], payload,
            note="Adiciona uma fonte na aula.",
        )

        latest = get_syllabus_version(db, imported["syllabus_id"])
        class_sources = latest["lessons"][0]["sources"]
        assert [source["title"] for source in class_sources] == ["Leitura", "Fonte nova"]
        assert (class_sources[0]["week_order"], class_sources[0]["activity_order"]) == (1, 3)
        authored = class_sources[1]
        assert authored["week_order"] == 1
        assert authored["activity_order"] > 4
        assert authored["parent_activity_uuid"] == class_uuid
        assert authored["parent_inference"] == "curated_explicit_parent"
        assert self._order_keys(latest)[1] == (1, 4, [(1, 4, None, None)])
