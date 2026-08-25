"""Browser regressions for the syllabus curation interface."""

from contextlib import contextmanager
from copy import deepcopy
from pathlib import Path
import socket
import threading
import time
import uuid
from urllib.parse import quote, urlparse

import psycopg
from openpyxl import Workbook
from playwright.sync_api import expect, sync_playwright
import uvicorn

from universe.syllabus import (
    LEGACY_COLUMNS,
    PROJECT_COLUMNS,
    get_syllabus_history,
    get_syllabus_version,
    import_workbook,
)
from universe.web.app import create_app


LOCAL_KC_STAGES = (
    "blocks",
    "passage-cuts",
    "passage-triage",
    "task-generation",
    "task-granularity",
    "task-revision",
    "task-triage",
    "task-substance",
    "kc-statement",
    "task-modality",
    "task-knowledge",
)


def _editable_workbook(path: Path) -> Path:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "All"
    sheet.append(LEGACY_COLUMNS)

    def append(**values):
        sheet.append([values.get(column) for column in LEGACY_COLUMNS])

    append(
        Week=1,
        Sort=1,
        Type="Class",
        Title="Primeira aula",
        Description="Descrição curta.",
    )
    append(
        Week=1,
        Sort=2,
        Type="Self-study",
        Title="Artigo da primeira aula",
        **{
            "Parent class": "Primeira aula",
            "Description": "Uma fonte de teste.",
            "URL": "https://example.com/primeira",
        },
    )
    append(
        Week=2,
        Sort=3,
        Type="Class",
        Title="Segunda aula",
        Description=(
            "Uma descrição longa o bastante para ocupar várias linhas no editor. "
            "Ela representa o conteúdo real que precisa permanecer inteiramente "
            "visível durante a curadoria da aula. " * 4
        ),
    )
    append(
        Week=2,
        Sort=4,
        Type="Self-study",
        Title="Livro da segunda aula",
        **{
            "Parent class": "Segunda aula",
            "Description": "Leitura do capítulo indicado.",
            "URL": "https://integrada.minhabiblioteca.com.br/reader/books/9788521622888",
            "Resource code": "9788521622888",
        },
    )
    workbook.save(path)
    workbook.close()
    return path


def _subject_filter_workbook(path: Path) -> Path:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "All"
    sheet.append(LEGACY_COLUMNS)

    def append(**values):
        sheet.append([values.get(column) for column in LEGACY_COLUMNS])

    append(
        Week=1,
        Sort=1,
        Type="Class",
        Title="Aula de comunicação",
        Axis="COM",
    )
    append(
        Week=1,
        Sort=2,
        Type="Orientation",
        Title="Sprint Planning",
    )
    append(
        Week=1,
        Sort=3,
        Type="Deliverable",
        Title="Apresentação do artefato",
        **{"Grade weight": 3},
    )
    append(
        Week=1,
        Sort=4,
        Type="Evaluation",
        Title="Avaliação em pares",
        **{"Grade weight": 2},
    )
    workbook.save(path)
    workbook.close()
    return path


def _type2_workbook(path: Path) -> Path:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Projetos"
    sheet.append(PROJECT_COLUMNS)

    def append(**values):
        sheet.append([values.get(column) for column in PROJECT_COLUMNS])

    common = {"Projeto": "GRAD CC07", "Semana": "Semana 02"}
    append(
        **common,
        Ordem=1,
        Atividade="Programação e Desenvolvimento de Banco de Dados",
        **{
            "Tipo da atividade": "Encontro de instrução",
            "Descrição da atividade": "Criação e manipulação de bancos relacionais.",
            "Eixo": "Computação",
            "Assuntos": "Banco de dados relacional\n,SQL Básico",
        },
    )
    append(
        **common,
        Ordem=2,
        Atividade="Tutorial MySQL",
        **{
            "Tipo da atividade": "Autoestudo",
            "Encontro pai": "Programação e Desenvolvimento de Banco de Dados",
            "URL": "https://example.com/mysql",
        },
    )
    append(
        **common,
        Ordem=3,
        Atividade="Sprint Planning",
        **{"Tipo da atividade": "Encontro de orientação"},
    )
    append(
        **common,
        Ordem=4,
        Atividade="Entrega do artefato",
        **{"Tipo da atividade": "Desenvolvimento projeto"},
    )
    append(
        **common,
        Ordem=5,
        Atividade="Avaliação geral",
        **{"Tipo da atividade": "Avaliação / pesquisa"},
    )
    workbook.save(path)
    workbook.close()
    return path


@contextmanager
def _serve(app):
    sock = socket.socket()
    sock.bind(("127.0.0.1", 0))
    sock.listen(128)
    port = sock.getsockname()[1]
    server = uvicorn.Server(
        uvicorn.Config(app, log_level="error", lifespan="on")
    )
    thread = threading.Thread(
        target=server.run, kwargs={"sockets": [sock]}, daemon=True
    )
    thread.start()
    deadline = time.monotonic() + 5
    while not server.started and time.monotonic() < deadline:
        time.sleep(0.01)
    if not server.started:
        raise RuntimeError("test web server did not start")
    try:
        yield f"http://127.0.0.1:{port}"
    finally:
        server.should_exit = True
        thread.join(timeout=5)
        sock.close()


def _route_detail(page, syllabus_id, mutate):
    """Keep the real projection, replacing only the not-yet-landed KC fields."""

    cached = {"payload": None}

    def handler(route):
        if urlparse(route.request.url).path != f"/api/syllabi/{syllabus_id}":
            route.continue_()
            return
        response = route.fetch()
        payload = response.json()
        status = response.status
        if "lessons" not in payload and cached["payload"] is not None:
            payload = deepcopy(cached["payload"])
            status = 200
        mutate(payload)
        cached["payload"] = deepcopy(payload)
        route.fulfill(status=status, json=payload)

    page.route(f"**/api/syllabi/{syllabus_id}*", handler)


def _snapshot(
    source_id,
    source_title,
    artifact_id,
    *,
    completed,
    statement=None,
    task=None,
    answer=None,
):
    stages = {
        name: {"status": "done" if index < completed else "pending"}
        for index, name in enumerate(LOCAL_KC_STAGES)
    }
    components = []
    if statement:
        components.append(
            {
                "id": f"task-{source_id}",
                "kind": "singleton",
                "canonical": {"verdict": "stated", "statement": statement},
                "members": [
                    {
                        "task_id": f"task-{source_id}",
                        "source_id": source_id,
                        "task": task,
                        "answer": answer,
                        "statement": statement,
                    }
                ],
            }
        )
    return {
        "source": {
            "id": source_id,
            "title": source_title,
            "artifact_id": artifact_id,
        },
        "stages": stages,
        "components": components,
        "relationships": [],
    }


def test_upload_dialog_previews_companion_identity_and_offers_conflict_choices(
    test_database_url, applied_migrations, tmp_path
):
    workbook_path = _editable_workbook(tmp_path / "syllabus-identity.xlsx")
    marker = uuid.uuid4().hex[:10]
    name = f"Upload identity {marker}"
    occupied = f"graph-inteli-upload-identity-{marker}"
    also_occupied = f"{occupied}-reserved"
    namespace = {
        "schema_version": "companion_graph_namespace.v1",
        "institutions": [{"slug": "inteli", "name": "Inteli"}],
        "graph_ids": [occupied, also_occupied],
    }
    app = create_app(
        lambda: psycopg.connect(test_database_url),
        companion_namespace_provider=lambda: namespace,
    )

    with _serve(app) as base_url, sync_playwright() as playwright:
        browser = playwright.chromium.launch(headless=True)
        page = browser.new_page()
        page.goto(f"{base_url}/syllabi")
        page.locator("[data-new-syllabus]").first.click()

        dialog = page.locator("[data-upload-dialog]")
        expect(dialog).to_be_visible()
        expect(dialog.locator('[name="display_name"]')).to_have_count(0)
        expect(dialog.locator('[name="lesson_subject_ids"]')).to_have_count(0)
        graph_id = dialog.locator('[name="graph_id"]')
        expect(graph_id).to_be_hidden()
        assert graph_id.is_disabled()

        dialog.locator('[name="institution_id"]').select_option("inteli")
        dialog.locator('[name="name"]').fill(name)

        preview = dialog.locator("[data-graph-preview]")
        expect(preview).to_be_visible()
        expect(preview.locator("[data-graph-display-name]")).to_have_text(name)
        expect(preview.locator("[data-proposed-graph-id]")).to_have_text(occupied)

        conflict = dialog.locator("[data-graph-conflict]")
        expect(conflict).to_be_visible()
        expect(conflict).to_contain_text(occupied)

        dialog.locator('[name="name"]').fill("C")
        assert preview.evaluate("element => element.hidden") is False
        expect(preview.locator("[data-proposed-graph-id]")).to_have_text("Calculando…")
        expect(preview.locator("[data-proposed-graph-id]")).to_have_text("Continue digitando…")
        assert preview.evaluate("element => element.hidden") is False

        changed_name = f"{name} changed"
        dialog.locator('[name="name"]').fill(changed_name)
        assert preview.evaluate("element => element.hidden") is False
        expect(preview.locator("[data-graph-display-name]")).to_have_text(changed_name)
        expect(preview.locator("[data-proposed-graph-id]")).to_have_text("Calculando…")
        expect(dialog.get_by_role("button", name="Escolher outro ID")).to_be_disabled()
        assert conflict.evaluate("element => element.hidden") is True
        expect(preview).to_be_visible()
        expect(preview.locator("[data-graph-display-name]")).to_have_text(changed_name)

        dialog.locator('[name="name"]').fill(name)
        assert preview.evaluate("element => element.hidden") is False
        expect(preview.locator("[data-proposed-graph-id]")).to_have_text("Calculando…")
        expect(conflict).to_be_visible()
        expect(dialog.get_by_role("button", name="Mudar nome do syllabus")).to_have_count(0)
        dialog.get_by_role("button", name="Editar somente o ID").click()
        expect(graph_id).to_be_visible()
        expect(graph_id).to_be_enabled()
        expect(dialog.get_by_role("button", name="Salvar ID")).to_be_visible()
        expect(dialog.get_by_role("button", name="Cancelar edição")).to_be_visible()

        graph_id.fill(f"{occupied}-rascunho")
        expect(conflict).to_be_visible()
        expect(preview.locator("[data-proposed-graph-id]")).to_have_text(occupied)
        dialog.get_by_role("button", name="Cancelar edição").click()
        expect(graph_id).to_be_hidden()
        assert graph_id.is_disabled()
        expect(conflict).to_be_visible()

        dialog.get_by_role("button", name="Editar somente o ID").click()
        graph_id.fill(also_occupied)
        expect(conflict).to_be_visible()
        dialog.get_by_role("button", name="Salvar ID").click()
        expect(dialog.locator("[data-graph-id-error]")).to_contain_text("já está em uso")
        expect(conflict).to_be_visible()
        expect(graph_id).to_be_visible()

        graph_id.fill(f"{occupied}-2")
        dialog.get_by_role("button", name="Salvar ID").click()
        expect(preview.locator("[data-graph-id-status]")).to_have_text(
            "ID verificado. Será salvo ao adicionar o syllabus."
        )
        expect(graph_id).to_be_visible()
        assert graph_id.is_enabled()
        assert graph_id.evaluate("input => input.readOnly") is True
        expect(preview.locator("[data-proposed-graph-id]")).to_have_text(f"{occupied}-2")
        expect(dialog.get_by_role("button", name="Salvar ID")).to_be_hidden()
        expect(dialog.get_by_role("button", name="Cancelar edição")).to_be_hidden()
        expect(conflict).to_be_hidden()

        dialog.locator('[name="file"]').set_input_files(workbook_path)
        dialog.get_by_role("button", name="Adicionar syllabus").click()

        page.wait_for_url("**/syllabi?id=*")
        expect(page.get_by_role("button", name="Enviar nova versão")).to_be_visible()
        browser.close()


def test_upload_dialog_blocks_an_existing_syllabus_and_can_open_it(
    test_database_url, applied_migrations, tmp_path
):
    workbook_path = _editable_workbook(tmp_path / "existing-syllabus.xlsx")
    marker = uuid.uuid4().hex[:10]
    name = f"Existing syllabus {marker}"
    with psycopg.connect(test_database_url) as conn:
        conn.execute(
            "INSERT INTO institution (id, name) VALUES ('inteli', 'Inteli')"
            " ON CONFLICT (id) DO UPDATE SET name = EXCLUDED.name"
        )
        imported = import_workbook(conn, workbook_path, name, institution_id="inteli")

    namespace = {
        "schema_version": "companion_graph_namespace.v1",
        "institutions": [{"slug": "inteli", "name": "Inteli"}],
        "graph_ids": [imported["syllabus_id"]],
    }
    app = create_app(
        lambda: psycopg.connect(test_database_url),
        companion_namespace_provider=lambda: namespace,
    )

    with _serve(app) as base_url, sync_playwright() as playwright:
        browser = playwright.chromium.launch(headless=True)
        page = browser.new_page()
        page.goto(f"{base_url}/syllabi")
        page.locator("[data-new-syllabus]").first.click()
        dialog = page.locator("[data-upload-dialog]")
        dialog.locator('[name="institution_id"]').select_option("inteli")
        dialog.locator('[name="name"]').fill(name)

        conflict = dialog.locator("[data-syllabus-conflict]")
        expect(conflict).to_be_visible()
        expect(conflict).to_contain_text(name)
        expect(conflict).not_to_contain_text("mude o nome")
        expect(dialog.get_by_role("button", name="Mudar nome do syllabus")).to_have_count(0)
        expect(dialog.get_by_role("button", name="Escolher outro ID")).to_be_disabled()
        actions = conflict.locator(".syl-graph-conflict__actions")
        assert actions.evaluate("element => getComputedStyle(element).display") == "grid"
        assert actions.locator(".button").count() == 1
        open_existing = dialog.get_by_role("link", name="Abrir syllabus existente")
        assert "button--primary" not in (open_existing.get_attribute("class") or "")
        assert "button--quiet" not in (open_existing.get_attribute("class") or "")
        button_geometry = open_existing.evaluate(
            """element => {
                const style = getComputedStyle(element);
                const box = element.getBoundingClientRect();
                const range = document.createRange();
                range.selectNodeContents(element);
                const text = range.getBoundingClientRect();
                return {
                    display: style.display,
                    alignItems: style.alignItems,
                    verticalOffset: Math.abs(
                        (box.top + box.height / 2) - (text.top + text.height / 2)
                    ),
                };
            }"""
        )
        assert button_geometry["display"] == "flex"
        assert button_geometry["alignItems"] == "center"
        assert button_geometry["verticalOffset"] < 1

        dialog.locator('[name="name"]').fill(f"{name} changed")
        expect(conflict).to_be_hidden()
        dialog.locator('[name="name"]').fill(name)
        expect(conflict).to_be_visible()
        expect(dialog.get_by_role("button", name="Escolher outro ID")).to_be_disabled()
        open_existing.click()

        page.wait_for_url(f"**/syllabi?id={imported['syllabus_id']}")
        expect(page.get_by_role("button", name="Enviar nova versão")).to_be_visible()
        browser.close()


def test_targeted_lesson_editor_stays_scoped_and_can_hide_that_lesson(
    test_database_url, applied_migrations, tmp_path
):
    name = f"Browser editor {uuid.uuid4().hex[:8]}"
    with psycopg.connect(test_database_url) as conn:
        imported = import_workbook(
            conn,
            _editable_workbook(tmp_path / "editor.xlsx"),
            name,
            require_syllabus_metadata=False,
        )

    app = create_app(lambda: psycopg.connect(test_database_url))
    with _serve(app) as base_url, sync_playwright() as playwright:
        browser = playwright.chromium.launch(headless=True)
        page = browser.new_page()
        page.goto(f"{base_url}/syllabi?id={imported['syllabus_id']}")

        page.get_by_role("button", name="Editar somente esta aula").nth(1).click()

        expect(page.locator(".syl-lesson--editor")).to_have_count(1)
        expect(page.locator("[data-lesson-field='title']")).to_have_value("Segunda aula")
        expect(
            page.locator(".syl-lesson:not(.syl-lesson--editor) .syl-lesson-edit")
        ).to_have_count(1)

        page.get_by_role("button", name="Ocultar aula").click()
        expect(page.locator(".syl-lesson--editor")).to_have_class(
            "syl-lesson syl-lesson--editor is-hidden-lesson"
        )
        expect(page.get_by_role("button", name="Desocultar aula")).to_be_visible()
        save = page.get_by_role("button", name="Salvar nova versão")
        expect(save).to_be_enabled()
        save.click()

        version_dialog = page.get_by_role("dialog", name="Registrar nova versão")
        expect(version_dialog).to_be_visible()
        reason = version_dialog.get_by_label("Razão da nova versão")
        expect(reason).to_have_attribute("maxlength", "500")
        reason.fill("A aula não faz parte do percurso curricular desta oferta.")
        expect(version_dialog.get_by_text("57/500", exact=True)).to_be_visible()
        version_dialog.get_by_role("button", name="Criar versão").click()
        expect(page.locator("[data-status]")).to_contain_text("Versão 2 salva")
        page.get_by_role("button", name="Versão 2").click()
        history_dialog = page.get_by_role("dialog", name="Versões do syllabus")
        expect(history_dialog).to_contain_text(
            "A aula não faz parte do percurso curricular desta oferta."
        )
        expect(history_dialog.get_by_text("Versão aberta", exact=True)).to_be_visible()
        history_dialog.get_by_role("button", name="Abrir versão 1").click()
        expect(page.get_by_role("button", name="Versão 1")).to_be_visible()
        expect(page.get_by_role("button", name="Editar syllabus")).to_be_disabled()
        browser.close()

    with psycopg.connect(test_database_url) as conn:
        latest = get_syllabus_version(conn, imported["syllabus_id"])
    assert [lesson["hidden"] for lesson in latest["lessons"]] == [False, True]


def test_lesson_description_expands_to_show_its_content_when_editing(
    test_database_url, applied_migrations, tmp_path
):
    name = f"Browser autosize {uuid.uuid4().hex[:8]}"
    with psycopg.connect(test_database_url) as conn:
        imported = import_workbook(
            conn,
            _editable_workbook(tmp_path / "autosize.xlsx"),
            name,
            require_syllabus_metadata=False,
        )

    app = create_app(lambda: psycopg.connect(test_database_url))
    with _serve(app) as base_url, sync_playwright() as playwright:
        browser = playwright.chromium.launch(headless=True)
        page = browser.new_page()
        page.goto(f"{base_url}/syllabi?id={imported['syllabus_id']}")
        page.get_by_role("button", name="Editar somente esta aula").nth(1).click()

        textarea = page.locator("[data-lesson-field='description']")
        expect(textarea).to_have_attribute("maxlength", "4000")
        dimensions = textarea.evaluate(
            "element => ({"
            "clientHeight: element.clientHeight, "
            "scrollHeight: element.scrollHeight, "
            "inlineHeight: element.style.height"
            "})"
        )
        assert dimensions["inlineHeight"]
        assert dimensions["clientHeight"] >= dimensions["scrollHeight"]
        browser.close()


def test_book_code_can_be_copied_exactly_from_the_source_card(
    test_database_url, applied_migrations, tmp_path
):
    name = f"Browser copy {uuid.uuid4().hex[:8]}"
    with psycopg.connect(test_database_url) as conn:
        imported = import_workbook(
            conn,
            _editable_workbook(tmp_path / "copy.xlsx"),
            name,
            require_syllabus_metadata=False,
        )

    app = create_app(lambda: psycopg.connect(test_database_url))
    with _serve(app) as base_url, sync_playwright() as playwright:
        browser = playwright.chromium.launch(headless=True)
        context = browser.new_context(
            permissions=["clipboard-read", "clipboard-write"]
        )
        page = context.new_page()
        page.goto(f"{base_url}/syllabi?id={imported['syllabus_id']}")

        page.locator(".syl-lesson").nth(1).get_by_role(
            "button", name="Expandir aula Segunda aula"
        ).click()
        page.get_by_role(
            "button", name="Copiar código do livro 9788521622888"
        ).click()

        assert page.evaluate("navigator.clipboard.readText()") == "9788521622888"
        expect(page.locator("[data-status]")).to_contain_text(
            "Código do livro copiado"
        )
        browser.close()


def test_syllabus_costs_live_in_the_top_bar(
    test_database_url, applied_migrations, tmp_path
):
    name = f"Browser heading {uuid.uuid4().hex[:8]}"
    with psycopg.connect(test_database_url) as conn:
        imported = import_workbook(
            conn,
            _editable_workbook(tmp_path / "heading.xlsx"),
            name,
            require_syllabus_metadata=False,
        )

    app = create_app(lambda: psycopg.connect(test_database_url))

    def mutate_detail(payload):
        payload["usage"] = {
            "openrouter": {"cost_usd": 0.15, "calls": 2, "total_tokens": 150},
            "firecrawl": {
                "extractions": 1,
                "attempts": 2,
                "succeeded": 1,
                "failed": 0,
            },
        }

    with _serve(app) as base_url, sync_playwright() as playwright:
        browser = playwright.chromium.launch(headless=True)
        page = browser.new_page(viewport={"width": 1440, "height": 900})
        _route_detail(page, imported["syllabus_id"], mutate_detail)
        page.goto(f"{base_url}/syllabi?id={imported['syllabus_id']}")

        costs = page.locator(".admin-shell").get_by_label(
            "Consumo das fontes desta versão"
        )
        expect(costs).to_contain_text("OpenRouter")
        expect(costs).to_contain_text("US$ 0,15")
        expect(costs).to_contain_text("Firecrawl")
        expect(costs).to_contain_text("1 extração")
        expect(page.locator(".syl-view .syl-usage-strip")).to_have_count(0)

        browser.close()


def test_subject_filter_includes_curricular_kinds_without_a_subject(
    test_database_url, applied_migrations, tmp_path
):
    name = f"Browser orientation {uuid.uuid4().hex[:8]}"
    with psycopg.connect(test_database_url) as conn:
        imported = import_workbook(
            conn,
            _subject_filter_workbook(tmp_path / "orientation.xlsx"),
            name,
            require_syllabus_metadata=False,
        )

    app = create_app(lambda: psycopg.connect(test_database_url))
    with _serve(app) as base_url, sync_playwright() as playwright:
        browser = playwright.chromium.launch(headless=True)
        page = browser.new_page()
        page.goto(f"{base_url}/syllabi?id={imported['syllabus_id']}")

        subjects = page.locator("[data-filter-subject]")
        expect(subjects.get_by_role("option", name="COM")).to_have_count(1)
        expect(subjects.get_by_role("option", name="Artefatos")).to_have_count(1)
        expect(subjects.get_by_role("option", name="Avaliações")).to_have_count(1)
        expect(subjects.get_by_role("option", name="Orientação")).to_have_count(1)
        subjects.select_option(label="Artefatos")

        expect(page.locator(".syl-lesson")).to_have_count(1)
        expect(page.locator(".syl-lesson h2")).to_have_text("Apresentação do artefato")
        browser.close()


def test_type2_lesson_shows_subjects_and_its_parented_source(
    test_database_url, applied_migrations, tmp_path
):
    name = f"Browser type 2 {uuid.uuid4().hex[:8]}"
    with psycopg.connect(test_database_url) as conn:
        imported = import_workbook(
            conn,
            _type2_workbook(tmp_path / "type2.xlsx"),
            name,
            require_syllabus_metadata=False,
        )

    app = create_app(lambda: psycopg.connect(test_database_url))
    with _serve(app) as base_url, sync_playwright() as playwright:
        browser = playwright.chromium.launch(headless=True)
        page = browser.new_page()
        page.goto(f"{base_url}/syllabi?id={imported['syllabus_id']}")

        lesson = page.locator(".syl-lesson").first
        lesson.get_by_role(
            "button", name="Expandir aula Programação e Desenvolvimento de Banco de Dados"
        ).click()

        expect(lesson.get_by_text("Criação e manipulação de bancos relacionais.")).to_be_visible()
        subjects = lesson.get_by_label("Assuntos")
        expect(subjects.get_by_text("Banco de dados relacional", exact=True)).to_be_visible()
        expect(subjects.get_by_text("SQL Básico", exact=True)).to_be_visible()
        expect(lesson.get_by_role("heading", name="Tutorial MySQL")).to_be_visible()

        subject_filter = page.locator("[data-filter-subject]")
        for label in ("COM", "Orientação", "Artefatos", "Avaliações"):
            expect(subject_filter.get_by_role("option", name=label)).to_have_count(1)
        expect(page.locator('.syl-lesson[data-subject="COM"]')).to_have_css(
            "border-left-color", "rgb(39, 93, 125)"
        )
        expect(page.locator('.syl-lesson[data-subject="ORIENTAÇÃO"]')).to_have_css(
            "border-left-color", "rgb(102, 114, 72)"
        )
        browser.close()


def test_a_fully_validated_lesson_can_be_unvalidated_from_its_compact_row(
    test_database_url, applied_migrations, tmp_path
):
    name = f"Browser unvalidate {uuid.uuid4().hex[:8]}"
    with psycopg.connect(test_database_url) as conn:
        imported = import_workbook(
            conn,
            _editable_workbook(tmp_path / "unvalidate.xlsx"),
            name,
            require_syllabus_metadata=False,
        )

    app = create_app(lambda: psycopg.connect(test_database_url))
    with _serve(app) as base_url, sync_playwright() as playwright:
        browser = playwright.chromium.launch(headless=True)
        page = browser.new_page()
        page.goto(f"{base_url}/syllabi?id={imported['syllabus_id']}")

        first_lesson = page.locator(".syl-lesson").first
        first_lesson.get_by_role(
            "button", name="Expandir aula Primeira aula"
        ).click()
        first_lesson.get_by_role("button", name="Validada").click()

        expect(first_lesson).to_have_class(
            "syl-lesson is-validated is-collapsed"
        )
        unvalidate = first_lesson.get_by_role("button", name="Desvalidar")
        expect(unvalidate).to_be_visible()
        unvalidate.click()

        expect(first_lesson).to_have_class("syl-lesson is-collapsed")
        first_lesson.get_by_role(
            "button", name="Expandir aula Primeira aula"
        ).click()
        expect(first_lesson.get_by_role("button", name="Validada")).to_be_visible()
        browser.close()


def test_syllabus_opens_with_every_lesson_collapsed(
    test_database_url, applied_migrations, tmp_path
):
    name = f"Browser collapsed default {uuid.uuid4().hex[:8]}"
    with psycopg.connect(test_database_url) as conn:
        imported = import_workbook(
            conn,
            _editable_workbook(tmp_path / "collapsed-default.xlsx"),
            name,
            require_syllabus_metadata=False,
        )

    app = create_app(lambda: psycopg.connect(test_database_url))
    with _serve(app) as base_url, sync_playwright() as playwright:
        browser = playwright.chromium.launch(headless=True)
        page = browser.new_page()
        page.goto(f"{base_url}/syllabi?id={imported['syllabus_id']}")

        lessons = page.locator(".syl-lesson")
        expect(lessons).to_have_count(2)
        expect(lessons.nth(0)).to_have_class("syl-lesson is-collapsed")
        expect(lessons.nth(1)).to_have_class("syl-lesson is-collapsed")
        expect(
            lessons.nth(0).get_by_role("button", name="Expandir aula Primeira aula")
        ).to_have_attribute("aria-expanded", "false")

        lessons.nth(0).get_by_role(
            "button", name="Expandir aula Primeira aula"
        ).click()
        expect(lessons.nth(0)).to_have_class("syl-lesson")
        expect(lessons.nth(1)).to_have_class("syl-lesson is-collapsed")
        browser.close()


def test_validating_the_last_source_refreshes_the_kc_offer(
    test_database_url, applied_migrations, tmp_path
):
    name = f"Browser KC gate {uuid.uuid4().hex[:8]}"
    with psycopg.connect(test_database_url) as conn:
        imported = import_workbook(
            conn,
            _editable_workbook(tmp_path / "kc-gate.xlsx"),
            name,
            require_syllabus_metadata=False,
        )

    app = create_app(lambda: psycopg.connect(test_database_url))

    def mutate_detail(payload):
        lesson = payload["lessons"][0]
        source = lesson["sources"][0]
        validated = bool(source.get("review", {}).get("validated"))
        source["has_markdown"] = True
        source["markdown"] = {"available": True}
        lesson["knowledge"] = {
            "lesson_id": lesson["id"],
            "active_reference_count": 1,
            "publication_count": 1,
            "eligibility": {
                "eligible": validated,
                "code": "ready" if validated else "references_not_validated",
            },
            "latest_build": None,
        }

    with _serve(app) as base_url, sync_playwright() as playwright:
        browser = playwright.chromium.launch(headless=True)
        page = browser.new_page()
        _route_detail(page, imported["syllabus_id"], mutate_detail)
        page.goto(f"{base_url}/syllabi?id={imported['syllabus_id']}")

        first_lesson = page.locator(".syl-lesson").first
        first_lesson.get_by_role(
            "button", name="Expandir aula Primeira aula"
        ).click()
        expect(
            first_lesson.get_by_role("button", name="Iniciar criação")
        ).to_have_count(0)

        first_lesson.get_by_role("button", name="Validada").click()

        expect(
            first_lesson.get_by_role("button", name="Iniciar criação")
        ).to_be_visible()
        expect(first_lesson).to_contain_text(
            "1 Source Publication pronta · 1 autoestudo validado"
        )
        browser.close()


def test_sequential_lesson_edits_are_saved_as_one_syllabus_edition(
    test_database_url, applied_migrations, tmp_path
):
    name = f"Browser serial edits {uuid.uuid4().hex[:8]}"
    with psycopg.connect(test_database_url) as conn:
        imported = import_workbook(
            conn,
            _editable_workbook(tmp_path / "serial.xlsx"),
            name,
            require_syllabus_metadata=False,
        )

    app = create_app(lambda: psycopg.connect(test_database_url))
    with _serve(app) as base_url, sync_playwright() as playwright:
        browser = playwright.chromium.launch(headless=True)
        page = browser.new_page()
        page.goto(f"{base_url}/syllabi?id={imported['syllabus_id']}")

        page.get_by_role("button", name="Editar somente esta aula").first.click()
        page.locator("[data-lesson-field='title']").fill("Primeira aula revisada")
        page.locator(
            ".syl-lesson:not(.syl-lesson--editor) .syl-lesson-edit"
        ).click()
        expect(page.locator(".syl-lesson--editor")).to_have_count(1)
        page.locator("[data-lesson-field='title']").fill("Segunda aula revisada")

        page.get_by_role("button", name="Salvar nova versão").click()
        version_dialog = page.get_by_role("dialog", name="Registrar nova versão")
        version_dialog.get_by_label("Razão da nova versão").fill(
            "Revisa os títulos de duas aulas."
        )
        version_dialog.get_by_role("button", name="Criar versão").click()
        expect(page.locator("[data-status]")).to_contain_text("Versão 2 salva")
        browser.close()

    with psycopg.connect(test_database_url) as conn:
        history = get_syllabus_history(conn, imported["syllabus_id"])
        latest = get_syllabus_version(conn, imported["syllabus_id"])
    assert [version["seq"] for version in history["versions"]] == [2, 1]
    assert [lesson["title"] for lesson in latest["lessons"]] == [
        "Primeira aula revisada",
        "Segunda aula revisada",
    ]


def test_kc_entry_points_confirm_before_post_and_poll_the_active_build(
    test_database_url, applied_migrations, tmp_path
):
    name = f"Browser KC start {uuid.uuid4().hex[:8]}"
    with psycopg.connect(test_database_url) as conn:
        imported = import_workbook(
            conn,
            _editable_workbook(tmp_path / "kc-start.xlsx"),
            name,
            require_syllabus_metadata=False,
        )

    app = create_app(lambda: psycopg.connect(test_database_url))
    current = {"build": None, "source_id": None, "reference_id": None}
    posts = []
    build_reads = []

    def mutate_detail(payload):
        lesson = payload["lessons"][0]
        source = lesson["sources"][0]
        current["source_id"] = source["source_id"]
        current["reference_id"] = source["reference_id"]
        source["review"] = {"validated": True, "complexity": "simple"}
        source["has_markdown"] = True
        source["markdown"] = {"available": True}
        lesson["knowledge"] = {
            "lesson_id": lesson["id"],
            "publication_count": 1,
            "eligibility": {"eligible": True, "code": "ready"},
            "latest_build": current["build"],
        }
        # Keep a non-KC source in its ordinary pre-extraction state so this
        # regression also proves those controls remain present.
        payload["lessons"][1]["sources"][0]["source_id"] = (
            payload["lessons"][1]["sources"][0].get("source_id")
            or "source-browser-book"
        )
        payload["knowledge_manifest_id"] = None

    with _serve(app) as base_url, sync_playwright() as playwright:
        browser = playwright.chromium.launch(headless=True)
        page = browser.new_page()
        _route_detail(page, imported["syllabus_id"], mutate_detail)

        def offer_handler(route):
            lesson_id = urlparse(route.request.url).path.split("/")[-2]
            route.fulfill(
                status=200,
                json={
                    "lesson_id": lesson_id,
                    "publication_count": 1,
                    "eligibility": {"eligible": True, "code": "ready"},
                    "latest_build": current["build"],
                },
            )

        def start_handler(route):
            posts.append(route.request.post_data_json)
            source_id = current["source_id"]
            reference_id = current["reference_id"]
            current["build"] = {
                "id": "build-browser-start",
                "status": "running",
                "references": [
                    {
                        "reference_id": reference_id,
                        "work_id": "work-browser-start",
                    }
                ],
                "work": [
                    {
                        "id": "work-browser-start",
                        "source_id": source_id,
                        "artifact_id": "artifact-browser-start",
                        "reference_ids": [reference_id],
                        "status": "running",
                        "snapshot": _snapshot(
                            source_id,
                            "Artigo da primeira aula",
                            "artifact-browser-start",
                            completed=0,
                        ),
                    }
                ],
            }
            route.fulfill(status=202, json=current["build"])

        def build_handler(route):
            build_reads.append(route.request.url)
            source_id = current["source_id"]
            reference_id = current["reference_id"]
            current["build"] = {
                "id": "build-browser-start",
                "status": "running",
                "references": [
                    {
                        "reference_id": reference_id,
                        "work_id": "work-browser-start",
                    }
                ],
                "work": [
                    {
                        "id": "work-browser-start",
                        "source_id": source_id,
                        "artifact_id": "artifact-browser-start",
                        "reference_ids": [reference_id],
                        "status": "running",
                        "snapshot": _snapshot(
                            source_id,
                            "Artigo da primeira aula",
                            "artifact-browser-start",
                            completed=9,
                            statement="Explica por que feedback reduz incerteza.",
                            task="Por que ciclos curtos ajudam a equipe?",
                            answer="Eles antecipam evidência sobre a solução.",
                        ),
                    }
                ],
            }
            route.fulfill(status=200, json=current["build"])

        page.route("**/api/knowledge-builds/build-browser-start", build_handler)
        page.route("**/knowledge-builds", start_handler)
        page.route("**/knowledge", offer_handler)
        page.goto(f"{base_url}/syllabi?id={imported['syllabus_id']}")

        first_lesson = page.locator(".syl-lesson").first
        expect(first_lesson.get_by_role("button", name="Visualizar")).to_be_visible()
        expect(
            first_lesson.get_by_role(
                "button", name="Ver KCs de Artigo da primeira aula"
            )
        ).to_have_count(0)
        expect(first_lesson.get_by_role("button", name="Desvalidar")).to_be_visible()
        expect(page.get_by_role("button", name="Universo")).to_be_disabled()
        page.locator(".syl-lesson").nth(1).get_by_role(
            "button", name="Expandir aula Segunda aula"
        ).click()
        expect(
            page.locator(".syl-lesson").nth(1).get_by_role(
                "button", name="Upload de PDF ou Imagem"
            )
        ).to_be_visible()

        expect(first_lesson).to_contain_text(
            "1 Source Publication pronta · 1 autoestudo validado"
        )
        first_lesson.get_by_role("button", name="Iniciar criação").click()
        dialog = page.locator("[data-knowledge-dialog]")
        expect(dialog).to_be_visible()
        expect(dialog).to_contain_text("1 Source Publication")
        expect(dialog.get_by_role("button", name="Confirmar e iniciar KCs")).to_be_visible()
        assert posts == []

        dialog.get_by_role("button", name="Confirmar e iniciar KCs").click()
        expect(dialog).to_contain_text("0/11 etapas locais")
        assert len(posts) == 1
        assert set(posts[0]) == {"request_key"}
        assert posts[0]["request_key"].startswith("syllabi-ui:")

        expect(dialog).to_contain_text("9/11 etapas locais", timeout=5_000)
        expect(dialog).to_contain_text("Explica por que feedback reduz incerteza.")
        expect(first_lesson).to_contain_text("9/11 etapas locais concluídas")
        expect(first_lesson.get_by_role("button", name="Acompanhar")).to_be_visible()
        expect(
            first_lesson.get_by_role(
                "button", name="Ver KCs de Artigo da primeira aula"
            )
        ).to_be_visible()
        assert build_reads
        assert dialog.locator("[data-knowledge-stage-details]").evaluate(
            "element => element.open"
        ) is False
        dialog.get_by_role("button", name="Fechar").click()
        expect(first_lesson.get_by_role("button", name="Acompanhar")).to_be_focused()
        browser.close()


def test_kc_modal_filters_one_publication_and_deduplicates_lesson_work(
    test_database_url, applied_migrations, tmp_path
):
    name = f"Browser KC results {uuid.uuid4().hex[:8]}"
    with psycopg.connect(test_database_url) as conn:
        imported = import_workbook(
            conn,
            _editable_workbook(tmp_path / "kc-results.xlsx"),
            name,
            require_syllabus_metadata=False,
        )

    app = create_app(lambda: psycopg.connect(test_database_url))
    shaped = {
        "offer": None,
        "build": None,
        "summary": None,
        "source_a": None,
        "reused_lesson_id": None,
    }

    def mutate_detail(payload):
        lesson = payload["lessons"][0]
        source_a = lesson["sources"][0]
        source_a["review"] = {"validated": True, "complexity": "simple"}
        source_a["has_markdown"] = True
        source_a["markdown"] = {"available": True}
        source_b = {
            **source_a,
            "source_id": "source-browser-neighbor",
            "reference_id": "reference-browser-neighbor",
            "title": "Fonte vizinha",
            "url": "https://example.com/vizinha",
            "review": {"validated": True, "complexity": "complex"},
        }
        lesson["sources"] = [source_a, source_b]
        shaped["source_a"] = source_a
        source_a_id = source_a["source_id"]
        reference_a_id = source_a["reference_id"]
        work_a = {
            "id": "work-a",
            "source_id": source_a_id,
            "artifact_id": "artifact-a",
            "reference_ids": [reference_a_id],
            "status": "succeeded",
            "snapshot": _snapshot(
                source_a_id,
                source_a["title"],
                "artifact-a",
                completed=9,
                statement="Distingue feedback de aprovação tardia.",
                task="Qual é a função do feedback?",
                answer="Reduzir incerteza enquanto ainda é barato mudar.",
            ),
        }
        duplicate_a = {**work_a, "id": "work-a-duplicate"}
        work_b = {
            "id": "work-b",
            "source_id": source_b["source_id"],
            "artifact_id": "artifact-b",
            "reference_ids": [source_b["reference_id"]],
            "status": "succeeded",
            "snapshot": _snapshot(
                source_b["source_id"],
                source_b["title"],
                "artifact-b",
                completed=11,
                statement="Relaciona descoberta contínua e decisões reversíveis.",
                task="O que torna uma decisão reversível?",
                answer="Baixo custo de correção após nova evidência.",
            ),
        }
        shaped["build"] = {
            "id": "build-browser-results",
            "status": "succeeded",
            "references": [
                {"reference_id": reference_a_id, "work_id": "work-a"},
                {
                    "reference_id": source_b["reference_id"],
                    "work_id": "work-b",
                },
            ],
            "work": [work_a, duplicate_a, work_b],
        }
        shaped["summary"] = {
            "id": shaped["build"]["id"],
            "status": "succeeded",
            "stage_progress": {"completed": 20, "total": 22},
            "references": shaped["build"]["references"],
            "work": [
                {
                    "id": "work-a",
                    "source_id": source_a_id,
                    "artifact_id": "artifact-a",
                    "reference_ids": [reference_a_id],
                    "status": "succeeded",
                    "kc_count": 1,
                },
                {
                    "id": "work-b",
                    "source_id": source_b["source_id"],
                    "artifact_id": "artifact-b",
                    "reference_ids": [source_b["reference_id"]],
                    "status": "succeeded",
                    "kc_count": 1,
                },
            ],
        }
        shaped["offer"] = {
            "lesson_id": lesson["id"],
            "publication_count": 2,
            "eligibility": {"eligible": True, "code": "ready"},
            "latest_build": shaped["summary"],
        }
        lesson["knowledge"] = shaped["offer"]
        source_a["knowledge"] = {
            "build_id": shaped["build"]["id"],
            "work_id": "work-a",
            "status": "succeeded",
            "current": True,
            "kc_count": 1,
        }
        source_b["knowledge"] = {
            "build_id": shaped["build"]["id"],
            "work_id": "work-b",
            "status": "succeeded",
            "current": True,
            "kc_count": 1,
        }
        reused_lesson = payload["lessons"][1]
        reused_source = reused_lesson["sources"][0]
        shaped["reused_lesson_id"] = reused_lesson["id"]
        reused_source["source_id"] = source_a_id
        reused_source["review"] = {"validated": True, "complexity": "simple"}
        reused_source["has_markdown"] = True
        reused_source["markdown"] = {"available": True}
        reused_source["knowledge"] = {
            "build_id": shaped["build"]["id"],
            "work_id": "work-a",
            "status": "succeeded",
            "current": True,
            "kc_count": 1,
        }
        reused_lesson["knowledge"] = {
            "lesson_id": reused_lesson["id"],
            "publication_count": 1,
            "eligibility": {"eligible": True, "code": "ready"},
            "latest_build": None,
        }
        payload["knowledge_manifest_id"] = "manifest-browser-results"

    with _serve(app) as base_url, sync_playwright() as playwright:
        browser = playwright.chromium.launch(headless=True)
        page = browser.new_page()
        _route_detail(page, imported["syllabus_id"], mutate_detail)
        page.route(
            "**/api/knowledge-builds/build-browser-results",
            lambda route: route.fulfill(status=200, json=shaped["build"]),
        )
        def offer_handler(route):
            lesson_id = urlparse(route.request.url).path.split("/")[-2]
            offer = shaped["offer"]
            if lesson_id == shaped["reused_lesson_id"]:
                offer = {
                    "lesson_id": lesson_id,
                    "publication_count": 1,
                    "eligibility": {"eligible": True, "code": "ready"},
                    "latest_build": None,
                }
            route.fulfill(status=200, json=offer)

        page.route("**/knowledge", offer_handler)
        page.goto(f"{base_url}/syllabi?id={imported['syllabus_id']}")

        first_lesson = page.locator(".syl-lesson").first
        expect(first_lesson).to_contain_text("2 KCs unitários produzidos")
        expect(first_lesson.get_by_role("button", name="Ver KCs da aula")).to_be_visible()
        universe = page.get_by_role("link", name="Universo", exact=True)
        expect(universe).to_have_attribute(
            "href",
            (
                "/graph?manifest_id=manifest-browser-results"
                f"&syllabus_id={imported['syllabus_id']}"
                f"&version_id={quote(imported['version_id'], safe='')}"
            ),
        )

        first_lesson.get_by_role(
            "button", name="Ver KCs de Artigo da primeira aula"
        ).click()
        dialog = page.locator("[data-knowledge-dialog]")
        expect(dialog).to_be_visible()
        expect(dialog.get_by_role("heading", name="KCs · Artigo da primeira aula")).to_be_visible()
        expect(dialog).to_contain_text("Distingue feedback de aprovação tardia.")
        expect(dialog).to_contain_text("Qual é a função do feedback?")
        expect(dialog).to_contain_text(
            "Reduzir incerteza enquanto ainda é barato mudar."
        )
        expect(dialog).to_contain_text("9/11 etapas locais")
        expect(dialog).to_contain_text("Artigo da primeira aula")
        expect(dialog).not_to_contain_text("Fonte vizinha")
        expect(dialog).not_to_contain_text(
            "Relaciona descoberta contínua e decisões reversíveis."
        )
        expect(dialog.locator(".syl-kc-item")).to_have_count(1)
        expect(dialog.locator(".syl-knowledge-work")).to_have_count(1)
        assert dialog.locator("[data-knowledge-stage-details]").evaluate(
            "element => element.open"
        ) is False

        page.keyboard.press("Escape")
        source_button = first_lesson.get_by_role(
            "button", name="Ver KCs de Artigo da primeira aula"
        )
        expect(source_button).to_be_focused()
        first_lesson.get_by_role("button", name="Ver KCs da aula").click()
        expect(dialog).to_contain_text("20/22 etapas locais")
        expect(dialog).to_contain_text(
            "Relaciona descoberta contínua e decisões reversíveis."
        )
        expect(dialog.locator(".syl-kc-item")).to_have_count(2)
        expect(dialog.locator(".syl-knowledge-work")).to_have_count(2)
        dialog.get_by_role("button", name="Fechar").click()
        expect(first_lesson.get_by_role("button", name="Ver KCs da aula")).to_be_focused()

        reused_lesson = page.locator(".syl-lesson").nth(1)
        reused_lesson.get_by_role("button", name="Ver KCs de Livro da segunda aula").click()
        expect(dialog.get_by_role("heading", name="KCs · Livro da segunda aula")).to_be_visible()
        expect(dialog).to_contain_text("9/11 etapas locais")
        expect(dialog).to_contain_text("Distingue feedback de aprovação tardia.")
        expect(dialog.locator(".syl-kc-item")).to_have_count(1)
        expect(dialog.locator(".syl-knowledge-work")).to_have_count(1)
        browser.close()


def test_universe_publication_is_a_second_explicit_checkpoint_with_shared_progress(
    test_database_url, applied_migrations, tmp_path
):
    name = f"Browser corpus checkpoint {uuid.uuid4().hex[:8]}"
    with psycopg.connect(test_database_url) as conn:
        imported = import_workbook(
            conn,
            _editable_workbook(tmp_path / "corpus-checkpoint.xlsx"),
            name,
            require_syllabus_metadata=False,
        )

    app = create_app(lambda: psycopg.connect(test_database_url))
    state = {
        "build": None,
        "published_build": None,
        "reads_after_start": 0,
        "auto_complete": True,
    }
    posts = []

    def mutate_detail(payload):
        build = state["build"]
        if build is not None:
            state["reads_after_start"] += 1
            if state["auto_complete"] and state["reads_after_start"] >= 2:
                build = {
                    **build,
                    "status": "succeeded",
                    "progress": {
                        "completed": 4,
                        "total": 4,
                        "pending": 0,
                        "partial": 0,
                        "running": 0,
                        "failed": 0,
                    },
                }
                state["build"] = build
                state["published_build"] = build
        payload["knowledge"] = {
            "eligibility": {"eligible": True, "code": "ready"},
            "complete_publication_count": 2,
            "publication_count": 2,
            "latest_build": build,
            "published_build": state["published_build"],
        }
        published_build = state["published_build"]
        if published_build and published_build.get("current", True):
            payload["knowledge_manifest_id"] = published_build["manifest_id"]
        else:
            payload.pop("knowledge_manifest_id", None)

    with _serve(app) as base_url, sync_playwright() as playwright:
        browser = playwright.chromium.launch(headless=True)
        page = browser.new_page()
        _route_detail(page, imported["syllabus_id"], mutate_detail)

        def start_handler(route):
            posts.append(route.request.post_data_json)
            state["build"] = {
                "id": "syllabus-build-browser",
                "manifest_id": "manifest-browser-published",
                "status": "running",
                "current": True,
                "progress": {
                    "completed": 0,
                    "total": 4,
                    "pending": 4,
                    "partial": 0,
                    "running": 0,
                    "failed": 0,
                },
            }
            route.fulfill(status=202, json=state["build"])

        page.route("**/versions/*/knowledge-builds", start_handler)
        page.goto(f"{base_url}/syllabi?id={imported['syllabus_id']}")

        page.get_by_role("button", name="Publicar Universo").click()
        dialog = page.locator("[data-knowledge-dialog]")
        expect(dialog).to_be_visible()
        expect(dialog).to_contain_text("segundo checkpoint explícito")
        expect(dialog.get_by_role("button", name="Confirmar publicação")).to_be_visible()
        assert posts == []

        dialog.get_by_role("button", name="Confirmar publicação").click()
        expect(dialog).to_be_visible()
        expect(dialog).to_contain_text("0/4 etapas compartilhadas")
        expect(page.get_by_role("button", name="0/4 · Publicando Universo")).to_be_visible()
        assert len(posts) == 1
        assert set(posts[0]) == {"request_key"}

        universe = page.get_by_role("link", name="Universo", exact=True)
        expect(universe).to_be_visible(timeout=7_000)
        expected_href = (
            "/graph?manifest_id=manifest-browser-published"
            f"&syllabus_id={imported['syllabus_id']}"
            f"&version_id={quote(imported['version_id'], safe='')}"
        )
        expect(universe).to_have_attribute("href", expected_href)
        expect(dialog).to_contain_text("Universo publicado")
        expect(dialog).to_contain_text("4/4 etapas compartilhadas")
        expect(dialog.get_by_role("link", name="Abrir Universo")).to_have_attribute(
            "href", expected_href
        )
        dialog.locator("[data-knowledge-close]").first.click()
        expect(universe).to_be_focused()

        state["auto_complete"] = False
        state["reads_after_start"] = 0
        state["build"] = {
            "id": "syllabus-build-browser-republish",
            "manifest_id": "manifest-browser-republish",
            "status": "queued",
            "current": True,
            "progress": {
                "completed": 0,
                "total": 4,
                "pending": 4,
                "partial": 0,
                "running": 0,
                "failed": 0,
            },
        }
        page.reload()
        expect(page.get_by_role("link", name="Universo", exact=True)).to_have_attribute(
            "href", expected_href
        )
        republishing = page.get_by_role("button", name="0/4 · Publicando Universo")
        expect(republishing).to_be_visible()
        republishing.click()
        expect(dialog).to_contain_text("0/4 etapas compartilhadas")
        expect(dialog.get_by_role("link", name="Abrir Universo publicado")).to_have_attribute(
            "href", expected_href
        )
        dialog.locator("[data-knowledge-close]").first.click()

        state["build"] = {**state["build"], "status": "failed"}
        state["published_build"] = {
            **state["published_build"],
            "current": False,
        }
        page.reload()
        previous = page.get_by_role("link", name="Universo anterior")
        expect(previous).to_have_attribute("href", expected_href)
        expect(
            page.get_by_role("button", name="Publicar novo Universo")
        ).to_be_visible()
        browser.close()
