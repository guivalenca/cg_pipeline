"""Versioned syllabus intake and its small domain-facing read interface.

A syllabus is named by a person. An Adalove observer XLSX export is the one
input adapter that authors a complete immutable version of it::

    syllabus -> version -> lesson -> source reference -> source

Import never queues acquisition.  A removed reference simply disappears from
the next version; the older version and every Source/Snapshot/Artifact fact
remain in their append-only ledgers (ADRs 0001 and 0006).
"""

from __future__ import annotations

import argparse
import hashlib
import json
import re
import unicodedata
from datetime import date, datetime
from io import BytesIO
from pathlib import Path
from urllib.parse import parse_qsl, unquote_plus, urlencode, urlparse, urlunparse

import psycopg
from openpyxl import Workbook, load_workbook
from psycopg.types.json import Jsonb

from universe.db import connect
from universe.graph_identity import (
    GraphIdConflict,
    graph_id_for,
    slug_component,
    validate_graph_id,
)

XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
HIDDEN_COLUMN = "Hidden"

ADALOVE_ACTIVITY_COLUMNS = (
    "Activity order",
    "Week",
    "Type",
    "Original label",
    "Title",
    "Description",
    "Date",
    "Professor",
    "Professor UUID",
    "Assistant",
    "Assistant UUID",
    "Lesson Subject code",
    "Related subjects",
    "Primary URL",
    "Resource code",
    "Study question",
    "Study answer / rubric",
    "Required",
    "Grade weight",
    "Duration minutes",
    "Self-study schedule",
    "Exam",
    "Makeup exam",
    "Parent activity UUID",
    "Parent title",
    "Parent date",
    "Parent inference",
    "Activity UUID",
    "Folder UUID",
    "Section UUID",
    "Active",
    "Detail error",
)
ADALOVE_SUBJECT_COLUMNS = (
    "Activity order",
    "Week",
    "Activity UUID",
    "Activity title",
    "Lesson Subject code",
    "Subject UUID",
    "Related subject",
)
ADALOVE_MATERIAL_COLUMNS = (
    "Activity order",
    "Week",
    "Activity UUID",
    "Activity title",
    "Label",
    "URL",
    "Source",
    "Source path",
    "Resource code",
    "Video",
)
ADALOVE_ORDER_AUDIT_COLUMNS = (
    "Activity order",
    "Week",
    "Order key",
    "Duplicate order key",
    "Missing orders in week",
    "Activity UUID",
    "Folder UUID",
    "Type",
    "Title",
    "Parent inference",
    "Detail error",
)
ADALOVE_SHEETS = {
    "Activities": ADALOVE_ACTIVITY_COLUMNS,
    "Subjects": ADALOVE_SUBJECT_COLUMNS,
    "Materials": ADALOVE_MATERIAL_COLUMNS,
    "Order audit": ADALOVE_ORDER_AUDIT_COLUMNS,
}


class SyllabusAlreadyExists(ValueError):
    """A create request resolved to an existing Syllabus."""

    def __init__(self, syllabus_id: str, title: str, graph_id: str | None) -> None:
        self.syllabus_id = syllabus_id
        self.title = title
        self.graph_id = graph_id
        super().__init__(
            "Este nome já existe. Você está adicionando uma versão a esse syllabus."
        )
LESSON_SUBJECT_CODES = {
    "com": "COM",
    "computacao": "COM",
    "lid": "LID",
    "lideranca": "LID",
    "neg": "NEG",
    "negocios": "NEG",
    "uex": "UEX",
    "user experience": "UEX",
    "mtf": "MTF",
    "matematica": "MTF",
    "matematica e fisica": "MTF",
}
LESSON_SUBJECT_NAMES = {
    "COM": "Computação",
    "LID": "Liderança",
    "NEG": "Negócios",
    "UEX": "User Experience",
    "MTF": "Matemática",
}
ADALOVE_ACTIVITY_TYPES = {"Class", "Orientation", "Self-study", "Deliverable", "Evaluation"}
BOOK_SCOPE = re.compile(
    r"(?P<label>cap[ií]tulos?|cap\.?|chapters?|p[aá]ginas?|p[aá]gs?\.?|pag(?:es?)?\.?|"
    r"p\.|pages?|unidades?|units?|exerc[ií]cios?|exercises?)"
    r"\s*(?:n(?:[º°o]|ro)?\.?\s*)?"
    r"(?P<value>\d+(?:\s*(?:[-–—]|a|à|at[eé]|to)\s*\d+)?"
    r"(?:\s*[,;]\s*\d+(?:\s*(?:[-–—]|a|à|at[eé]|to)\s*\d+)?)*)",
    re.IGNORECASE,
)
RESOURCE_CODE = re.compile(
    r"(?:reader/)?books?/([0-9Xx-]{8,})|(?:^|/)books?/([0-9Xx-]{8,})",
    re.IGNORECASE,
)
ISBN = re.compile(r"\bISBN\s*[:#]?\s*([0-9Xx\-\s]{10,22})", re.IGNORECASE)
PAGE_ID = re.compile(r"/pageid/(\d+)", re.IGNORECASE)
YOUTUBE_VIDEO_ID = re.compile(r"^[A-Za-z0-9_-]{11}$")
YOUTUBE_TEXTUAL_QUERY = re.compile(
    r"^(?P<video_id>[A-Za-z0-9_-]{11})"
    r"(?:\s+e\s+(?:list|index|t)=[^\s]+)+$",
    re.IGNORECASE,
)
SYLLABUS_ID = re.compile(r"^[a-z][a-z0-9_.-]{1,127}$")


def slugify(text: str) -> str:
    """Return a lowercase ASCII slug suitable for a stable syllabus id."""
    return slug_component(text)


def validate_syllabus_id(value: str) -> str:
    """Return a valid durable Syllabus id or reject it with the UI contract."""
    syllabus_id = str(value or "").strip()
    if SYLLABUS_ID.fullmatch(syllabus_id) is None:
        raise ValueError(
            "O identificador do syllabus deve começar com letra minúscula, ter "
            "de 2 a 128 caracteres e usar apenas letras minúsculas, números, "
            "ponto, hífen ou sublinhado."
        )
    return syllabus_id


def canonical_url(url: str) -> str:
    """Canonicalize an article URL without retaining tracking or fragments."""
    raw = (url or "").strip()
    if not raw:
        return ""
    parsed = urlparse(raw)
    if parsed.scheme.lower() not in {"http", "https"} or not parsed.hostname:
        return ""
    query = [
        (key, value)
        for key, value in parse_qsl(parsed.query, keep_blank_values=True)
        if not unquote_plus(key).lower().startswith("utm_")
    ]
    return urlunparse(
        (
            parsed.scheme.lower(),
            parsed.netloc.lower(),
            parsed.path,
            parsed.params,
            urlencode(query, doseq=True),
            "",
        )
    )


def media_type(url: str, resource_code: str | None = None) -> str:
    """Classify an assigned material without trusting its gateway URL alone."""
    if _clean_code(resource_code):
        return "book"
    raw = (url or "").strip()
    host = (urlparse(raw).hostname or "").lower()
    if any(
        host == candidate or host.endswith(f".{candidate}")
        for candidate in ("youtube.com", "youtu.be", "vimeo.com", "ted.com")
    ):
        return "video"
    if (
        host == "integrada.minhabiblioteca.com.br"
        or host.endswith(".integrada.minhabiblioteca.com.br")
        or "sophia" in raw.lower()
    ):
        return "book"
    return "article"


def extract_book_scope(text: str, url: str = "") -> tuple[str, str] | None:
    """Extract one explicit book scope into a stable ``(kind, value)`` pair."""
    match = BOOK_SCOPE.search(text or "")
    if match:
        label = _ascii(match.group("label")).lower().rstrip(".")
        if label.startswith(("cap", "chapter")):
            kind = "chapters"
        elif label.startswith(("unidade", "unit")):
            kind = "units"
        elif label.startswith(("exercicio", "exercise")):
            kind = "exercises"
        else:
            kind = "pages"
        return kind, _normalize_scope_value(match.group("value"))
    page = PAGE_ID.search(url or "")
    if page:
        return "pages", page.group(1)
    return None


def book_scope_missing(item: dict) -> bool:
    """Say whether a linked book lacks an explicit page/chapter/unit scope."""
    resource_code = item.get("resource_code")
    if media_type(item.get("url") or "", resource_code) != "book":
        return False
    if item.get("scope_kind") and item.get("scope_value"):
        return False
    text = f"{item.get('title') or ''}\n{item.get('description') or ''}"
    return extract_book_scope(text, item.get("url") or "") is None


def source_identity(
    url: str,
    *,
    media_kind: str | None = None,
    resource_code: str | None = None,
    scope_kind: str | None = None,
    scope_value: str | None = None,
) -> dict | None:
    """Return the stable identity of a logical assigned source.

    Articles use canonical URL, videos use provider + provider id, and books
    use resource code + normalized assigned scope.  An incomplete book has no
    logical Source yet: its syllabus reference remains visible and actionable.
    """
    kind = media_kind or media_type(url, resource_code)
    if kind == "book":
        code = _clean_code(resource_code) or _resource_code_from_url(url)
        if not code or not scope_kind or not scope_value:
            return None
        return {
            "kind": "book",
            "resource_code": code,
            "scope": {"kind": scope_kind, "value": _normalize_scope_value(scope_value)},
        }
    if kind == "video":
        provider, video_id = _video_identity(url)
        if not video_id:
            return None
        return {"kind": "video", "provider": provider, "video_id": video_id}
    canonical = canonical_url(url)
    return {"kind": "article", "canonical_url": canonical} if canonical else None


def youtube_video_id(value: object) -> str | None:
    """Return one real YouTube id, repairing the XLSX textual-query form."""
    candidate = unquote_plus(str(value or "")).strip()
    if YOUTUBE_VIDEO_ID.fullmatch(candidate):
        return candidate
    textual_query = YOUTUBE_TEXTUAL_QUERY.fullmatch(candidate)
    return textual_query.group("video_id") if textual_query else None


def _video_identity(url: str) -> tuple[str, str | None]:
    parsed = urlparse((url or "").strip())
    host = (parsed.hostname or "").lower().removeprefix("www.")
    parts = [part for part in parsed.path.split("/") if part]
    if host == "youtu.be":
        return "youtube", youtube_video_id(parts[0]) if parts else None
    if host.endswith("youtube.com"):
        query = dict(parse_qsl(parsed.query))
        if query.get("v"):
            return "youtube", youtube_video_id(query["v"])
        if len(parts) >= 2 and parts[0] in {"embed", "shorts", "live"}:
            return "youtube", youtube_video_id(parts[1])
        return "youtube", None
    if host.endswith("vimeo.com"):
        video_id = next((part for part in reversed(parts) if part.isdigit()), None)
        return "vimeo", video_id
    if host.endswith("ted.com"):
        return "ted", "/".join(parts) or None
    return host or "video", canonical_url(url) or None


def _resource_code_from_url(url: str) -> str | None:
    parsed = urlparse((url or "").strip())
    searchable = f"{parsed.path}/{parsed.fragment}"
    match = RESOURCE_CODE.search(searchable)
    if not match:
        return None
    return _clean_code(next(group for group in match.groups() if group))


def _resource_code(url: str, explicit: str | None, text: str) -> str | None:
    code = _clean_code(explicit) or _resource_code_from_url(url)
    if code:
        return code
    match = ISBN.search(text or "")
    return _clean_code(match.group(1)) if match else None


def _clean_code(value: str | None) -> str | None:
    cleaned = re.sub(r"[^0-9Xx]", "", str(value or ""))
    return cleaned.upper() or None


def _normalize_scope_value(value: str) -> str:
    normalized = re.sub(r"\s*(?:[-–—]|\ba\b|\bà\b|\bat[eé]\b|\bto\b)\s*", "-", value, flags=re.I)
    normalized = re.sub(r"\s*([,;])\s*", r"\1", normalized)
    return normalized.strip()


def _ascii(value: str) -> str:
    return unicodedata.normalize("NFKD", value).encode("ascii", "ignore").decode()


def _text(value) -> str | None:
    if value is None:
        return None
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    result = str(value).strip()
    return result or None


def _lesson_subject(value: object, *, row_number: int) -> str | None:
    """Translate an Adalove Eixo value to the shared Lesson Subject code."""
    subject = _text(value)
    if not subject:
        return None
    code = LESSON_SUBJECT_CODES.get(_ascii(subject).casefold())
    if code is None:
        accepted = ", ".join(
            f"{code} {name}" for code, name in LESSON_SUBJECT_NAMES.items()
        )
        raise ValueError(
            f"A linha {row_number} da aba Activities tem o Eixo {subject!r} na "
            "coluna 'Lesson Subject code'. O Eixo identifica a área curricular "
            f"da aula. Use um destes valores: {accepted}."
        )
    return code


def _adalove_activity_type(value: object, *, row_number: int) -> str:
    kind = _text(value) or ""
    if kind not in ADALOVE_ACTIVITY_TYPES:
        raise ValueError(
            f"A linha {row_number} da aba Activities tem o tipo {kind!r}. "
            "A coluna 'Type' aceita Class, Orientation, Self-study, "
            "Deliverable ou Evaluation."
        )
    return kind


def parse_subjects(value: object) -> list[str]:
    """Turn one submitted subjects value into an ordered list."""
    values = value if isinstance(value, (list, tuple)) else [value]
    subjects: list[str] = []
    for raw in values:
        for line in re.split(r"[\r\n]+", str(raw or "")):
            subject = re.sub(r"^\s*,\s*", "", line).strip()
            if subject and subject not in subjects:
                subjects.append(subject)
    return subjects


def _as_int(value, *, row_number: int, column: str) -> int:
    text = _text(value)
    try:
        return int(float(text or ""))
    except ValueError as exc:
        raise ValueError(
            f"A linha {row_number} tem {text!r} na coluna '{column}'. "
            "Use um número inteiro nessa coluna."
        ) from exc


def _as_date(value) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = _text(value)
    if not text:
        return None
    for fmt in ("%d/%m/%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            pass
    return None


def _header(sheet) -> tuple[str | None, ...]:
    return tuple(_text(value) for value in next(sheet.iter_rows(max_row=1, values_only=True), ()))


def _require_columns(header: tuple[str | None, ...], required: tuple[str, ...], label: str) -> None:
    missing = [column for column in required if column not in header]
    if missing:
        raise ValueError(
            f"A aba '{label}' não é a exportação completa do observador do "
            f"Adalove. Faltam estas colunas: {', '.join(missing)}. Gere uma nova "
            "planilha com tools/adalove_observer_export.js."
        )


def _sheet_records(sheet, required: tuple[str, ...]) -> list[dict]:
    header = _header(sheet)
    _require_columns(header, required, sheet.title)
    records = []
    rows = sheet.iter_rows(values_only=True)
    next(rows, None)
    for row_number, raw in enumerate(rows, start=2):
        if not any(_text(value) for value in raw):
            continue
        records.append(
            {
                "row_number": row_number,
                "fields": {column: _text(value) for column, value in zip(header, raw)},
            }
        )
    return records


def _read_me(workbook) -> dict[str, str | None]:
    if "Read me" not in workbook.sheetnames:
        return {}
    rows = workbook["Read me"].iter_rows(values_only=True)
    next(rows, None)
    return {
        _text(field) or "": _text(value)
        for field, value, *_ in rows
        if _text(field)
    }


def _week_order(fields: dict, *, row_number: int, sheet: str) -> int:
    value = fields.get("Week order") or fields.get("Week")
    text = _text(value) or ""
    match = re.fullmatch(r"Semana\s+(\d+)", text, re.IGNORECASE)
    if match:
        return int(match.group(1))
    return _as_int(value, row_number=row_number, column=f"{sheet}: Week")


def _yes(value: object) -> bool:
    return (_text(value) or "").casefold() in {"1", "true", "yes", "sim"}


def _as_bool(value: object) -> bool:
    return (_text(value) or "").casefold() in {
        "1", "true", "yes", "sim", "hidden", "oculta", "oculto"
    }


def _adalove_fields(
    activity: dict,
    *,
    subjects: list[dict],
    audit: dict,
    material: dict | None = None,
) -> dict:
    fields = {
        "adalove_activity": dict(activity),
        "adalove_subjects": [dict(row) for row in subjects],
        "adalove_order_audit": dict(audit),
    }
    if material is not None:
        fields["adalove_material"] = dict(material)
    return fields


def parse_workbook(path: str | Path) -> dict:
    """Parse one full-fidelity Adalove observer export."""
    try:
        workbook = load_workbook(Path(path), read_only=True, data_only=True)
    except Exception as exc:
        raise ValueError(
            "Não foi possível abrir a planilha. Envie um arquivo .xlsx gerado por "
            "tools/adalove_observer_export.js."
        ) from exc
    try:
        missing_sheets = [name for name in ADALOVE_SHEETS if name not in workbook.sheetnames]
        if missing_sheets:
            raise ValueError(
                "Esta planilha não é uma exportação completa do observador do "
                f"Adalove. Faltam estas abas: {', '.join(missing_sheets)}. Gere uma "
                "nova planilha com tools/adalove_observer_export.js."
            )
        activity_required = tuple(
            column
            for column in ADALOVE_ACTIVITY_COLUMNS
            if column not in {"Original label", "Self-study schedule"}
        )
        activities = _sheet_records(workbook["Activities"], activity_required)
        subjects = _sheet_records(workbook["Subjects"], ADALOVE_SUBJECT_COLUMNS)
        materials = _sheet_records(workbook["Materials"], ADALOVE_MATERIAL_COLUMNS)
        audits = _sheet_records(workbook["Order audit"], ADALOVE_ORDER_AUDIT_COLUMNS)
        return _assemble_adalove(activities, subjects, materials, audits, _read_me(workbook))
    finally:
        workbook.close()


def _assemble_adalove(
    activity_records: list[dict],
    subject_records: list[dict],
    material_records: list[dict],
    audit_records: list[dict],
    workbook_metadata: dict,
) -> dict:
    activities: dict[str, dict] = {}
    for record in activity_records:
        fields = record["fields"]
        row_number = record["row_number"]
        activity_uuid = fields.get("Activity UUID") or ""
        if not activity_uuid:
            raise ValueError(
                f"A linha {row_number} da aba Activities não tem 'Activity UUID'. "
                "Gere novamente a exportação antes de enviar."
            )
        if activity_uuid in activities:
            raise ValueError(
                f"A aba Activities repete o Activity UUID {activity_uuid!r} nas "
                f"linhas {activities[activity_uuid]['row_number']} e {row_number}."
            )
        folder_uuid = fields.get("Folder UUID") or ""
        if not folder_uuid:
            raise ValueError(
                f"A linha {row_number} da aba Activities não tem 'Folder UUID'. "
                "Gere novamente a exportação antes de enviar."
            )
        title = fields.get("Title") or ""
        if not title:
            raise ValueError(
                f"A linha {row_number} da aba Activities não tem título. "
                "Corrija a atividade no Adalove e gere outra exportação."
            )
        activity_type = _adalove_activity_type(fields.get("Type"), row_number=row_number)
        detail_error = fields.get("Detail error")
        if detail_error:
            raise ValueError(
                f"A linha {row_number} da aba Activities está incompleta porque o "
                f"Adalove respondeu com este erro: {detail_error}. Gere a exportação "
                "novamente; se o erro continuar, corrija essa atividade antes do upload."
            )
        week_order = _week_order(fields, row_number=row_number, sheet="Activities")
        activity_order = _as_int(
            fields.get("Activity order"), row_number=row_number, column="Activity order"
        )
        lesson_date = _as_date(fields.get("Date"))
        if fields.get("Date") and lesson_date is None:
            raise ValueError(
                f"A linha {row_number} da aba Activities tem a data "
                f"{fields.get('Date')!r}. Use DD/MM/AAAA ou AAAA-MM-DD."
            )
        activities[activity_uuid] = {
            "row_number": row_number,
            "fields": fields,
            "activity_uuid": activity_uuid,
            "folder_uuid": folder_uuid,
            "week_order": week_order,
            "activity_order": activity_order,
            "kind": activity_type,
            "title": title,
            "description": fields.get("Description"),
            "lesson_date": lesson_date,
            "parent_activity_uuid": fields.get("Parent activity UUID"),
            "parent_inference": fields.get("Parent inference"),
        }

    subjects_by_activity: dict[str, list[dict]] = {key: [] for key in activities}
    for record in subject_records:
        fields = record["fields"]
        activity_uuid = fields.get("Activity UUID") or ""
        if activity_uuid not in activities:
            raise ValueError(
                f"A linha {record['row_number']} da aba Subjects aponta para o "
                f"Activity UUID desconhecido {activity_uuid!r}."
            )
        subject = fields.get("Related subject")
        if not subject:
            raise ValueError(
                f"A linha {record['row_number']} da aba Subjects não informa o "
                "assunto relacionado."
            )
        subjects_by_activity[activity_uuid].append(fields)

    materials_by_activity: dict[str, list[dict]] = {key: [] for key in activities}
    for record in material_records:
        fields = record["fields"]
        activity_uuid = fields.get("Activity UUID") or ""
        if activity_uuid not in activities:
            raise ValueError(
                f"A linha {record['row_number']} da aba Materials aponta para o "
                f"Activity UUID desconhecido {activity_uuid!r}."
            )
        url = fields.get("URL") or ""
        if not url:
            raise ValueError(
                f"A linha {record['row_number']} da aba Materials não tem URL. "
                "Corrija o material no Adalove e gere outra exportação."
            )
        materials_by_activity[activity_uuid].append(fields)

    audits_by_activity: dict[str, dict] = {}
    for record in audit_records:
        fields = record["fields"]
        activity_uuid = fields.get("Activity UUID") or ""
        if activity_uuid not in activities:
            raise ValueError(
                f"A linha {record['row_number']} da aba Order audit aponta para o "
                f"Activity UUID desconhecido {activity_uuid!r}."
            )
        if activity_uuid in audits_by_activity:
            raise ValueError(f"A aba Order audit repete o Activity UUID {activity_uuid!r}.")
        if fields.get("Detail error"):
            raise ValueError(
                f"A linha {record['row_number']} da aba Order audit registra este "
                f"erro de leitura: {fields['Detail error']}. Gere a exportação novamente."
            )
        if _yes(fields.get("Duplicate order key")):
            raise ValueError(
                f"A linha {record['row_number']} da aba Order audit marca a chave "
                f"{fields.get('Order key')!r} como duplicada. Corrija a ordem no "
                "Adalove e gere outra exportação."
            )
        activity = activities[activity_uuid]
        audit_week = _week_order(
            fields, row_number=record["row_number"], sheet="Order audit"
        )
        audit_order = _as_int(
            fields.get("Activity order"),
            row_number=record["row_number"],
            column="Order audit: Activity order",
        )
        expected_order_key = f"{audit_week}:{audit_order}"
        if (
            audit_week != activity["week_order"]
            or audit_order != activity["activity_order"]
            or fields.get("Order key") != expected_order_key
            or fields.get("Folder UUID") != activity["folder_uuid"]
        ):
            raise ValueError(
                f"A linha {record['row_number']} da aba Order audit não corresponde "
                f"à atividade {activity_uuid!r} da aba Activities. Esperado: semana "
                f"{activity['week_order']}, ordem {activity['activity_order']}, chave "
                f"{activity['week_order']}:{activity['activity_order']} e Folder UUID "
                f"{activity['folder_uuid']!r}. Gere outra exportação."
            )
        audits_by_activity[activity_uuid] = fields
    missing_audits = sorted(set(activities) - set(audits_by_activity))
    if missing_audits:
        raise ValueError(
            "A aba Order audit não registra todas as atividades. Faltam estes "
            f"Activity UUIDs: {', '.join(missing_audits[:5])}. Gere outra exportação."
        )

    ordered = sorted(
        activities.values(),
        key=lambda item: (item["week_order"], item["activity_order"], item["activity_uuid"]),
    )
    orientation_ids = {
        activity["activity_uuid"] for activity in ordered if activity["kind"] == "Orientation"
    }
    dropped = []
    lessons_by_activity: dict[str, dict] = {}
    for activity in ordered:
        if activity["kind"] == "Orientation":
            dropped.append(
                {
                    "activity_uuid": activity["activity_uuid"],
                    "type": activity["kind"],
                    "title": activity["title"],
                    "reason": "orientation",
                }
            )
            continue
        if activity["kind"] == "Self-study":
            continue
        if activity.get("parent_activity_uuid"):
            raise ValueError(
                f"A linha {activity['row_number']} da aba Activities liga "
                f"{activity['kind']} {activity['title']!r} a uma atividade pai. "
                "Somente Self-study recebe pai no exportador."
            )
        subject = _lesson_subject(
            activity["fields"].get("Lesson Subject code"),
            row_number=activity["row_number"],
        )
        if activity["kind"] == "Class" and subject is None:
            raise ValueError(
                f"A linha {activity['row_number']} da aba Activities descreve uma "
                "Class sem 'Lesson Subject code'. Informe o Eixo no Adalove e exporte novamente."
            )
        subject_rows = subjects_by_activity[activity["activity_uuid"]]
        lesson = {
            "week": activity["week_order"],
            "seq": activity["activity_order"],
            "week_order": activity["week_order"],
            "activity_order": activity["activity_order"],
            "activity_uuid": activity["activity_uuid"],
            "folder_uuid": activity["folder_uuid"],
            "kind": activity["kind"],
            "title": activity["title"],
            "subject": subject,
            "subjects": list(dict.fromkeys(row["Related subject"] for row in subject_rows)),
            "lesson_date": activity["lesson_date"],
            "description": activity["description"],
            "is_hidden": _as_bool(activity["fields"].get(HIDDEN_COLUMN)),
            "fields": _adalove_fields(
                activity["fields"],
                subjects=subject_rows,
                audit=audits_by_activity[activity["activity_uuid"]],
            ),
            "row_number": activity["row_number"],
            "source_references": [],
        }
        lessons_by_activity[activity["activity_uuid"]] = lesson

    for activity in ordered:
        if activity["kind"] not in {"Self-study", "Deliverable"}:
            continue
        parent_uuid = None
        inference = None
        if activity["kind"] == "Self-study":
            _lesson_subject(
                activity["fields"].get("Lesson Subject code"),
                row_number=activity["row_number"],
            )
            parent_uuid = activity.get("parent_activity_uuid") or ""
            inference = activity.get("parent_inference") or ""
            if not parent_uuid or not inference:
                raise ValueError(
                    f"A linha {activity['row_number']} da aba Activities descreve o "
                    f"Self-study {activity['title']!r} sem pai inferido e identificado. "
                    "Gere outra exportação e confira a ordem das atividades da semana."
                )
            if parent_uuid in orientation_ids:
                dropped.append(
                    {
                        "activity_uuid": activity["activity_uuid"],
                        "type": activity["kind"],
                        "title": activity["title"],
                        "parent_activity_uuid": parent_uuid,
                        "parent_inference": inference,
                        "reason": "parent_orientation",
                    }
                )
                continue
            lesson = lessons_by_activity.get(parent_uuid)
            if lesson is None or lesson["kind"] != "Class":
                raise ValueError(
                    f"A linha {activity['row_number']} da aba Activities liga o "
                    f"Self-study {activity['title']!r} ao Activity UUID {parent_uuid!r}, "
                    "mas esse pai não é uma Class da planilha."
                )
        else:
            lesson = lessons_by_activity[activity["activity_uuid"]]
        subject_rows = subjects_by_activity[activity["activity_uuid"]]
        activity_materials = materials_by_activity[activity["activity_uuid"]]
        if not activity_materials:
            if activity["kind"] == "Deliverable":
                continue
            activity_materials = [
                {
                    "Activity order": str(activity["activity_order"]),
                    "Week": str(activity["week_order"]),
                    "Activity UUID": activity["activity_uuid"],
                    "Activity title": activity["title"],
                    "Label": activity["title"],
                    "URL": activity["fields"].get("Primary URL"),
                    "Source": "activity_without_material_row",
                    "Source path": "Primary URL",
                    "Resource code": activity["fields"].get("Resource code"),
                    "Video": None,
                }
            ]
        for material_index, material in enumerate(activity_materials, 1):
            url = material.get("URL") or activity["fields"].get("Primary URL") or ""
            title = material.get("Label") or activity["title"]
            code = _resource_code(
                url,
                material.get("Resource code") or activity["fields"].get("Resource code"),
                f"{title}\n{activity['description'] or ''}",
            )
            kind = "video" if _yes(material.get("Video")) else media_type(url, code)
            scope = (
                extract_book_scope(f"{title}\n{activity['description'] or ''}", url)
                if kind == "book"
                else None
            )
            lesson["source_references"].append(
                {
                    "seq": activity["activity_order"],
                    "week_order": activity["week_order"],
                    "activity_order": activity["activity_order"],
                    "activity_uuid": activity["activity_uuid"],
                    "folder_uuid": activity["folder_uuid"],
                    "parent_activity_uuid": parent_uuid,
                    "parent_inference": inference,
                    "material_index": material_index,
                    "title": title,
                    "description": activity["description"],
                    "url": url or None,
                    "media_type": kind,
                    "resource_code": code,
                    "scope_kind": scope[0] if scope else None,
                    "scope_value": scope[1] if scope else None,
                    "is_hidden": _as_bool(activity["fields"].get(HIDDEN_COLUMN)),
                    "fields": _adalove_fields(
                        activity["fields"],
                        subjects=subject_rows,
                        audit=audits_by_activity[activity["activity_uuid"]],
                        material=material,
                    ),
                    "row_number": activity["row_number"],
                }
            )

    lessons = sorted(
        lessons_by_activity.values(),
        key=lambda item: (item["week_order"], item["activity_order"], item["activity_uuid"]),
    )
    if not lessons:
        raise ValueError(
            "A exportação não tem nenhuma Class, Deliverable ou Evaluation para armazenar."
        )
    for lesson in lessons:
        lesson["source_references"].sort(
            key=lambda item: (item["activity_order"], item["activity_uuid"], item["material_index"])
        )
    orientation_count = sum(item["reason"] == "orientation" for item in dropped)
    child_count = sum(item["reason"] == "parent_orientation" for item in dropped)
    return {
        "format": "adalove-observer",
        "workbook_title": workbook_metadata.get("Project"),
        "workbook_metadata": workbook_metadata,
        "lessons": lessons,
        "lesson_count": len(lessons),
        "source_count": sum(len(lesson["source_references"]) for lesson in lessons),
        "dropped": dropped,
        "dropped_summary": {
            "orientation_count": orientation_count,
            "orientation_self_study_count": child_count,
            "total_count": len(dropped),
        },
    }


# --- Database write/read interface -------------------------------------------------


def next_curation_event_id(conn: psycopg.Connection) -> str:
    """Allocate the next append-only curation event id."""
    conn.execute("LOCK TABLE curation_event IN SHARE ROW EXCLUSIVE MODE")
    number = conn.execute(
        "SELECT coalesce(max(substring(id from '^ce([0-9]+)$')::bigint), 0) + 1"
        " FROM curation_event WHERE id ~ '^ce[0-9]+$'"
    ).fetchone()[0]
    return f"ce{number:04d}"


def _version_counts(conn: psycopg.Connection, version_id: str) -> tuple[int, int]:
    """Return lesson and source-reference counts for a complete version."""
    return conn.execute(
        "SELECT"
        " (SELECT count(*) FROM syllabus_lesson WHERE version_id = %s),"
        " (SELECT count(*) FROM syllabus_source_reference WHERE version_id = %s)",
        (version_id, version_id),
    ).fetchone()


def resolve_source(
    conn: psycopg.Connection,
    url: str,
    title: str,
    *,
    media_kind: str | None = None,
    resource_code: str | None = None,
    scope_kind: str | None = None,
    scope_value: str | None = None,
) -> tuple[str | None, bool]:
    """Resolve or mint a logical Source from a media-aware identity."""
    if (
        media_kind is None
        and resource_code is None
        and scope_kind is None
        and scope_value is None
    ):
        canonical = canonical_url(url)
        if not canonical:
            return None, False
        identity = {"canonical_url": canonical}
        encoded = json.dumps(
            identity, ensure_ascii=False, sort_keys=True, separators=(",", ":")
        )
        source_id = "src-" + hashlib.sha256(encoded.encode()).hexdigest()[:16]
        cursor = conn.execute(
            "INSERT INTO source (id, identity, title, media_type)"
            " VALUES (%s, %s, %s, %s) ON CONFLICT (id) DO NOTHING",
            (source_id, Jsonb(identity), title, media_type(url)),
        )
        return source_id, bool(cursor.rowcount)
    identity = source_identity(
        url,
        media_kind=media_kind,
        resource_code=resource_code,
        scope_kind=scope_kind,
        scope_value=scope_value,
    )
    if identity is None:
        return None, False
    encoded = json.dumps(identity, ensure_ascii=False, sort_keys=True, separators=(",", ":"))
    source_id = "src-" + hashlib.sha256(encoded.encode()).hexdigest()[:16]
    cursor = conn.execute(
        "INSERT INTO source (id, identity, title, media_type)"
        " VALUES (%s, %s, %s, %s) ON CONFLICT (id) DO NOTHING",
        (source_id, Jsonb(identity), title, identity["kind"]),
    )
    return source_id, bool(cursor.rowcount)


def import_workbook(
    conn: psycopg.Connection,
    path: str | Path,
    name: str,
    *,
    syllabus_id: str | None = None,
    institution_id: str | None = None,
    occupied_graph_ids: set[str] | tuple[str, ...] | list[str] = (),
    require_syllabus_metadata: bool = True,
    actor: str = "founder",
) -> dict:
    """Author one complete uploaded version under a manually named syllabus.

    New named Syllabi require a Companion Institution. Their graph id derives
    from that Institution and the Syllabus name.
    Passing ``require_syllabus_metadata=False`` is the explicit compatibility
    path for historical fixtures and migrations that predate that model.
    """
    name = (name or "").strip()
    if not name:
        raise ValueError("syllabus name is required")
    resolved_id = (syllabus_id or slugify(name)).strip()
    if not resolved_id:
        raise ValueError("syllabus id is empty after normalizing its name")
    stored = conn.execute(
        "SELECT title, institution_id, graph_id, group_id"
        " FROM syllabus WHERE id = %s",
        (resolved_id,),
    ).fetchone()
    syllabus_exists = stored is not None
    creation_requested = require_syllabus_metadata and syllabus_id is None
    exact_title = (
        conn.execute(
            "SELECT id, title, graph_id FROM syllabus WHERE title = %s"
            " ORDER BY created_at, id LIMIT 1",
            (name,),
        ).fetchone()
        if creation_requested
        else None
    )
    clean_institution_id = str(institution_id or "").strip()
    if syllabus_exists and not clean_institution_id:
        clean_institution_id = stored[1] or ""
    if require_syllabus_metadata and not clean_institution_id:
        raise ValueError("Selecione uma instituição do Companion.")
    if clean_institution_id:
        institution = conn.execute(
            "SELECT id, name FROM institution WHERE id = %s",
            (clean_institution_id,),
        ).fetchone()
        if institution is None:
            raise ValueError("Selecione uma instituição existente no Companion.")
    if clean_institution_id:
        resolved_id = validate_syllabus_id(resolved_id)
    if exact_title is not None:
        raise SyllabusAlreadyExists(exact_title[0], exact_title[1], exact_title[2])
    if creation_requested and syllabus_exists:
        raise GraphIdConflict(graph_id_for(clean_institution_id, name))
    resolved_graph_id = (stored[2] or "") if syllabus_exists else ""
    if not resolved_graph_id and clean_institution_id:
        resolved_graph_id = graph_id_for(clean_institution_id, name)
    if resolved_graph_id:
        resolved_graph_id = validate_graph_id(resolved_graph_id)
        local_owner = conn.execute(
            "SELECT id FROM syllabus WHERE graph_id = %s",
            (resolved_graph_id,),
        ).fetchone()
        if local_owner is not None and local_owner[0] != resolved_id:
            raise GraphIdConflict(resolved_graph_id)
        graph_id_is_new = not syllabus_exists or not stored[2]
        if graph_id_is_new and resolved_graph_id in set(occupied_graph_ids):
            raise GraphIdConflict(resolved_graph_id)
    elif require_syllabus_metadata:
        raise ValueError("Não foi possível gerar o graph ID do syllabus.")
    path = Path(path)
    file_body = path.read_bytes()
    file_sha = hashlib.sha256(file_body).hexdigest()
    parsed = parse_workbook(path)

    try:
        inserted = conn.execute(
            "INSERT INTO syllabus"
            " (id, title, institution_id, graph_id)"
            " VALUES (%s, %s, %s, %s) ON CONFLICT (id) DO NOTHING",
            (
                resolved_id,
                name,
                clean_institution_id or None,
                resolved_graph_id or None,
            ),
        ).rowcount
    except psycopg.errors.UniqueViolation as exc:
        if exc.diag.constraint_name == "syllabus_graph_id_key":
            raise GraphIdConflict(resolved_graph_id) from exc
        raise
    stored = conn.execute(
        "SELECT title, institution_id, graph_id, group_id"
        " FROM syllabus WHERE id = %s FOR UPDATE",
        (resolved_id,),
    ).fetchone()
    if creation_requested and not inserted:
        if stored[0] == name:
            raise SyllabusAlreadyExists(resolved_id, stored[0], stored[2])
        raise GraphIdConflict(resolved_graph_id)
    if stored[0] != name:
        raise ValueError(
            f"syllabus id {resolved_id!r} already belongs to {stored[0]!r}, not {name!r}"
        )
    if clean_institution_id and not inserted:
        if stored[1] not in {None, clean_institution_id}:
            raise ValueError("a instituição não corresponde ao syllabus existente")
        if stored[2] not in {None, resolved_graph_id}:
            raise ValueError("o graph ID não corresponde ao syllabus existente")
        if stored[3] is not None:
            group_institution = conn.execute(
                "SELECT institution_id FROM study_group WHERE id = %s",
                (stored[3],),
            ).fetchone()
            if group_institution and group_institution[0] != clean_institution_id:
                raise ValueError("a instituição não corresponde ao grupo do syllabus")
        conn.execute(
            "UPDATE syllabus SET institution_id = %s, graph_id = %s"
            " WHERE id = %s",
            (clean_institution_id, resolved_graph_id or None, resolved_id),
        )
    previous = _latest_version_row(conn, resolved_id)
    if previous and previous["file_sha"] == file_sha:
        lesson_count, reference_count = _version_counts(conn, previous["id"])
        conn.commit()
        return {
            "syllabus_id": resolved_id,
            "version_id": previous["id"],
            "seq": previous["seq"],
            "unchanged": True,
            "lesson_count": lesson_count,
            "reference_count": reference_count,
            "source_count": reference_count,
            "new_source_count": 0,
            "diff": {},
            "dropped": parsed["dropped"],
            "dropped_summary": parsed["dropped_summary"],
        }

    next_seq = (previous["seq"] if previous else 0) + 1
    version_id = f"{resolved_id}:v{next_seq:04d}"
    conn.execute(
        "INSERT INTO syllabus_version"
        " (id, syllabus_id, seq, origin, input_format, file_name, file_mime, file_sha, file_body)"
        " VALUES (%s, %s, %s, 'upload', %s, %s, %s, %s, %s)",
        (
            version_id,
            resolved_id,
            next_seq,
            parsed["format"],
            path.name,
            XLSX_MIME,
            file_sha,
            file_body,
        ),
    )

    new_sources = 0
    reference_count = 0
    for lesson_number, lesson in enumerate(parsed["lessons"], start=1):
        lesson_id = f"{version_id}:lesson:{lesson_number:04d}"
        conn.execute(
            "INSERT INTO syllabus_lesson"
            " (id, version_id, week, seq, kind, title, subject, subjects, lesson_date,"
            "  description, is_hidden, fields, activity_uuid, folder_uuid, week_order,"
            "  activity_order)"
            " VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)",
            (
                lesson_id,
                version_id,
                lesson["week"],
                lesson["seq"],
                lesson["kind"],
                lesson["title"],
                lesson.get("subject"),
                lesson.get("subjects") or [],
                lesson.get("lesson_date"),
                lesson.get("description"),
                bool(lesson.get("is_hidden")),
                Jsonb(lesson["fields"]),
                lesson.get("activity_uuid"),
                lesson.get("folder_uuid"),
                lesson.get("week_order"),
                lesson.get("activity_order"),
            ),
        )
        for reference in lesson["source_references"]:
            reference_count += 1
            reference_id = f"{version_id}:source:{reference_count:04d}"
            source_id, created = resolve_source(
                conn,
                reference.get("url") or "",
                reference["title"],
                media_kind=reference["media_type"],
                resource_code=reference.get("resource_code"),
                scope_kind=reference.get("scope_kind"),
                scope_value=reference.get("scope_value"),
            )
            new_sources += int(created)
            conn.execute(
                "INSERT INTO syllabus_source_reference"
                " (id, version_id, lesson_id, seq, title, description, url, media_type,"
                "  resource_code, scope_kind, scope_value, source_id, is_hidden, fields,"
                "  activity_uuid, folder_uuid, week_order, activity_order,"
                "  parent_activity_uuid, parent_inference)"
                " VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s,"
                " %s, %s, %s, %s, %s, %s)",
                (
                    reference_id,
                    version_id,
                    lesson_id,
                    reference["seq"],
                    reference["title"],
                    reference.get("description"),
                    reference.get("url"),
                    reference["media_type"],
                    reference.get("resource_code"),
                    reference.get("scope_kind"),
                    reference.get("scope_value"),
                    source_id,
                    bool(reference.get("is_hidden")),
                    Jsonb(reference["fields"]),
                    reference.get("activity_uuid"),
                    reference.get("folder_uuid"),
                    reference.get("week_order"),
                    reference.get("activity_order"),
                    reference.get("parent_activity_uuid"),
                    reference.get("parent_inference"),
                ),
            )

    event_id = next_curation_event_id(conn)
    conn.execute(
        "INSERT INTO curation_event (id, actor, action, subject)"
        " VALUES (%s, %s, 'syllabus_upload', %s)",
        (
            event_id,
            actor,
            Jsonb(
                {
                    "syllabus_id": resolved_id,
                    "version_id": version_id,
                    "file_sha": file_sha,
                    "input_format": parsed["format"],
                }
            ),
        ),
    )
    diff = diff_versions(conn, previous["id"], version_id) if previous else {}
    conn.commit()
    return {
        "syllabus_id": resolved_id,
        "version_id": version_id,
        "seq": next_seq,
        "unchanged": False,
        "lesson_count": parsed["lesson_count"],
        "reference_count": reference_count,
        "source_count": reference_count,
        "new_source_count": new_sources,
        "diff": diff,
        "dropped": parsed["dropped"],
        "dropped_summary": parsed["dropped_summary"],
    }


def _latest_version_row(conn: psycopg.Connection, syllabus_id: str) -> dict | None:
    row = conn.execute(
        "SELECT id, seq, origin, input_format, file_name, file_sha, note, created_at"
        " FROM syllabus_version WHERE syllabus_id = %s ORDER BY seq DESC LIMIT 1",
        (syllabus_id,),
    ).fetchone()
    if row is None:
        return None
    return dict(
        zip(
            ("id", "seq", "origin", "input_format", "file_name", "file_sha", "note", "created_at"),
            row,
        )
    )


def _lesson_subjects_by_syllabus(
    conn: psycopg.Connection,
    syllabus_ids: list[str] | tuple[str, ...] | None = None,
) -> dict[str, list[dict]]:
    query = (
        "WITH latest AS ("
        " SELECT syllabus_id, max(seq) AS seq FROM syllabus_version"
        " GROUP BY syllabus_id"
        ")"
        " SELECT version.syllabus_id, lesson.subject, lesson.fields"
        " FROM latest"
        " JOIN syllabus_version version"
        " ON version.syllabus_id = latest.syllabus_id AND version.seq = latest.seq"
        " JOIN syllabus_lesson lesson ON lesson.version_id = version.id"
        " WHERE lesson.subject IS NOT NULL"
    )
    params: tuple[object, ...] = ()
    if syllabus_ids is not None:
        if not syllabus_ids:
            return {}
        query += " AND version.syllabus_id = ANY(%s)"
        params = (list(syllabus_ids),)
    query += " ORDER BY version.syllabus_id, lesson.subject, lesson.id"
    by_code: dict[str, dict[str, dict]] = {}
    for syllabus_id, code, fields in conn.execute(query, params).fetchall():
        fields = fields if isinstance(fields, dict) else {}
        activity = fields.get("adalove_activity") or {}
        workbook_name = str(activity.get("Lesson Subject code") or "").strip()
        workbook_code = LESSON_SUBJECT_CODES.get(_ascii(workbook_name).casefold())
        display_name = (
            workbook_name
            if workbook_name and workbook_code is None
            else LESSON_SUBJECT_NAMES.get(code, code)
        )
        by_code.setdefault(syllabus_id, {}).setdefault(
            code,
            {"code": code, "display_name": display_name},
        )
    result = {
        syllabus_id: [subjects[code] for code in sorted(subjects)]
        for syllabus_id, subjects in by_code.items()
    }
    return result


def _syllabus_payload(row: tuple, lesson_subjects: list[dict]) -> dict:
    syllabus_id, title, graph_id, institution_id, institution_name, created_at = row
    institution = (
        {"id": institution_id, "name": institution_name}
        if institution_id is not None
        else None
    )
    metadata_complete = bool(
        institution_id
        and graph_id
        and SYLLABUS_ID.fullmatch(syllabus_id)
    )
    export_identity = (
        {
            "graph_id": graph_id,
            "display_name": title,
            "institution_slug": institution_id,
        }
        if metadata_complete
        else None
    )
    return {
        "id": syllabus_id,
        "title": title,
        "display_name": title,
        "institution_id": institution_id,
        "institution": institution,
        "lesson_subjects": lesson_subjects,
        "metadata_complete": metadata_complete,
        "graph_id": graph_id,
        "institution_slug": institution_id,
        "export_identity": export_identity,
        "created_at": created_at,
    }


def list_syllabi(conn: psycopg.Connection) -> list[dict]:
    """List syllabi with a compact summary of only their latest version."""
    rows = conn.execute(
        "SELECT syllabus.id, syllabus.title, syllabus.graph_id,"
        " institution.id, institution.name, syllabus.created_at"
        " FROM syllabus LEFT JOIN institution"
        " ON institution.id = syllabus.institution_id"
        " ORDER BY syllabus.created_at DESC, syllabus.id"
    ).fetchall()
    subjects = _lesson_subjects_by_syllabus(
        conn, [row[0] for row in rows]
    )
    result = []
    for row in rows:
        syllabus_id = row[0]
        latest = _latest_version_row(conn, syllabus_id)
        if latest:
            lesson_count, source_count = _version_counts(conn, latest["id"])
            latest = {**latest, "lesson_count": lesson_count, "source_count": source_count}
        result.append({**_syllabus_payload(row, subjects.get(syllabus_id, [])), "latest": latest})
    return result


def get_syllabus_history(conn: psycopg.Connection, syllabus_id: str) -> dict:
    """Return immutable versions newest first, each with full-state counts."""
    syllabus = conn.execute(
        "SELECT syllabus.id, syllabus.title, syllabus.graph_id,"
        " institution.id, institution.name, syllabus.created_at"
        " FROM syllabus LEFT JOIN institution"
        " ON institution.id = syllabus.institution_id"
        " WHERE syllabus.id = %s",
        (syllabus_id,),
    ).fetchone()
    if syllabus is None:
        raise LookupError(f"unknown syllabus {syllabus_id!r}")
    rows = conn.execute(
        "SELECT sv.id, sv.seq, sv.origin, sv.input_format, sv.file_name, sv.file_sha,"
        " sv.note, sv.created_at,"
        " (SELECT count(*) FROM syllabus_lesson sl WHERE sl.version_id = sv.id),"
        " (SELECT count(*) FROM syllabus_source_reference sr WHERE sr.version_id = sv.id)"
        " FROM syllabus_version sv WHERE sv.syllabus_id = %s ORDER BY sv.seq DESC",
        (syllabus_id,),
    ).fetchall()
    versions = [
        dict(
            zip(
                (
                    "id", "seq", "origin", "input_format", "file_name", "file_sha",
                    "note", "created_at", "lesson_count", "source_count",
                ),
                row,
            )
        )
        for row in rows
    ]
    subjects = _lesson_subjects_by_syllabus(conn, [syllabus_id])
    return {
        **_syllabus_payload(syllabus, subjects.get(syllabus_id, [])),
        "versions": versions,
    }


def get_syllabus_workbook(conn: psycopg.Connection, version_id: str) -> dict:
    """Return the exact uploaded XLSX for download or audit."""
    row = conn.execute(
        "SELECT file_name, file_mime, file_sha, file_body"
        " FROM syllabus_version WHERE id = %s",
        (version_id,),
    ).fetchone()
    if row is None:
        raise LookupError(f"unknown syllabus version {version_id!r}")
    if row[3] is None:
        raise LookupError(f"syllabus version {version_id!r} has no uploaded workbook")
    return {"file_name": row[0], "mime_type": row[1], "sha256": row[2], "body": bytes(row[3])}


def get_syllabus_version(
    conn: psycopg.Connection, syllabus_id: str, version_id: str | None = None
) -> dict:
    """Return one full version; omitting ``version_id`` means latest."""
    syllabus = conn.execute(
        "SELECT syllabus.id, syllabus.title, syllabus.graph_id,"
        " institution.id, institution.name, syllabus.created_at"
        " FROM syllabus LEFT JOIN institution"
        " ON institution.id = syllabus.institution_id"
        " WHERE syllabus.id = %s",
        (syllabus_id,),
    ).fetchone()
    if syllabus is None:
        raise LookupError(f"unknown syllabus {syllabus_id!r}")
    if version_id is None:
        version = _latest_version_row(conn, syllabus_id)
    else:
        row = conn.execute(
            "SELECT id, seq, origin, input_format, file_name, file_sha, note, created_at"
            " FROM syllabus_version WHERE syllabus_id = %s AND id = %s",
            (syllabus_id, version_id),
        ).fetchone()
        version = (
            dict(zip(("id", "seq", "origin", "input_format", "file_name", "file_sha", "note", "created_at"), row))
            if row
            else None
        )
    if version is None:
        raise LookupError(f"unknown version {version_id!r} for syllabus {syllabus_id!r}")

    lesson_rows = conn.execute(
        "SELECT sl.id, sl.week, sl.seq, sl.kind, sl.title, sl.subject, sl.subjects,"
        " sl.lesson_date, sl.description, sl.fields, sl.created_at, sl.is_hidden,"
        " sl.activity_uuid, sl.folder_uuid, sl.week_order, sl.activity_order"
        " FROM syllabus_lesson sl"
        " WHERE sl.version_id = %s ORDER BY sl.week NULLS LAST, sl.seq, sl.id",
        (version["id"],),
    ).fetchall()
    lessons = []
    for row in lesson_rows:
        lesson = dict(
            zip(
                (
                    "id", "week", "seq", "kind", "title", "subject", "subjects",
                    "date", "description", "fields", "created_at", "hidden",
                    "activity_uuid", "folder_uuid", "week_order", "activity_order",
                ),
                row,
            )
        )
        reference_rows = conn.execute(
            "SELECT sr.id, sr.source_id, sr.seq, sr.title, sr.description, sr.url,"
            " sr.media_type, sr.resource_code, sr.scope_kind, sr.scope_value, sr.is_hidden,"
            " sr.fields, sr.created_at, s.identity,"
            " coalesce(rr.is_validated, false), rr.complexity,"
            " sr.activity_uuid, sr.folder_uuid, sr.week_order, sr.activity_order,"
            " sr.parent_activity_uuid, sr.parent_inference"
            " FROM syllabus_source_reference sr"
            " LEFT JOIN source s ON s.id = sr.source_id"
            " LEFT JOIN syllabus_source_review rr ON rr.reference_id = sr.id"
            " WHERE sr.version_id = %s AND sr.lesson_id = %s"
            " ORDER BY sr.seq, sr.id",
            (version["id"], lesson["id"]),
        ).fetchall()
        sources = []
        for reference_row in reference_rows:
            source = dict(
                zip(
                    (
                        "reference_id", "source_id", "seq", "title", "description", "url",
                        "media_type", "resource_code", "scope_kind", "scope_value", "hidden",
                        "fields", "created_at", "identity",
                        "validated", "complexity", "activity_uuid", "folder_uuid",
                        "week_order", "activity_order", "parent_activity_uuid",
                        "parent_inference",
                    ),
                    reference_row,
                )
            )
            source["review"] = {
                "validated": source.pop("validated"),
                "complexity": source.pop("complexity"),
            }
            source["scope"] = (
                {"kind": source["scope_kind"], "value": source["scope_value"]}
                if source["scope_kind"]
                else None
            )
            sources.append(source)
        lesson["sources"] = sources
        lessons.append(lesson)
    subjects = _lesson_subjects_by_syllabus(conn, [syllabus_id]).get(
        syllabus_id, []
    )
    by_code = {subject["code"]: subject for subject in subjects}
    for lesson in lessons:
        lesson["lesson_subject"] = by_code.get(lesson["subject"])
    return {
        **_syllabus_payload(syllabus, subjects),
        "version": version,
        "lessons": lessons,
    }


def update_source_review(
    conn: psycopg.Connection,
    syllabus_id: str,
    reference_id: str,
    changes: object,
) -> dict:
    """Update the small operational review state for one source reference."""
    if not isinstance(changes, dict) or not changes:
        raise ValueError("Informe ao menos um marcador do autoestudo.")
    unknown = set(changes) - {"validated", "complexity"}
    if unknown:
        raise ValueError("A revisão contém campos desconhecidos.")
    reference = conn.execute(
        "SELECT 1 FROM syllabus_source_reference sr"
        " JOIN syllabus_version sv ON sv.id = sr.version_id"
        " WHERE sr.id = %s AND sv.syllabus_id = %s",
        (reference_id, syllabus_id),
    ).fetchone()
    if reference is None:
        raise LookupError(
            f"unknown source reference {reference_id!r} for syllabus {syllabus_id!r}"
        )

    current = conn.execute(
        "SELECT is_validated, complexity FROM syllabus_source_review WHERE reference_id = %s",
        (reference_id,),
    ).fetchone() or (False, None)
    validated, complexity = current
    if "validated" in changes:
        if not isinstance(changes["validated"], bool):
            raise ValueError("O marcador de validação deve ser verdadeiro ou falso.")
        validated = changes["validated"]
    if "complexity" in changes:
        complexity = changes["complexity"]
        if complexity not in {None, "simple", "complex"}:
            raise ValueError("A complexidade deve ser simples, complexa ou vazia.")

    conn.execute(
        "INSERT INTO syllabus_source_review (reference_id, is_validated, complexity)"
        " VALUES (%s, %s, %s)"
        " ON CONFLICT (reference_id) DO UPDATE SET"
        " is_validated = excluded.is_validated, complexity = excluded.complexity,"
        " updated_at = now()",
        (reference_id, validated, complexity),
    )
    conn.commit()
    return {"validated": validated, "complexity": complexity}


class SyllabusVersionConflict(ValueError):
    """Raised when a curator tries to save over a newer immutable version."""


def _clean_edit_text(value, *, field: str, required: bool = False, limit: int = 20_000) -> str | None:
    text = _text(value)
    if required and not text:
        raise ValueError(f"{field} is required")
    if text and len(text) > limit:
        raise ValueError(f"{field} exceeds {limit} characters")
    return text


def _clean_subjects(value: object, *, lesson_index: int) -> list[str]:
    if value in (None, ""):
        return []
    if not isinstance(value, (str, list, tuple)):
        raise ValueError(f"lesson {lesson_index}: subjects must be a list")
    subjects = parse_subjects(value)
    if len(subjects) > 200:
        raise ValueError(f"lesson {lesson_index}: subjects exceed the limit of 200")
    for subject in subjects:
        _clean_edit_text(
            subject,
            field=f"lesson {lesson_index} subject",
            required=True,
            limit=500,
        )
    return subjects


def _base_fields(item: dict | None) -> dict:
    fields = (item or {}).get("fields") or {}
    return dict(fields) if isinstance(fields, dict) else {}


def _normalize_curation_projection(
    base: dict,
    submitted_lessons: object,
    *,
    trust_workbook_metadata: bool = False,
) -> list[dict]:
    """Validate an editor payload and inherit non-editable workbook metadata.

    The browser submits the complete desired projection, but arbitrary client
    ``fields`` are never trusted. Existing rows inherit their original field
    bag by id; newly authored rows start empty and receive the canonical typed
    values below. This keeps professor/assessment metadata lossless without
    letting stale form state overwrite unrelated columns.
    """
    if not isinstance(submitted_lessons, list) or not submitted_lessons:
        raise ValueError("the syllabus must contain at least one lesson")
    if len(submitted_lessons) > 500:
        raise ValueError("the syllabus exceeds the limit of 500 lessons")

    base_lessons = {lesson["id"]: lesson for lesson in base.get("lessons", [])}
    base_references = {
        source["reference_id"]: source
        for lesson in base.get("lessons", [])
        for source in lesson.get("sources", [])
    }
    normalized: list[dict] = []
    source_total = 0
    curated_activity_order_by_week: dict[int | None, int] = {}
    for lesson_index, raw_lesson in enumerate(submitted_lessons, 1):
        if not isinstance(raw_lesson, dict):
            raise ValueError(f"lesson {lesson_index} is invalid")
        base_lesson = base_lessons.get(str(raw_lesson.get("id") or ""))
        try:
            week = int(raw_lesson.get("week")) if raw_lesson.get("week") not in {None, ""} else None
        except (TypeError, ValueError) as exc:
            raise ValueError(f"lesson {lesson_index}: week must be a whole number") from exc
        if week is not None and not 0 < week <= 1000:
            raise ValueError(f"lesson {lesson_index}: week must be between 1 and 1000")
        lesson_date = _as_date(raw_lesson.get("date"))
        if raw_lesson.get("date") and lesson_date is None:
            raise ValueError(f"lesson {lesson_index}: date must use YYYY-MM-DD")
        raw_sources = raw_lesson.get("sources", [])
        if not isinstance(raw_sources, list) or len(raw_sources) > 500:
            raise ValueError(f"lesson {lesson_index}: invalid source list")
        source_total += len(raw_sources)
        if source_total > 5000:
            raise ValueError("the syllabus exceeds the limit of 5000 sources")

        lesson_metadata = raw_lesson if trust_workbook_metadata else (base_lesson or {})
        curated_activity_order_by_week[week] = (
            curated_activity_order_by_week.get(week, 0) + 1
        )
        lesson_fields = (
            dict(raw_lesson.get("fields") or {})
            if trust_workbook_metadata
            else _base_fields(base_lesson)
        )
        lesson = {
            "id": (base_lesson or {}).get("id"),
            "week": week,
            "seq": lesson_index,
            "activity_uuid": lesson_metadata.get("activity_uuid"),
            "folder_uuid": lesson_metadata.get("folder_uuid"),
            "week_order": (
                lesson_metadata.get("week_order", week)
                if trust_workbook_metadata
                else week
            ),
            "activity_order": (
                lesson_metadata.get("activity_order", lesson_index)
                if trust_workbook_metadata
                else curated_activity_order_by_week[week]
            ),
            "kind": _clean_edit_text(
                raw_lesson.get("kind") or (base_lesson or {}).get("kind") or "Class",
                field=f"lesson {lesson_index} kind",
                required=True,
                limit=200,
            ),
            "title": _clean_edit_text(
                raw_lesson.get("title"),
                field=f"lesson {lesson_index} title",
                required=True,
                limit=1000,
            ),
            "subject": _clean_edit_text(
                raw_lesson.get("subject"), field=f"lesson {lesson_index} subject", limit=500
            ),
            "subjects": _clean_subjects(
                raw_lesson.get(
                    "subjects", (base_lesson or {}).get("subjects") or []
                ),
                lesson_index=lesson_index,
            ),
            "lesson_date": lesson_date,
            "description": _clean_edit_text(
                raw_lesson.get("description"),
                field=f"lesson {lesson_index} description",
                limit=20_000,
            ),
            "is_hidden": bool(raw_lesson.get("hidden")),
            "fields": lesson_fields,
            "source_references": [],
        }
        curated_source_orders: dict[str, int] = {}
        for source_index, raw_source in enumerate(raw_sources, 1):
            if not isinstance(raw_source, dict):
                raise ValueError(f"lesson {lesson_index}, source {source_index} is invalid")
            base_source = base_references.get(str(raw_source.get("reference_id") or ""))
            kind = str(raw_source.get("media_type") or "article").strip().lower()
            if kind not in {"article", "video", "book"}:
                raise ValueError(
                    f"lesson {lesson_index}, source {source_index}: invalid media type"
                )
            code = _clean_code(raw_source.get("resource_code"))
            scope_kind = _clean_edit_text(
                raw_source.get("scope_kind"),
                field=f"lesson {lesson_index}, source {source_index} scope kind",
                limit=100,
            )
            scope_value = _clean_edit_text(
                raw_source.get("scope_value"),
                field=f"lesson {lesson_index}, source {source_index} scope value",
                limit=500,
            )
            if bool(scope_kind) != bool(scope_value):
                raise ValueError(
                    f"lesson {lesson_index}, source {source_index}: scope kind and value must be provided together"
                )
            if scope_value:
                scope_value = _normalize_scope_value(scope_value)
            source_metadata = raw_source if trust_workbook_metadata else (base_source or {})
            source_fields = (
                dict(raw_source.get("fields") or {})
                if trust_workbook_metadata
                else _base_fields(base_source)
            )
            source_group_identity = source_metadata.get("activity_uuid") or (
                f"new:{lesson_index}:{source_index}"
            )
            if source_group_identity not in curated_source_orders:
                curated_activity_order_by_week[week] += 1
                curated_source_orders[source_group_identity] = (
                    curated_activity_order_by_week[week]
                )
            reference = {
                "seq": source_index,
                "activity_uuid": source_metadata.get("activity_uuid"),
                "folder_uuid": source_metadata.get("folder_uuid"),
                "week_order": (
                    source_metadata.get("week_order", week)
                    if trust_workbook_metadata
                    else week
                ),
                "activity_order": (
                    source_metadata.get("activity_order", source_index)
                    if trust_workbook_metadata
                    else curated_source_orders[source_group_identity]
                ),
                "parent_activity_uuid": (
                    source_metadata.get("parent_activity_uuid")
                    or lesson_metadata.get("activity_uuid")
                ),
                "parent_inference": source_metadata.get("parent_inference")
                or "curated_explicit_parent",
                "title": _clean_edit_text(
                    raw_source.get("title"),
                    field=f"lesson {lesson_index}, source {source_index} title",
                    required=True,
                    limit=1000,
                ),
                "description": _clean_edit_text(
                    raw_source.get("description"),
                    field=f"lesson {lesson_index}, source {source_index} description",
                    limit=4000,
                ),
                "url": _clean_edit_text(
                    raw_source.get("url"),
                    field=f"lesson {lesson_index}, source {source_index} URL",
                    limit=8000,
                ),
                "media_type": kind,
                "resource_code": code,
                "scope_kind": scope_kind,
                "scope_value": scope_value,
                "is_hidden": bool(raw_source.get("hidden")),
                "fields": source_fields,
                "_base_review": dict((base_source or {}).get("review") or {}),
            }
            reference["_content_unchanged"] = bool(
                base_source
                and _projection_signature(
                    [{"source_references": [reference]}]
                ) == _projection_signature(
                    [{"sources": [base_source]}]
                )
            )
            lesson["source_references"].append(reference)
        normalized.append(lesson)
    return normalized


def _projection_signature(lessons: list[dict]) -> list[dict]:
    """Return the authored syllabus meaning and stable Lesson identities."""
    return [
        {
            "id": lesson.get("id"),
            "week": lesson.get("week"),
            "kind": lesson.get("kind"),
            "title": lesson.get("title"),
            "subject": lesson.get("subject"),
            "subjects": list(lesson.get("subjects") or []),
            "date": str(lesson.get("lesson_date") or lesson.get("date") or ""),
            "description": lesson.get("description"),
            "hidden": bool(lesson.get("is_hidden", lesson.get("hidden", False))),
            "activity_uuid": lesson.get("activity_uuid"),
            "folder_uuid": lesson.get("folder_uuid"),
            "week_order": lesson.get("week_order"),
            "activity_order": lesson.get("activity_order"),
            "sources": [
                {
                    "title": source.get("title"),
                    "description": source.get("description"),
                    "url": source.get("url"),
                    "media_type": source.get("media_type"),
                    "resource_code": source.get("resource_code"),
                    "scope_kind": source.get("scope_kind"),
                    "scope_value": source.get("scope_value"),
                    "hidden": bool(source.get("is_hidden", source.get("hidden", False))),
                    "activity_uuid": source.get("activity_uuid"),
                    "folder_uuid": source.get("folder_uuid"),
                    "week_order": source.get("week_order"),
                    "activity_order": source.get("activity_order"),
                    "parent_activity_uuid": source.get("parent_activity_uuid"),
                    "parent_inference": source.get("parent_inference"),
                }
                for source in lesson.get("source_references", lesson.get("sources", []))
            ],
        }
        for lesson in lessons
    ]


def _curated_uuid(kind: str, *parts: object) -> str:
    encoded = "\0".join(str(part or "") for part in parts).encode()
    return f"curated-{kind}-" + hashlib.sha256(encoded).hexdigest()[:24]


def _formatted_date(value: object) -> str | None:
    parsed = _as_date(value)
    return parsed.strftime("%d/%m/%Y") if parsed else None


def _append_sheet(workbook: Workbook, name: str, columns: tuple[str, ...], rows: list[dict]) -> None:
    sheet = workbook.active if name == "Activities" else workbook.create_sheet()
    sheet.title = name
    sheet.append(columns)
    for row in rows:
        sheet.append([row.get(column) for column in columns])
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions


def compile_syllabus_workbook(syllabus_title: str, lessons: list[dict]) -> bytes:
    """Compile one curated projection using the Adalove observer workbook shape."""
    activities: list[dict] = []
    subjects: list[dict] = []
    materials: list[dict] = []
    audits: list[dict] = []

    def add_activity(
        item: dict,
        *,
        kind: str,
        activity_uuid: str,
        folder_uuid: str,
        week_order: int,
        activity_order: int,
        parent: dict | None = None,
    ) -> dict:
        stored = dict((item.get("fields") or {}).get("adalove_activity") or {})
        stored.update(
            {
                "Activity order": activity_order,
                "Week": week_order,
                "Type": kind,
                "Original label": stored.get("Original label")
                or stored.get("Type caption")
                or kind,
                "Title": item.get("title"),
                "Description": item.get("description"),
                "Date": _formatted_date(item.get("lesson_date", item.get("date"))),
                "Lesson Subject code": item.get("subject"),
                "Related subjects": "; ".join(item.get("subjects") or []),
                "Primary URL": item.get("url"),
                "Resource code": item.get("resource_code"),
                "Parent activity UUID": (parent or {}).get("activity_uuid"),
                "Parent title": (parent or {}).get("title"),
                "Parent inference": item.get("parent_inference")
                if parent
                else None,
                "Activity UUID": activity_uuid,
                "Folder UUID": folder_uuid,
                HIDDEN_COLUMN: "yes"
                if item.get("is_hidden", item.get("hidden", False))
                else "no",
                "Detail error": None,
            }
        )
        activities.append(stored)
        audit = dict((item.get("fields") or {}).get("adalove_order_audit") or {})
        audit.update(
            {
                "Activity order": activity_order,
                "Week": week_order,
                "Order key": f"{week_order}:{activity_order}",
                "Activity UUID": activity_uuid,
                "Folder UUID": folder_uuid,
                "Type": kind,
                "Title": item.get("title"),
                "Parent inference": item.get("parent_inference") if parent else None,
                "Detail error": None,
            }
        )
        audits.append(audit)
        return stored

    for lesson_index, lesson in enumerate(lessons, 1):
        week_order = lesson.get("week_order") or lesson.get("week") or lesson_index
        activity_order = lesson.get("activity_order") or lesson.get("seq") or lesson_index
        activity_uuid = lesson.get("activity_uuid") or _curated_uuid(
            "lesson", week_order, activity_order, lesson.get("title")
        )
        folder_uuid = lesson.get("folder_uuid") or _curated_uuid("week", week_order)
        lesson_identity = {
            "activity_uuid": activity_uuid,
            "folder_uuid": folder_uuid,
            "title": lesson.get("title"),
        }
        add_activity(
            lesson,
            kind=lesson.get("kind") or "Class",
            activity_uuid=activity_uuid,
            folder_uuid=folder_uuid,
            week_order=week_order,
            activity_order=activity_order,
        )
        stored_subjects = (lesson.get("fields") or {}).get("adalove_subjects") or []
        subject_uuid_by_name = {
            row.get("Related subject"): row.get("Subject UUID") for row in stored_subjects
        }
        for subject in lesson.get("subjects") or []:
            subjects.append(
                {
                    "Activity order": activity_order,
                    "Week": week_order,
                    "Activity UUID": activity_uuid,
                    "Activity title": lesson.get("title"),
                    "Lesson Subject code": lesson.get("subject"),
                    "Subject UUID": subject_uuid_by_name.get(subject),
                    "Related subject": subject,
                }
            )

        source_references = lesson.get("source_references") or []
        if lesson.get("kind") == "Deliverable":
            for source in source_references:
                material = dict((source.get("fields") or {}).get("adalove_material") or {})
                material.update(
                    {
                        "Activity order": activity_order,
                        "Week": week_order,
                        "Activity UUID": activity_uuid,
                        "Activity title": lesson.get("title"),
                        "Label": source.get("title"),
                        "URL": source.get("url"),
                        "Resource code": source.get("resource_code"),
                        "Video": "Sim" if source.get("media_type") == "video" else "Não",
                    }
                )
                materials.append(material)
            continue

        source_groups: dict[str, list[dict]] = {}
        for source_index, source in enumerate(source_references, 1):
            source_uuid = source.get("activity_uuid") or _curated_uuid(
                "source", activity_uuid, source_index, source.get("title"), source.get("url")
            )
            source_groups.setdefault(source_uuid, []).append(source)
        for group_index, (source_uuid, group) in enumerate(source_groups.items(), 1):
            first = group[0]
            source_order = first.get("activity_order") or activity_order + group_index
            source_week = first.get("week_order") or week_order
            source_folder = first.get("folder_uuid") or folder_uuid
            source_item = {
                **first,
                "title": (
                    (first.get("fields") or {}).get("adalove_activity") or {}
                ).get("Title")
                or first.get("title"),
                "parent_inference": first.get("parent_inference")
                or "curated_explicit_parent",
            }
            add_activity(
                source_item,
                kind="Self-study",
                activity_uuid=source_uuid,
                folder_uuid=source_folder,
                week_order=source_week,
                activity_order=source_order,
                parent=lesson_identity,
            )
            for subject_row in (first.get("fields") or {}).get("adalove_subjects") or []:
                row = dict(subject_row)
                row.update(
                    {
                        "Activity order": source_order,
                        "Week": source_week,
                        "Activity UUID": source_uuid,
                        "Activity title": source_item["title"],
                    }
                )
                subjects.append(row)
            for source in group:
                material = dict((source.get("fields") or {}).get("adalove_material") or {})
                material.update(
                    {
                        "Activity order": source_order,
                        "Week": source_week,
                        "Activity UUID": source_uuid,
                        "Activity title": source_item["title"],
                        "Label": source.get("title"),
                        "URL": source.get("url"),
                        "Resource code": source.get("resource_code"),
                        "Video": "Sim" if source.get("media_type") == "video" else "Não",
                    }
                )
                materials.append(material)

    workbook = Workbook()
    _append_sheet(workbook, "Activities", (*ADALOVE_ACTIVITY_COLUMNS, HIDDEN_COLUMN), activities)
    _append_sheet(workbook, "Subjects", ADALOVE_SUBJECT_COLUMNS, subjects)
    _append_sheet(workbook, "Materials", ADALOVE_MATERIAL_COLUMNS, materials)
    _append_sheet(workbook, "Order audit", ADALOVE_ORDER_AUDIT_COLUMNS, audits)
    read_me = workbook.create_sheet("Read me")
    read_me.append(("Field", "Value / note"))
    read_me.append(("Project", syllabus_title))
    read_me.append(("Format", "Adalove observer export with curated values"))
    stream = BytesIO()
    workbook.save(stream)
    workbook.close()
    return stream.getvalue()


def curate_syllabus(
    conn: psycopg.Connection,
    syllabus_id: str,
    base_version_id: str,
    lessons: object,
    *,
    actor: str = "founder",
    note: str | None = None,
    trust_workbook_metadata: bool = False,
) -> dict:
    """Author a full immutable curation version from the latest projection."""
    syllabus = conn.execute(
        "SELECT id, title FROM syllabus WHERE id = %s FOR UPDATE", (syllabus_id,)
    ).fetchone()
    if syllabus is None:
        raise LookupError(f"unknown syllabus {syllabus_id!r}")
    latest = _latest_version_row(conn, syllabus_id)
    if latest is None:
        raise LookupError(f"syllabus {syllabus_id!r} has no version")
    if latest["id"] != base_version_id:
        raise SyllabusVersionConflict(
            "Este syllabus recebeu uma versão mais nova. Recarregue antes de salvar suas mudanças."
        )
    base = get_syllabus_version(conn, syllabus_id, base_version_id)
    normalized = _normalize_curation_projection(
        base,
        lessons,
        trust_workbook_metadata=trust_workbook_metadata,
    )
    if _projection_signature(normalized) == _projection_signature(base["lessons"]):
        conn.commit()
        lesson_count, source_count = _version_counts(conn, base_version_id)
        return {
            "syllabus_id": syllabus_id,
            "version_id": base_version_id,
            "seq": latest["seq"],
            "unchanged": True,
            "lesson_count": lesson_count,
            "reference_count": source_count,
            "source_count": source_count,
            "new_source_count": 0,
            "diff": {"added": [], "removed": [], "changed": []},
        }

    raw_note = str(note or "").strip()
    if not raw_note:
        raise ValueError("A razão da nova versão é obrigatória.")
    if len(raw_note) > 500:
        raise ValueError(
            "A razão da nova versão deve ter no máximo 500 caracteres."
        )

    next_seq = latest["seq"] + 1
    version_id = f"{syllabus_id}:v{next_seq:04d}"
    body = compile_syllabus_workbook(syllabus[1], normalized)
    file_name = f"{syllabus_id}-v{next_seq:04d}.xlsx"
    file_sha = hashlib.sha256(body).hexdigest()
    clean_note = _clean_edit_text(
        raw_note, field="curation note", required=True, limit=500
    )
    conn.execute(
        "INSERT INTO syllabus_version"
        " (id, syllabus_id, seq, origin, input_format, file_name, file_mime, file_sha,"
        "  file_body, note)"
        " VALUES (%s, %s, %s, 'curation', %s, %s, %s, %s, %s, %s)",
        (
            version_id, syllabus_id, next_seq, latest.get("input_format") or "adalove-observer",
            file_name, XLSX_MIME, file_sha, body, clean_note,
        ),
    )

    reference_count = 0
    new_sources = 0
    for lesson_number, lesson in enumerate(normalized, 1):
        lesson_id = lesson.get("id") or f"{version_id}:lesson:{lesson_number:04d}"
        conn.execute(
            "INSERT INTO syllabus_lesson"
            " (id, version_id, week, seq, kind, title, subject, subjects, lesson_date,"
            "  description, is_hidden, fields, activity_uuid, folder_uuid, week_order,"
            "  activity_order)"
            " VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)",
            (
                lesson_id, version_id, lesson["week"], lesson_number, lesson["kind"],
                lesson["title"], lesson.get("subject"), lesson.get("subjects") or [],
                lesson.get("lesson_date"), lesson.get("description"),
                lesson["is_hidden"], Jsonb(lesson["fields"]),
                lesson.get("activity_uuid"), lesson.get("folder_uuid"),
                lesson.get("week_order"), lesson.get("activity_order"),
            ),
        )
        for source_number, reference in enumerate(lesson["source_references"], 1):
            reference_count += 1
            reference_id = f"{version_id}:source:{reference_count:04d}"
            source_id, created = resolve_source(
                conn,
                reference.get("url") or "",
                reference["title"],
                media_kind=reference["media_type"],
                resource_code=reference.get("resource_code"),
                scope_kind=reference.get("scope_kind"),
                scope_value=reference.get("scope_value"),
            )
            new_sources += int(created)
            conn.execute(
                "INSERT INTO syllabus_source_reference"
                " (id, version_id, lesson_id, seq, title, description, url, media_type,"
                "  resource_code, scope_kind, scope_value, source_id, is_hidden, fields,"
                "  activity_uuid, folder_uuid, week_order, activity_order,"
                "  parent_activity_uuid, parent_inference)"
                " VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s,"
                " %s, %s, %s, %s, %s, %s)",
                (
                    reference_id, version_id, lesson_id, source_number,
                    reference["title"], reference.get("description"), reference.get("url"),
                    reference["media_type"], reference.get("resource_code"),
                    reference.get("scope_kind"), reference.get("scope_value"), source_id,
                    reference["is_hidden"], Jsonb(reference["fields"]),
                    reference.get("activity_uuid"), reference.get("folder_uuid"),
                    reference.get("week_order"), reference.get("activity_order"),
                    reference.get("parent_activity_uuid"), reference.get("parent_inference"),
                ),
            )
            base_review = reference.get("_base_review") or {}
            complexity = base_review.get("complexity")
            validated = bool(
                base_review.get("validated") and reference.get("_content_unchanged")
            )
            if complexity is not None or validated:
                conn.execute(
                    "INSERT INTO syllabus_source_review"
                    " (reference_id, is_validated, complexity) VALUES (%s, %s, %s)",
                    (reference_id, validated, complexity),
                )

    diff = diff_versions(conn, base_version_id, version_id)
    event_id = next_curation_event_id(conn)
    conn.execute(
        "INSERT INTO curation_event (id, actor, action, subject, note)"
        " VALUES (%s, %s, 'syllabus_curated', %s, %s)",
        (
            event_id,
            actor,
            Jsonb(
                {
                    "syllabus_id": syllabus_id,
                    "base_version_id": base_version_id,
                    "version_id": version_id,
                    "diff": diff,
                }
            ),
            clean_note,
        ),
    )
    conn.commit()
    return {
        "syllabus_id": syllabus_id,
        "version_id": version_id,
        "seq": next_seq,
        "unchanged": False,
        "lesson_count": len(normalized),
        "reference_count": reference_count,
        "source_count": reference_count,
        "new_source_count": new_sources,
        "diff": diff,
    }


def diff_versions(conn: psycopg.Connection, version_a: str, version_b: str) -> dict:
    """Compare source references in two full syllabus versions."""
    query = (
        "SELECT sl.week, sl.title, sl.activity_uuid, sr.seq, sr.title, sr.url,"
        " sr.description, sr.resource_code, sr.scope_kind, sr.scope_value,"
        " sr.is_hidden, sr.activity_uuid,"
        " sr.fields->'adalove_material'->>'Source path'"
        " FROM syllabus_source_reference sr"
        " JOIN syllabus_lesson sl"
        "   ON sl.version_id = sr.version_id AND sl.id = sr.lesson_id"
        " WHERE sr.version_id = %s ORDER BY sl.week, sl.seq, sr.seq, sr.id"
    )
    rows_a = conn.execute(query, (version_a,)).fetchall()
    rows_b = conn.execute(query, (version_b,)).fetchall()
    def keys(row):
        if row[11]:
            return ("adalove", row[11], row[12] or row[3])
        return ("authored", row[0], row[1], row[4])

    map_a, map_b = {keys(row): row for row in rows_a}, {keys(row): row for row in rows_b}

    def record(row):
        return {
            "week": row[0],
            "lesson": row[1],
            "seq": row[3],
            "title": row[4],
            "url": row[5],
            "description": row[6],
            "resource_code": row[7],
            "scope_kind": row[8],
            "scope_value": row[9],
            "hidden": row[10],
        }

    added = [record(map_b[key]) for key in map_b.keys() - map_a.keys()]
    removed = [record(map_a[key]) for key in map_a.keys() - map_b.keys()]
    changed = []
    for key in map_a.keys() & map_b.keys():
        if record(map_a[key]) != record(map_b[key]):
            changed.append({"before": record(map_a[key]), "after": record(map_b[key])})
    sort_key = lambda item: (item["week"], item["lesson"], item["seq"], item["title"])
    added.sort(key=sort_key)
    removed.sort(key=sort_key)
    changed.sort(key=lambda item: sort_key(item["after"]))
    return {"added": added, "removed": removed, "changed": changed}


# --- CLI ---------------------------------------------------------------------------


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(prog="universe.syllabus")
    sub = parser.add_subparsers(dest="command", required=True)
    import_cmd = sub.add_parser("import", help="import a complete workbook version")
    import_cmd.add_argument("path")
    import_cmd.add_argument("--name", required=True, help="human name of the syllabus")
    import_cmd.add_argument("--syllabus-id", help="existing syllabus id for a new version")
    import_cmd.add_argument(
        "--institution-id",
        help="Companion Institution slug for a new Syllabus",
    )
    import_cmd.set_defaults(func=cmd_import)
    sub.add_parser("list", help="list syllabi").set_defaults(func=cmd_list)
    return parser


def cmd_import(args: argparse.Namespace) -> None:
    from universe import companion_seam

    namespace = companion_seam.graph_namespace()
    with connect() as conn:
        if args.institution_id:
            companion_seam.remember_institution(conn, namespace, args.institution_id)
        result = import_workbook(
            conn,
            args.path,
            args.name,
            syllabus_id=args.syllabus_id,
            institution_id=args.institution_id,
            occupied_graph_ids=namespace["graph_ids"],
            require_syllabus_metadata=True,
        )
    print(json.dumps(result, default=str, ensure_ascii=False, indent=2))


def cmd_list(_args: argparse.Namespace) -> None:
    with connect() as conn:
        result = list_syllabi(conn)
    print(json.dumps(result, default=str, ensure_ascii=False, indent=2))


def main(argv: list[str] | None = None) -> None:
    args = build_parser().parse_args(argv)
    args.func(args)


if __name__ == "__main__":
    main()
